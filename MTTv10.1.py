import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math

# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
"""print('학생들의 이름이 담긴 길이 s짜리 xlsx파일의 이름이 뭔가요?')"""
StudentNamesFile_path = 'StudentNames.xlsx'
dfStudentNames = pd.read_excel(StudentNamesFile_path)
StudentNames = dfStudentNames.values[:, 1]
s = len(StudentNames)

"""print('과목들의 이름이 담긴 길이 l짜리 xlsx파일의 이름이 뭔가요?')"""
LectureNamesFile_path = 'LectureNames.xlsx'
dfLectureNames = pd.read_excel(LectureNamesFile_path)
LN = dfLectureNames.values[0]
l = len(LN)

"""print('학생들의 수강 유무가 있는 s x l짜리 xlsx파일의 이름이 뭔가요?')"""
StudentLectureFile_path = 'SL.xlsx'
dfStudentLecture = pd.read_excel(StudentLectureFile_path)
SL = dfStudentLecture.values

"""print('각 과목별로 시수,종류가 있는 l x 2짜리 xlsx파일의 이름이 뭔가요?')"""
KLFile_path = 'KindLecture.xlsx'
dfKindLecture = pd.read_excel(KLFile_path)
KL = dfKindLecture.values[0]

"""print('각 과목별로 한 교시에 가능한 최대 분반의 개수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
MCPLFile_path = 'MCPL.xlsx'
dfMCPL = pd.read_excel(MCPLFile_path)
MLP = dfMCPL.values[0]

"""print('각 과목별 분반의 수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
LCNFile_path = 'LCN.xlsx'
dfLCN = pd.read_excel(LCNFile_path)
LCN = dfLCN.values[0]
c = np.sum(LCN)

"""print('1주일동안 수업이 진행되는 날들의 교시수들이 담긴 xlsx파일의 이름은 뭔가요?')"""
DPNFile_path = 'DailyPN.xlsx'
dfDailyPN = pd.read_excel(DPNFile_path)
DPN = dfDailyPN.values[0]

"""print('과목별로 진행되는 강의실의 기하학적 중심에서 다른 과목의 강의실로 가는데 이동해야 하는 실거리가 담긴 l x l짜리 xlsx파일의 이름이 뭔가요?')"""
DistanceFile_path = 'DI.xlsx'
dfDistance = pd.read_excel(DistanceFile_path)
DI = dfDistance.values

"""print('각 학생별로 모든 과목들을 자신에게 중요하다고 생각하는 것들의 순위(0부터시작하는 순위)가 담긴 s x l짜기 xlsx파일의 이름이 뭔가요?')"""
StudentimpFile_path = 'Studentimp.xlsx'
dfStudentimp = pd.read_excel(StudentimpFile_path)
Simp = dfStudentimp.values

"""print('매 요일의 이름이 담긴 xlsx파일의 이름은 뭔가요?')"""
datenamesFile_path = 'DateNames.xlsx'
dfdatenames = pd.read_excel(datenamesFile_path)
datenames = dfdatenames.values[0]


# 추가 함수
def SelT(k, arr1, n):
    if k == 4:
        return np.array([arr1[n]])
    elif k == 3:
        return np.array([arr1[n], arr1[n] + 1])
    elif k == 2:
        return np.array([arr1[n][0], arr1[n][1]])
    elif k == 1:
        return np.array([arr1[n][0], arr1[n][0] + 1, arr1[n][1]])
    else:
        return np.array([arr1[n][0], arr1[n][0] + 1, arr1[n][1], arr1[n][1] + 1])


def abletime(n):
    l1 = PIL(ICL, n)[0]
    return np.arange(wpn)[(np.sum(GP[np.flatnonzero(GC[:, n])], axis=0) == 0) * (LP[l1] < MLP[l1])]


def Even(arr):
    timeset = np.intersect1d(EP[CiP[cN, EP] * CiP[cN, VEP] == 1], abletime(cN))
    eap = np.array([], int)
    for i in range(weekly):
        if i not in CAD or arr[SDPI[np.where(CAD == i)[0]]] == 1:
            eap = np.append(eap, np.intersect1d(np.arange(IPD[i], arr[i] + 2, 2), timeset))
    return np.sort(eap)


def Odd(arr):
    timeset = abletime(cN)
    oap = np.array([], int)
    for i in range(weekly):
        n1 = arr[SDPI[np.where(CAD == i)[0]]]
        n2 = NEP[np.where(NEPI == i)[0]]
        if i in CAD:
            if n1 == 1:
                oap = np.append(oap, np.intersect1d(np.arange(IPD[i], arr[i] + 2, 2), timeset))
        else:
            oap = np.append(oap, np.intersect1d(np.arange(IPD[i], arr[i] + 2, 2), timeset))
        if i not in CAD or n1 == 1 and DPN[i] % 2 == 1:
            if len(n2) > 0 and n2 in timeset:
                oap = np.append(oap, n2)
    for j in NZBF(arr[-1]):
        if VEP[j] in timeset:
            oap = np.append(oap, VEP[j])
    return np.sort(oap)


def APTType(arr0):
    k = KC[cN]
    if k == 3:
        ap = Even(arr0)
    elif k == 4:
        ap = Odd(arr0)
    elif k == 2:
        ap = np.empty((0, 2), int)
        AP1 = Odd(arr0)
        if len(AP1) > 0:
            for p1 in AP1:
                AP2 = Odd(CTType(4, p1, arr0))
                AP2 = AP2[AP2 > FPD[date(p1)[0]]]
                if len(AP2) > 0:
                    for p2 in AP2:
                        ap = np.append(ap, [[p1, p2]], axis=0)
    elif k == 1:
        ap = np.empty((0, 2), int)
        AP1 = Even(arr0)
        if len(AP1) > 0:
            for p1 in AP1:
                AP2 = Odd(CTType(3, p1, arr0))
                AP2 = AP2[(AP2 > FPD[date(p1)[0]]) + (AP2 < IPD[date(p1)[0]])]
                if len(AP2) > 0:
                    for p2 in AP2:
                        ap = np.append(ap, [[p1, p2]], axis=0)
    else:
        ap = np.empty((0, 2), int)
        AP1 = Even(arr0)
        if len(AP1) > 0:
            for p1 in AP1:
                AP2 = Even(CTType(3, p1, arr0))
                AP2 = AP2[AP2 > FPD[date(p1)[0]]]
                if len(AP2) > 0:
                    for p2 in AP2:
                        ap = np.append(ap, [[p1, p2]], axis=0)
    if len(ap) == 0:
        return -1
    return ap


# 행렬-원소 선택 함수들#
def PIL(arr1, ip):
    arr2 = np.array(ip, int).reshape(-1)
    pil1 = np.array([], int)
    for i in arr2:
        pil1 = np.append(pil1, np.sum(arr1 <= i) - 1)
    return pil1


# DPN 활용#
daily = np.max(DPN)
weekly = len(DPN)
IPD = np.array([0], int)
FPD = np.array([], int)
for i3 in range(weekly - 1):
    IPD = np.append(IPD, IPD[-1] + DPN[i3])
    FPD = np.append(FPD, IPD[-1] - 1)
wpn = np.sum(DPN)
FPD = np.append(FPD, wpn)


def date(ip):
    return PIL(IPD, np.array(ip).reshape(-1))


def CTType(k, inputp, arr):
    selp = np.array(inputp).reshape(-1)
    arr1 = np.copy(arr)
    for i in range(weekly):
        if arr[i] < DMEP[i]:
            for j in selp:
                if j == EP[PIL(EP, arr[i])]:
                    arr1[i] += 2
    dates = np.unique(date(inputp))
    for i in dates:
        if i in DCD:
            ind = np.where(DCD == i)[0]
            arr1[weekly + ind] = 1
    if k == 1:
        if selp[1] in EP or selp[1] in VEP:
            arr1[-1] = BFN(np.union1d(NZBF(arr[-1]), PIL(EP, selp[1])))
    elif k == 2 or k == 4:
        for i in selp:
            if i in EP or i in VEP:
                arr1[-1] = BFN(np.union1d(NZBF(arr[-1]), PIL(EP, i)))
    if k == 0 or k == 2:
        if np.all(np.isin(dates, SDPD)):
            i0 = np.where(DCD == dates[0])[0] + weekly
            if arr[i0] == 0:
                if inputp[0] + np.sum(DPN[dates[0]:dates[1]]) == inputp[2 - k // 2]:
                    arr1[i0] = 0
    return arr1


# making classes#
ICL = np.array([0], int)
FCL = np.array([], int)
for i6 in range(l - 1):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL = np.append(ICL, a1 + a2)
    FCL = np.append(FCL, a1 + a2 - 1)
FCL = np.append(FCL, c - 1)

StS = np.argsort(s * np.count_nonzero(SL, axis=1) + np.arange(s))
SL = SL[StS]
SC = np.zeros((s, c), int)
cnow = 0
for l1 in range(l):
    SinL = np.nonzero(SL[:, l1])[0]
    SinLC = np.array_split(SinL, LCN[l1])
    for i in range(LCN[l1]):
        SinC = SinLC[i]
        for j in range(len(SinC)):
            SC[SinC[j]][cnow] = 1
        cnow += 1
KC = np.repeat(KL, LCN)

ICK = np.array([], int)
for i in range(5):
    ICK = np.append(ICK, np.nonzero(KC == i)[0][0])

RSC = np.array([np.copy(SC)], int)
REGC = np.empty((0, s, c), int)
REg = np.array([], int)
LMC = np.array([], int)
LLC = np.array([], int)
LB = np.array([], int)
RE = np.array([], int)
SCbr = np.empty((s, c), int)

for i in range(l):
    if '실' in LN[i] and LN[i] != '천문학실습':
        for j in range(LCN[i]):
            LB = np.append(LB, ICL[i] + j)


def Allin(arr1, arr2):
    for i in range(len(arr1)):
        if arr1[i] not in arr2:
            return 0
    return 1


def Pprediction(arr1, arr2, arr3):
    for i in range(len(arr1)):
        arr4 = arr1[i, :arr2[i], ]
        for j in range(arr2[i]):
            if Allin(arr4, arr3) == 1:
                return 1
    return 0


def GTest(arr3):
    timeset = abletime(cN)
    for i in arr3:
        if i not in timeset:
            arr1 = GC[:, cN]
            return np.argmax(arr1 > 0) + s * cN
    if cN in LB:
        Focus2 = np.flatnonzero(SC[:, cN])
        Target2 = SbD[Focus2, PIL(IPD, arr3[0])]
        if np.max(Target2) > 0:
            return (np.argmax(Target2) + s * cN)
    return -1


def sortition(arr):
    colsort = np.zeros(c, int)
    for l1 in range(l):
        ind = np.arange(LCN[l1]) + ICL[l1]
        np.put(colsort, ind, np.argsort(np.argmax(arr[:, ind], axis=0)) + ICL[l1])
    return arr[:, colsort]


def swr(arr, indi, n):
    csort = np.arange(len(arr.T))
    np.put(csort, np.arange(indi, indi + n), csort[indi:indi + n][::-1])
    return arr[:, csort]


def plusing(arr, xind, yind, num):
    for i in xind:
        for j in yind:
            arr[i, j] += num


def findEG(n):
    for j in np.flatnonzero(GC[:, n]):
        DCiP = np.zeros(wpn, int)
        np.putmask(DCiP, GC.T[n] @ GP - GP[j] == 0, 1)
        if np.min(np.sum(DCiP[EP] * DCiP[VEP] == 1)) >= 2 - kind % 2:
            return j
    return -1


def findEG2(n):
    for j in np.flatnonzero(GC[:, n]):
        DCiP = np.zeros(wpn, int)
        np.putmask(DCiP, GC.T[n] @ GP - GP[j] == 0, 1)
        if np.min(np.sum(DCiP == 1)) >= (12 - 2 * kind) // 3:
            return j
    return -1


# 소수,진법 변환 관련 함수들#
def NZBF(n):
    nzbf = np.array([], int)
    nbf = format(n, 'b')
    for i in range(len(nbf)):
        if nbf[i] == '1':
            nzbf = np.append(nzbf, i)
    return nzbf


def BFN(arr):
    n = 0
    for i in arr:
        n += 2 ** (len(EP) - i - 1)
    return n


# 교시교환 관련 함수들#
def LW1P(da, arr):
    a = IPD[da]
    tw = np.zeros((2, 2), int)
    for i in range(2):
        for j in range(2):
            for s1 in range(s):
                c1 = arr[s1][a + 1 - i]
                c2 = arr[s1][a + 2 + j]
                if max(c1, c2) < c:
                    l1 = PIL(ICL, c1)
                    l2 = PIL(ICL, c2)
                    tw[i][j] += DI[l1][l2]
    p2 = np.argmin(tw)
    y = p2 % 2
    x = p2 // 2
    mwp = np.array([a + x, a + 1 - x, a + 2 + y, a + 3 - y], int)
    if DPN[da] % 2 == 1:
        mwp = np.append(mwp, a + 4)
    return mwp


def LW2P(da, arr):
    a = IPD[da]
    tw = np.zeros((3, 2, 2), int)
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for s1 in range(s):
                    c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
                    c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
                    if max(c1, c2) < c:
                        l1 = PIL(ICL, c1)
                        l2 = PIL(ICL, c2)
                        tw[i][j] += DI[l1][l2]
    p3 = np.argmin(tw)
    z = p3 % 2
    y = (p3 // 2) % 2
    x = p3 // 4
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1], int)


def LW3P(da, arr):
    a = IPD[da]
    tw = np.zeros((3, 2, 2, 2), int)
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for o in range(2):
                    for s1 in range(s):
                        c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
                        c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
                        c3 = arr[s1][a + 2 * ((i + 2) % 3) + 1 - o]
                        c4 = arr[s1][a + 6]
                        if max(c1, c2) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
                        if max(c3, c4) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
    p4 = np.argmin(tw)
    w = p4 % 2
    z = (p4 // 2) % 2
    y = (p4 // 4) % 2
    x = p4 // 8
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6], int)


def LWP(arr):
    tmwp = np.array([], int)
    for i in range(weekly):
        if DPN[i] < 6:
            tmwp = np.append(tmwp, LW1P(i, arr))
        elif DPN[i] == 6:
            tmwp = np.append(tmwp, LW2P(i, arr))
        else:
            tmwp = np.append(tmwp, LW3P(i, arr))
    return tmwp


# 그래프 색깔#
"""
SPC = np.zeros(12)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
Math = []
Phy = []
Che = []
CoSc = []
LiSc = []
EaSc = []
ConS = []
ScEx = []
CCSL = []
Lan = []
Soc = []
MAS = []


def CSP(n):
    if n == 0:
        return Math
    elif n == 1:
        return Phy
    elif n == 2:
        return Che
    elif n == 3:
        return CoSc
    elif n == 4:
        return LiSc
    elif n == 5:
        return EaSc
    elif n == 6:
        return ConS
    elif n == 7:
        return ScEx
    elif n == 8:
        return CCSL
    elif n == 9:
        return Lan
    elif n == 10:
        return Soc
    else:
        return MAS


for c1 in range(c):
    partn = KC[c1][1]
    CSP(partn).append(c1)
    SPC[partn] += 1


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = []
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1.append(i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
        else:
            for G in range(k - R, -1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
    return RGB
"""
PTTcolorset = ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff',
               'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']

# TType 설정 코드#
DP = np.setdiff1d(np.arange(wpn), np.union1d(IPD, IPD + 4))
DEPN = DPN // 2
EP = np.array([], int)
for i in range(weekly):
    EP = np.append(EP, IPD[i])
    for j in range(DEPN[i] - 1):
        EP = np.append(EP, EP[-1] + 2)
VEP = EP + 1
NEP = np.setdiff1d(np.arange(wpn), np.union1d(EP, VEP))
NEPI = date(NEP)

DMEP = IPD + 2 * DEPN - 2
SUDPI = np.array([], int)
SDPI = np.array([], int)
SUDPD = np.array([], int)
SDPD = np.empty((0, 2), int)
DCD = np.array([], int)
CAD = np.array([], int)

fTType = np.copy(IPD)
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            if DPN[i] % 2 == 1:
                SUDPI = np.append(SUDPI, len(fTType))
                SUDPD = np.append(SUDPD, j)
            SDPI = np.append(SDPI, len(fTType))
            fTType = np.append(fTType, 0)
            SDPD = np.append(SDPD, [[i, j]], axis=0)
            DCD = np.append(DCD, i)
            CAD = np.append(CAD, j)
            fi = 0
fTType = np.append(fTType, 0)

ADS = np.arange(weekly).reshape(1, -1)
nds = 1
done1 = 1
i1111 = 0
l1111 = 0
while done1 == 1:
    j = 0
    re = 1
    while re * (nds - i1111) > 0:
        ADS1 = ADS[i1111]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = np.copy(ADS1)
            a1 = ADS1[k1]
            a2 = ADS1[k0]
            ADS2[k0] = a1
            ADS2[k1] = a2
            i1 = 0
            l1111 = len(ADS)
            while i1 < l1111:
                ads2 = ADS[i1]
                if np.array_equiv(ads2, ADS2):
                    j += 1
                    i1 = len(ADS)
                else:
                    i1 += 1
                    if i1 == len(ADS):
                        ADS = np.append(ADS, [ADS2], axis=0)
                        nds += 1
                        re = 0
        if re == 1:
            i1111 += 1
        elif i1111 < l1111:
            i1111 += 1
            j = 0
    if re == 1:
        done1 = 0


# 엑셀 조작 관련#
def style(workbook, n, i, j, arr0, arr1):
    newsheet = workbook[sheetname(n, arr1)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif arr0[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(arr0.reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == arr0[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[0], fill_type='solid')


def sheetname(b, arr):
    return str(arr[np.where(arr == b)[0]]) + '번째 학생'


def makingsheetdone(arr, wbname, arr1):
    wb = openpyxl.Workbook()
    for i in range(len(arr)):
        wb.create_sheet(sheetname(i, arr1))
    list3 = []
    for s3 in range(len(arr)):
        list0 = arr[s3].tolist()
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
    list4 = list3[:]
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(arr)):
            for i in range(weekly):
                for j in range(daily):
                    c5 = list3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, c5)
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LN[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0, arr1))
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(arr)):
        i = 0
        s2 = wb[sheetname(s1, arr1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list4[s1], arr1)
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


datepernum = ' ' * 5
for i in range(wpn):
    datepernum = datepernum + str(i - IPD[PIL(IPD, i)[0]] + 1) + ' ' * (len(str(i)) - 1) + ' '
datepernum = datepernum + ' '

strnan = ''
for i in range(wpn):
    strnan = strnan + ' ' * len(str(i)) + ' '


def printmtt(n1, n2):
    printstr = ''
    for i in range(cN):
        if i == cN - 1:
            printstr = printstr + str(i) + ' ' * (3 - len(str(i))) + ':V' + CP[i] + ' '
        else:
            printstr = printstr + str(i) + ' ' * (3 - len(str(i))) + ': ' + CP[i] + ' '
    for i in range(cN, maxcN):
        printstr = printstr + str(i) + ' ' * (3 - len(str(i))) + ': ' + strnan + ' '
    printstr = printstr + datepernum
    printone = printstr + 'Found:' + str(n1) + '    checked SC: ' + str(n2) + '    back: ' + str(back) + '    RE: ' + str(len(RE))
    print(printone, end='\r')


dead = 0
out = 0
distinguished = 0
old = 0
go = 0
back = 0
newface = 0
fastsend = 0

cN = -1
Found = 0
pN = 0
SbD = np.zeros((s, weekly), int)
RSbD = np.empty((0, s, weekly), int)
CP = np.array([], str)
CiP = np.empty((0, wpn), int)
maxcN = -1

while dead == 0:
    GC = np.empty((0, c), int)
    GS = np.array([], int)
    CC = np.zeros((c, c), int)
    for s1 in range(s):
        SS1 = SC[s1]
        CC += np.outer(SS1, SS1, out=None)
    for s2 in range(s):
        NZ = np.flatnonzero(SC[s2])
        if np.min(CC[NZ][:, NZ]) > 1:
            SS2 = SC[s2]
            CC -= np.outer(SS2, SS2, out=None)
        else:
            GS = np.append(GS, s2)
            GC = np.append(GC, [SC[s2]], axis=0)
    g = len(GS)
    GC = sortition(GC)
    out = Pprediction(REGC, REg, GC)
    if out == 1:
        fastsend = 1
    else:
        distinguished = 0
        old = 0
        RE = np.array([], int)
        GP = np.zeros((g, wpn), int)
        RGP = np.empty((0, g, wpn), int)
        SbD = np.zeros((s, weekly), int)
        RSbD = np.empty((0, s, weekly), int)
        RpN = np.array([], int)
        RpL = np.array([], int)
        LP = np.zeros((l, wpn), int)
        RLP = np.empty((0, l, wpn), int)
        TType = np.copy(fTType)
        RTType = np.empty((0, len(TType)), int)
        CP = np.array([], str)
        maxcN = -1
        CiP = np.empty((0, wpn), int)
        while out == 0:
            if old == 0:
                cN += 1
                maxcN = max(maxcN, cN)
                pN = 0
            lN = PIL(ICL, cN)[0]
            kind = KL[lN]
            go = 0
            if cN in ICL:
                if old == 0:
                    CiP = np.append(CiP, np.zeros((LCN[lN], wpn), int), axis=0)
                if lN == l - 1:
                    np.putmask(CiP[cN:], GC.T[cN:] @ GP == 0, 1)
                else:
                    np.putmask(CiP[cN:], GC.T[cN:cN + LCN[lN]] @ GP == 0, 1)
                EC1 = np.array([], int)
                EC2 = np.array([], int)
                if kind == 0 or kind == 1 or kind == 3:
                    sums = np.sum(CiP[cN:, EP] * CiP[cN:, VEP] == 1, axis=1)
                    if np.min(sums) < 2 - kind % 2:
                        EC1 = ICL[lN] + np.argwhere(sums == np.min(sums)).reshape(-1)
                sums2 = np.sum(CiP[cN:] == 1, axis=1)
                if np.min(sums2) < (12 - 2 * kind) // 3:
                    EC2 = (ICL[lN] + np.argwhere(sums2 == np.min(sums2)).reshape(-1))
                EC = np.unique(np.append(EC1, EC2))
                if len(EC)>0:
                    go = 1
                    back = 1
                for i in EC:
                    if i in EC1:
                        if findEG(i) > -1:
                            RE = np.unique(np.append(RE, i * s + findEG(i)))
                    else:
                        if findEG2(i) > -1:
                            RE = np.unique(np.append(RE, i * s + findEG2(i)))
            AP = APTType(TType)
            if type(AP) is int:
                RpL = np.append(RpL, 0)
            else:
                RpL = np.append(RpL, len(AP))
            if cN == c:
                ScT = np.ones((s, wpn), int) * c
                SiT = np.ones((s, wpn), int) * l
                for c1 in range(c):
                    TType = RTType[c1]
                    ki = KC[c1]
                    Selt2 = SelT(ki, APTType(TType), RpN[c1])
                    l1 = PIL(ICL, c1)
                    carray = np.flatnonzero(SC[:, c1])
                    plusing(ScT, carray, Selt2, c1)
                    for i in carray:
                        for j in Selt2:
                            SiT[i][j] += Simp[i][l1]
                ScT = ScT[:, LWP(ScT)]
                SiT = SiT[:, LWP(ScT)]
                for d2 in range(weekly):
                    if DPN[d2] > 3:
                        p4 = IPD[d2]
                        if np.sum(SiT[p4:p4 + 2]) < np.sum(SiT[p4 + 2:p4 + 4]):
                            ScT = swr(ScT, p4, 4)
                        if DPN[d2] == 6 and np.sum(SiT[p4 + 4]) < np.sum(SiT[p4 + 5]):
                            ScT = swr(ScT, p4 + 4, 2)
                qdsds = np.array([], int)
                for i in range(nds):
                    basic = np.arange(wpn)
                    basic1 = np.arange(wpn)
                    for j in range(weekly):
                        np.put(basic, np.arange(DPN[j]) + IPD[j], basic1[IPD[ADS[j]]:IPD[ADS[j]] + DPN[j]])
                    ScT1 = ScT[:, basic]
                    qdsd = 0
                    for cn in range(ICK[3]):
                        b = np.transpose(np.nonzero(ScT1 == cn))[:, 1]
                        kind = PIL(ICK, cn)
                        for i in range(s):
                            qdsd += PIL(IPD, b[(4 - kind) * i + (3 - kind)]) - PIL(IPD, b[(4 - kind) * i])
                    qdsds = np.append(qdsds, qdsd)
                BSP = ADS[np.argmin(qdsds)]
                basic2 = np.arange(wpn)
                basic3 = np.arange(wpn)
                for j in range(weekly):
                    np.put(basic2, np.arange(DPN[j]) + IPD[j], basic3[IPD[BSP[j]]:IPD[BSP[j]] + DPN[j]])
                ScT = ScT[:, basic2]
                tdw = 0
                for s1 in range(s):
                    sctd = ScT[s1]
                    for t1 in DP:
                        l1 = PIL(ICL, sctd[t1 - 1])
                        l2 = PIL(ICL, sctd[t1])
                        tdw += DI[l1][l2]
                workbookname = str(tdw) + '가 총 이동거리인 시간표'
                makingsheetdone(ScT, workbookname, StS)
            while go == 0:
                if RpL[-1] == 0 or pN == RpL[-1]:
                    back = 1
                    go = 1
                else:
                    SelP = SelT(kind, AP, pN)
                    gtest = GTest(SelP)
                    if gtest == -1:
                        go = 1
                        RpN = np.append(RpN, pN)
                        printmtt(Found, len(RSC))
                        RLP = np.append(RLP, [LP], axis=0)
                        plusing(LP, [lN], SelP, 1)
                        str1 = ''
                        for i in range(wpn):
                            if i not in SelP:
                                str1 = str1 + ' ' * len(str(i)) + ' '
                            else:
                                str1 = str1 + str(i) + ' '
                        CP = np.append(CP, str1)
                        RGP = np.append(RGP, [GP], axis=0)
                        plusing(GP, np.flatnonzero(GC[:, cN]), SelP, 1)
                        RTType = np.append(RTType, [TType], axis=0)
                        TType = CTType(kind, SelP, TType)
                        if cN in LB:
                            RSbD = np.append(RSbD, [SbD], axis=0)
                            plusing(SbD, np.flatnonzero(SC[:, cN]), PIL(IPD, SelP[0]), 1)
                    else:
                        RE = np.unique(np.append(RE, gtest))
                        pN += 1
            old = 0
            while back == 1:
                back = 0
                cN -= 1
                if cN == -1:
                    out = 1
                else:
                    RpL = np.delete(RpL, -1)
                    pN = RpN[-1] + 1
                    RpN = np.delete(RpN, -1)
                    if pN == RpL[-1]:
                        back = 1
                    else:
                        old = 1
                        GP = np.copy(RGP[cN])
                        RGP = RGP[:cN]
                        TType = np.copy(RTType[cN])
                        RTType = RTType[:cN]
                        LP = np.copy(RLP[cN])
                        RLP = RLP[:cN]
                        RpL = np.delete(RpL, -1)
                        printmtt(Found, len(RSC))
                        CP = CP[:cN]
                        if lN > PIL(ICL, cN)[0]:
                            CiP = CiP[:ICL[PIL(ICL, cN)[0] + 1]]
                        if cN in LB:
                            SbD = np.copy(RSbD[PIL(LB, cN)])
                            RSbD = RSbD[:PIL(LB, cN)]
    if fastsend == 0:
        delin = np.array([])
        for i in range(len(REGC)):
            if Allin(GC, REGC[i, :REg[i], :]) == 1:
                delin = np.append(delin, i)
        REGC = np.delete(REGC, delin, axis=0)
        REg = np.delete(REg, delin)
        REGC = np.append(REGC, [np.append(GC, np.zeros((s - g, c)), axis=0)], axis=0)
        REg = np.append(REg, g)
    fastsend = 0
    if distinguished == 0:
        SCbr = np.copy(SC)
        delind = np.array([])
        for i in range(len(RE)):
            classnumber = RE[i] // s
            lecture = PIL(ICL, classnumber)
            if LCN[lecture] == 1:
                LLC = np.append(LLC, lecture)
                delind = np.append(delind, i)
            elif np.sum(SC[:, classnumber]) == 1:
                LMC = np.append(LMC, lecture)
                delind = np.append(delind, i)
        RE = np.delete(RE, delind)
        RE = np.sort(RE)
        distinguished = 1
    newface = 0
    while newface == 0:
        classnum = RE[-1] // s
        lecture = PIL(ICL, classnum)
        student = GS[RE[-1] % s]
        SC = np.copy(SCbr)
        SC[student, classnum] = 0
        SC[student, classnum + 1 - LCN[lecture] * (classnum == FCL[lecture])] = 1
        RE = np.delete(RE, -1)
        SC = sortition(SC)
        if SC in RSC:
            if len(RE) == 0:
                dead = 1
                newface = 1
        else:
            RSC = np.append(RSC, [SC], axis=0)
            out = 0
            newface = 1
print("Lectures with less classes: " + str(LLC))
print("Lectures with much classes: " + str(LMC))
