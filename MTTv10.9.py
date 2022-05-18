import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook, chart
from openpyxl.utils import get_column_letter
import math
import matplotlib.pyplot as plt
from collections import Counter

# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
"""print('학생들의 이름이 담긴 길이 s짜리 xlsx파일의 이름이 뭔가요?')"""
StudentNames = pd.read_excel('StudentNames.xlsx').values[:, 1]
s = len(StudentNames)

"""print('과목들의 이름이 담긴 길이 l짜리 xlsx파일의 이름이 뭔가요?')"""
LN = pd.read_excel('LectureNames.xlsx').values[0]
l = len(LN)

"""print('학생들의 수강 유무가 있는 s x l짜리 xlsx파일의 이름이 뭔가요?')"""
SL = pd.read_excel('SL.xlsx').values

"""print('각 과목별로 시수가 있는 2 x l짜리 xlsx파일의 이름이 뭔가요?')"""
KL = pd.read_excel('KindLecture.xlsx').values[0]

"""print('각 과목별로 종류가 있는 2 x l짜리 xlsx파일의 이름이 뭔가요?')"""
KA = pd.read_excel('AreaLecture.xlsx').values[0]

"""print('각 과목별로 한 교시에 가능한 최대 분반의 개수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
MLP = pd.read_excel('MCPL.xlsx').values[0]

"""print('각 과목별 분반의 수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
LCN = pd.read_excel('LCN.xlsx').values[0]
c = np.sum(LCN)

"""print('1주일동안 수업이 진행되는 날들의 교시수들이 담긴 xlsx파일의 이름은 뭔가요?')"""
DPN = pd.read_excel('DailyPN.xlsx').values[0]

"""print('과목별로 진행되는 강의실의 기하학적 중심에서 다른 과목의 강의실로 가는데 이동해야 하는 실거리가 담긴 l x l짜리 xlsx파일의 이름이 뭔가요?')"""
DI = pd.read_excel('DI.xlsx').values

"""print('각 학생별로 모든 과목들을 자신에게 중요하다고 생각하는 것들의 순위(0부터시작하는 순위)가 담긴 s x l짜기 xlsx파일의 이름이 뭔가요?')"""
Simp = pd.read_excel('Studentimp.xlsx').values

"""print('매 요일의 이름이 담긴 xlsx파일의 이름은 뭔가요?')"""
datenames = pd.read_excel('DateNames.xlsx').values[0]
periodnames = np.array([], str)
for i in range(len(datenames)):
    for j in range(DPN[i]):
        periodnames = np.append(periodnames, datenames[i] + ' ' + str(j + 1) + '교시')


# 추가 함수
def SelT(kc1, aptype1, pn1):
    if kc1 == 4:
        return np.array([aptype1[pn1]])
    elif kc1 == 3:
        return np.array([aptype1[pn1], aptype1[pn1] + 1])
    elif kc1 == 2:
        return np.array([aptype1[pn1][0], aptype1[pn1][1]])
    elif kc1 == 1:
        return np.array([aptype1[pn1][0], aptype1[pn1][0] + 1, aptype1[pn1][1]])
    else:
        return np.array([aptype1[pn1][0], aptype1[pn1][0] + 1, aptype1[pn1][1], aptype1[pn1][1] + 1])


def CAiP(AiP1, cn7, kc4, ttype4, selp1):  # TType을 변경하기 전에 진행을 해야 한다.
    AiP2 = np.copy(AiP1)
    ln7 = PIL(ICL, cn7)[0]
    ttype5 = CTType(kc4, selp1, ttype4)
    remainc = np.flatnonzero(NcN)
    Dd = np.flatnonzero((ttype5 - ttype4)[weekly:-1]) + weekly
    for k in Dd:
        j = CAD[SDPI == k][0]
        i = DCD[SDPI == k][0]
        while i in CAD and ttype4[SDPI[CAD == i]] == 0:
            i = DCD[CAD == i][0]
        AiP2[remainc, IPD[j]:FPD[j] + 1] = np.copy(AiP2[remainc, IPD[i]:FPD[i] + 1])
    Ck = ttype5[np.flatnonzero((ttype5 - ttype4)[:weekly])]
    for i in Ck:
        AiP2[remainc, i:i + 2] *= -1
    for d1 in np.unique(date(selp1)):
        AiP2[cn7, IPD[d1]:FPD[d1] + 1] = 0
    GinC = np.flatnonzero(GC[:, cn7])
    for i in np.intersect1d(np.flatnonzero(np.sum(GC[GinC], axis=0)), remainc):
        AiP2[i, selp1] = 0
        AiP2[i] *= (np.sum(GP[GinC], axis=0) == 0)
    for i in np.intersect1d(np.arange(ICL[ln7], FCL[ln7] + 1), remainc):
        AiP2[i] *= (LP[ln7] < MLP[ln7])
    return AiP2


def APTType(AiP3, cn9, ttype6, num=0):
    kc4 = PIL(ICK, cn9)[0]
    if num > 0:
        kc4 = 5 - num
    if kc4 == 4:
        ape = EP[AiP3[cn9, EP] == 1]
        apve = VEP[np.intersect1d(np.argwhere(AiP3[cn9, EP] == 1).reshape(-1), NZBF(ttype6[-1]))]
        apne = NEP[AiP3[cn9, NEP] == 1]
        return np.append(np.append(ape, apve), apne)
    elif kc4 == 3:
        return EP[AiP3[cn9, EP] == 1]
    elif kc4 == 2:
        ap2 = np.empty((0, 2), int)
        AP1 = APTType(AiP3, cn9, ttype6, num=1)
        for i in AP1:
            AiP4 = CAiP(AiP3, cn9, 4, ttype6, i)
            AP2 = APTType(AiP4, cn9, CTType(4, i, ttype6), num=1)
            AP2 = AP2[AP2 > i]
            for j in AP2:
                ap2 = np.append(ap2, [[i, j]], axis=0)
        return ap2
    elif kc4 == 1:
        ap3 = np.empty((0, 2), int)
        AP1 = APTType(AiP3, cn9, ttype6, num=2)
        for i in AP1:
            AiP4 = CAiP(AiP3, cn9, 3, ttype6, [i, i + 1])
            AP2 = APTType(AiP4, cn9, CTType(3, i, ttype6), num=1)
            for j in AP2:
                ap3 = np.append(ap3, [[i, j]], axis=0)
        return ap3
    else:
        ap4 = np.empty((0, 2), int)
        AP1 = APTType(AiP3, cn9, ttype6, num=2)
        for i in AP1:
            AiP4 = CAiP(AiP3, cn9, 3, ttype6, [i, i + 1])
            AP2 = APTType(AiP4, cn9, CTType(3, i, ttype6), num=2)
            AP2 = AP2[AP2 > i]
            for j in AP2:
                ap4 = np.append(ap4, [[i, j]], axis=0)
        return ap4


def LPTType(AiP3, cn9, ttype6, num=0):
    kc4 = PIL(ICK, cn9)[0]
    if num > 0:
        kc4 = 5 - num
    if kc4 == 4:
        lpe = np.argwhere(AiP3[cn9, EP] == 1).reshape(-1)
        return np.sum(AiP3[cn9, EP] == 1) + len(np.intersect1d(lpe, NZBF(ttype6[-1]))) + np.sum(AiP3[cn9, NEP] == 1)
    elif kc4 == 3:
        return np.sum(AiP3[cn9, EP] == 1)
    elif kc4 == 2:
        lp2 = 0
        AP1 = APTType(AiP3, cn9, ttype6, num=1)
        for i in AP1:
            AP2 = APTType(CAiP(AiP3, cn9, 4, ttype6, i), cn9, CTType(4, i, ttype6), num=1)
            AP2 = AP2[AP2 > i]
            lp2 += len(AP2)
        return lp2
    elif kc4 == 1:
        lp3 = 0
        AP1 = APTType(AiP3, cn9, ttype6, num=2)
        for i in AP1:
            lp3 += LPTType(CAiP(AiP3, cn9, 3, ttype6, [i, i + 1]), cn9, CTType(3, i, ttype6), num=1)
        return lp3
    else:
        lp4 = 0
        AP1 = APTType(AiP3, cn9, ttype6, num=2)
        for i in AP1:
            AP2 = APTType(CAiP(AiP3, cn9, 3, ttype6, [i, i + 1]), cn9, CTType(3, i, ttype6), num=2)
            AP2 = AP2[AP2 > i]
            lp4 += len(AP2)
        return lp4


# 행렬-원소 선택 함수들#
def PIL(arr1, ip):
    return np.array([Counter(arr1 <= i)[True] for i in np.atleast_1d(ip)]) - 1


# DPN 활용#
daily = np.max(DPN)
weekly = len(DPN)
IPD = np.array([0], int)
FPD = np.array([], int)
for i3 in range(weekly - 1):
    IPD = np.append(IPD, IPD[-1] + DPN[i3])
    FPD = np.append(FPD, IPD[-1] - 1)
wpn = np.sum(DPN)
FPD = np.append(FPD, wpn - 1)


def date(ip):
    return PIL(IPD, np.array(ip).reshape(-1))


def CTType(kc4, inputp, ttype4):
    selp1 = np.array(inputp).reshape(-1)
    arr1 = np.copy(ttype4)
    for i in range(weekly):
        if ttype4[i] < DMEP[i]:
            for j in selp1:
                if j == EP[PIL(EP, ttype4[i])]:
                    arr1[i] += 2
    dates = np.unique(date(inputp))
    for i in dates:
        if i in DCD:
            arr1[weekly + np.where(DCD == i)[0]] = 1
    if kc4 == 1:
        if selp1[1] in EP or selp1[1] in VEP:
            arr1[-1] = BFN(np.union1d(NZBF(ttype4[-1]), PIL(EP, selp1[1])))
    elif kc4 == 2 or kc4 == 4:
        for i in selp1:
            if i in EP or i in VEP:
                arr1[-1] = BFN(np.union1d(NZBF(ttype4[-1]), PIL(EP, i)))
    if kc4 == 0 or kc4 == 2:
        for i in SDPD:
            if np.array_equiv(dates, i):
                i0 = np.where(DCD == dates[0])[0] + weekly
                if ttype4[i0] == 0:
                    if inputp[0] + np.sum(DPN[dates[0]:dates[1]]) == inputp[2 - kc4 // 2]:
                        arr1[i0] = 0
    return arr1


# making classes#
ICL = np.array([0], int)
FCL = np.array([], int)
for i6 in range(l - 1):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL = np.append(ICL, a1 + a2)
    FCL = np.append(FCL, a1 + a2 - 1)
FCL = np.append(FCL, c - 1)

StS = np.argsort(s * np.count_nonzero(SL, axis=1) + np.arange(s))
SL = SL[StS]
SC = np.zeros((s, c), int)
cnow = 0
for l1 in range(l):
    SinL = np.nonzero(SL[:, l1])[0]
    SinLC = np.array_split(SinL, LCN[l1])
    for i in SinLC:
        for j in i:
            SC[j, cnow] += 1
        cnow += 1
KC = np.repeat(KL, LCN)

ICK = np.array([], int)
for i in range(5):
    ICK = np.append(ICK, np.sum(KC < i))

RSC = np.array([np.copy(SC)], int)
REGC = np.empty((0, s, c), int)
REg = np.array([], int)
LMC = np.array([], int)
LLC = np.array([], int)
RE = np.array([], int)
SCbr = np.empty((s, c), int)


def Allin(arr1, arr2):
    for i in arr1:
        if i not in arr2:
            return 0
    return 1


def Pprediction(arr1, arr2, arr3):
    for i in range(len(arr1)):
        arr4 = arr1[i, :arr2[i], ]
        for j in range(arr2[i]):
            if Allin(arr4, arr3) == 1:
                return 1
    return 0


def sortition(arr):
    colsort = np.zeros(c, int)
    for l1 in range(l):
        ind = np.arange(LCN[l1]) + ICL[l1]
        np.put(colsort, ind, np.argsort(np.argmax(arr[:, ind], axis=0)) + ICL[l1])
    return arr[:, colsort]


def swr(arr, indi, n):
    csort = np.arange(len(arr.T))
    np.put(csort, np.arange(indi, indi + n), csort[indi:indi + n][::-1])
    return arr[:, csort]


def plusing(arr, xind, yind, num):
    for i in xind:
        for j in yind:
            arr[i, j] += num


# 소수,진법 변환 관련 함수들#
def NZBF(n):
    nzbf = np.array([], int)
    nbf = format(n, 'b')
    for i in range(len(nbf)):
        if nbf[i] == '1':
            nzbf = np.append(nzbf, i)
    return nzbf


def BFN(arr):
    n = 0
    for i in arr:
        n += 2 ** (len(EP) - i - 1)
    return n


# 교시교환 관련 함수들#
def LW1P(da1, arr):
    a = IPD[da1]
    tw = np.zeros((2, 2), int)
    for i in range(2):
        for j in range(2):
            for s1 in range(s):
                c1 = arr[s1][a + 1 - i]
                c2 = arr[s1][a + 2 + j]
                if max(c1, c2) < c:
                    l1 = PIL(ICL, c1)
                    l2 = PIL(ICL, c2)
                    tw[i][j] += DI[l1][l2]
    p2 = np.argmin(tw)
    y = p2 % 2
    x = p2 // 2
    mwp = np.array([a + x, a + 1 - x, a + 2 + y, a + 3 - y], int)
    if DPN[da1] % 2 == 1:
        mwp = np.append(mwp, a + 4)
    return mwp


def LW2P(da2, arr):
    a = IPD[da2]
    tw = np.zeros((3, 2, 2), int)
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for s1 in range(s):
                    c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
                    c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
                    if max(c1, c2) < c:
                        l1 = PIL(ICL, c1)
                        l2 = PIL(ICL, c2)
                        tw[i][j] += DI[l1][l2]
    p3 = np.argmin(tw)
    z = p3 % 2
    y = (p3 // 2) % 2
    x = p3 // 4
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1], int)


def LW3P(da3, arr):
    a = IPD[da3]
    tw = np.zeros((3, 2, 2, 2), int)
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for o in range(2):
                    for s1 in range(s):
                        c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
                        c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
                        c3 = arr[s1][a + 2 * ((i + 2) % 3) + 1 - o]
                        c4 = arr[s1][a + 6]
                        if max(c1, c2) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
                        if max(c3, c4) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
    p4 = np.argmin(tw)
    w = p4 % 2
    z = (p4 // 2) % 2
    y = (p4 // 4) % 2
    x = p4 // 8
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6], int)


def LWP(arr):
    tmwp = np.array([], int)
    for i in range(weekly):
        if DPN[i] < 6:
            tmwp = np.append(tmwp, LW1P(i, arr))
        elif DPN[i] == 6:
            tmwp = np.append(tmwp, LW2P(i, arr))
        else:
            tmwp = np.append(tmwp, LW3P(i, arr))
    return tmwp


# 그래프 색깔#
KCA = np.zeros((c,), int)
for c1 in range(c):
    l1 = PIL(ICL, c1)[0]
    KCA[c1] = KA[l1]
SPC = np.zeros((12,), int)
for i in range(12):
    SPC[i] = np.sum(KCA == i)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
CAL = []
for i in range(12):
    CAL.append(np.argwhere(KCA == 0).tolist())


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = np.array([], int)
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1 = np.append(RGBset1, i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                B = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[B], 'x'))
        else:
            for G in range(k - R, -1, -1):
                B = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[B], 'x'))
    return RGB


PTTcolorset = ['ffe4e1', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff', 'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']

# TType 설정 코드#
DP = np.setdiff1d(np.arange(wpn), np.union1d(IPD, IPD + 4))
DEPN = DPN // 2
EP = np.array([], int)
for i in range(weekly):
    EP = np.append(EP, IPD[i])
    for j in range(DEPN[i] - 1):
        EP = np.append(EP, EP[-1] + 2)
VEP = EP + 1
NEP = np.setdiff1d(np.arange(wpn), np.union1d(EP, VEP))
NEPI = date(NEP)

DMEP = IPD + 2 * DEPN - 2
SDPI = np.array([], int)
SDPD = np.empty((0, 2), int)
DCD = np.array([], int)
CAD = np.array([], int)

fTType = np.copy(IPD)
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            SDPI = np.append(SDPI, len(fTType))
            fTType = np.append(fTType, 0)
            SDPD = np.append(SDPD, [[i, j]], axis=0)
            DCD = np.append(DCD, i)
            CAD = np.append(CAD, j)
            fi = 0
fTType = np.append(fTType, 0)

oAiP = fTType[np.setdiff1d(np.arange(weekly), CAD, assume_unique=True)]
fAiP = -np.ones((c, wpn), int)
fAiP[:, np.sort(np.append(np.append(oAiP, oAiP + 1), NEP))] *= -1

lAPcNset = np.array([], int)
for i in range(np.max(KL) + 1):
    lAPcNset = np.append(lAPcNset, ((1 + 2 * (i > 0) * (i < 4)) * (4 - i) + i * (i == 4)) * np.ones(np.sum(KC == i), int))

ADS = np.arange(weekly).reshape(1, -1)
nds = 1
done1 = 1
i1111 = 0
l1111 = 0
while done1 == 1:
    j = 0
    re = 1
    while re * (nds - i1111) > 0:
        ADS1 = ADS[i1111]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = np.copy(ADS1)
            a1 = ADS1[k1]
            a2 = ADS1[k0]
            ADS2[k0] = a1
            ADS2[k1] = a2
            i1 = 0
            l1111 = len(ADS)
            while i1 < l1111:
                ads2 = ADS[i1]
                if np.array_equiv(ads2, ADS2):
                    j += 1
                    i1 = len(ADS)
                else:
                    i1 += 1
                    if i1 == len(ADS):
                        ADS = np.append(ADS, [ADS2], axis=0)
                        nds += 1
                        re = 0
        if re == 1:
            i1111 += 1
        elif i1111 < l1111:
            i1111 += 1
            j = 0
    if re == 1:
        done1 = 0


# 엑셀 조작 관련#
def style(workbook, n, i, j, arr0, arr1):
    newsheet = workbook[sheetname(n, arr1)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif arr0[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(arr0.reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == arr0[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[c4], fill_type='solid')


def graphstyle(workbook, a, arr1, arr2):
    newsheet = workbook[GT[a]]
    for c1 in arr1:
        arr3 = np.flatnonzero(np.sum(arr2 == c1, axis=0))
        for p1 in arr3:
            newsheet.cell(c1 + 2, p1 + 2).fill = PatternFill(start_color=CColorSet(arr1)[c1], end_color=CColorSet(arr1)[c1], fill_type='solid')


def sheetname(b, arr):
    return str(arr[np.where(arr == b)[0]]) + '번째 학생'


def makingsheetdone(arr, wbname, arr1, arr2):
    wb = openpyxl.Workbook()
    for i in range(len(arr)):
        wb.create_sheet(sheetname(i, arr1))
    for area in range(12):
        wb.create_sheet(GT[area])
    wb.create_sheet('student walking distance')
    list3 = []
    for s3 in range(len(arr)):
        list0 = arr[s3].tolist()
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
    list4 = list3[:]
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(arr)):
            for i in range(weekly):
                for j in range(daily):
                    c5 = list3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, c5)
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LN[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0, arr1))
        for area0 in range(12):
            df = pd.DataFrame(index=list(map(str, CAL[area0])), columns=periodnames)
            df.to_excel(writer, sheet_name=GT[area0])
        dfc = pd.DataFrame(arr2, columns=StudentNames[arr1])
        dfc.to_excel(writer, sheet_name='student walking distance')
        chart = openpyxl.chart.LineChart()
        chart.title = 'student walking distance graph'
        chart.x_axis.title = 'student'
        chart.y_axis.title = 'walking distance'
        ws1 = wb['student walking distance']
        datas = openpyxl.chart.Reference(ws1, min_col=2, min_row=2, max_col=s + 1, max_row=2)
        chart.add_data(datas, from_rows=True, titles_from_data=False)
        cats = openpyxl.chart.Reference(ws1, min_col=2, min_row=1, max_col=s + 1, max_row=1)
        chart.set_categories(cats)
        ws1.add_chart(chart, "A1")
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(arr)):
        i = 0
        s2 = wb[sheetname(s1, arr1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list4[s1], arr1)
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    for a1 in range(12):
        graphstyle(wb, a1, CAL[a1], arr)
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


datepernum = ' ' * 8
for i in range(wpn):
    datepernum = datepernum + str(i - IPD[PIL(IPD, i)[0]] + 1) + ' ' * (len(str(i)) - 1) + ' '
datepernum = datepernum + ' '

strnan = ''
for i in range(wpn):
    strnan = strnan + ' ' * len(str(i)) + ' '


def printmtt(n1, n2):
    printstr = ''
    for i in range(len(RcN)):
        if i == len(RcN) - 1:
            printstr = printstr + str(i) + ' ' * (3 - len(str(i))) + str(RcN[i]) + ' ' * (3 - len(str(RcN[i]))) + ':V' + CP[i] + ' '
        else:
            printstr = printstr + str(i) + ' ' * (3 - len(str(i))) + str(RcN[i]) + ' ' * (3 - len(str(RcN[i]))) + ': ' + CP[i] + ' '
    for i in np.arange(len(RcN), maxcN):
        printstr = printstr + str(i) + ' ' * (3 - len(str(i))) + ' ' * 3 + ': ' + strnan + ' '
    printstr = printstr + datepernum
    printstr = printstr + 'Found:' + str(n1) + '       checked SC: ' + str(n2) + '         maxcN: ' + str(maxcN - 1) + '        RE: ' + str(len(RE))
    print(printstr)


dead = 0
out = 0
distinguished = 0
old = 0
go = 0
back = 0
newface = 0
fastsend = 0

cN = -1
Found = 0
pN = 0
SbD = np.zeros((s, weekly), int)
RSbD = np.empty((0, s, weekly), int)
CP = np.array([], str)
AiP = np.copy(fAiP)
RAiP = np.empty((0, c, wpn), int)
maxcN = -1
RcN = np.array([], int)
gRcN = np.array([], int)
fig = plt.figure()
NcN = np.ones(c, int)

while dead == 0:
    GC = np.empty((0, c), int)
    GS = np.array([], int)
    CC = np.zeros((c, c), int)
    for s1 in range(s):
        SS1 = SC[s1]
        CC += np.outer(SS1, SS1, out=None)
    for s2 in range(s):
        NZ = np.flatnonzero(SC[s2])
        if np.min(CC[NZ][:, NZ]) > 1:
            SS2 = SC[s2]
            CC -= np.outer(SS2, SS2, out=None)
        else:
            GS = np.append(GS, s2)
            GC = np.append(GC, [SC[s2]], axis=0)
    g = len(GS)
    GC = sortition(GC)
    out = Pprediction(REGC, REg, GC)
    if out == 1:
        fastsend = 1
    else:
        distinguished = 0
        old = 0
        RE = np.array([], int)
        GP = np.zeros((g, wpn), int)
        RGP = np.empty((0, g, wpn), int)
        SbD = np.zeros((s, weekly), int)
        RSbD = np.empty((0, s, weekly), int)
        RpN = np.array([], int)
        LP = np.zeros((l, wpn), int)
        RLP = np.empty((0, l, wpn), int)
        TType = np.copy(fTType)
        RTType = np.empty((0, len(TType)), int)
        CP = np.array([], str)
        maxcN = -1
        figinum = 0
        AiP = np.copy(fAiP)
        RAiP = np.empty((0, c, wpn), int)
        gRcN = np.array([], int)
        RcN = np.array([], int)
        NgN = np.ones(g, int)
        RNgN = np.empty((0, g), int)
        NcN = np.ones(c, int)
        RNcN = np.empty((0, c), int)
        lAPcN = np.copy(lAPcNset)
        RlAPcN = np.empty((0, c), int)
        while out == 0:
            if old == 0:
                if np.min(lAPcN[NcN == 1]) == 0:
                    for i in np.argwhere(lAPcN[NcN == 1] == 0):
                        RE = np.unique(np.append(RE, s * np.flatnonzero(GC[:, RcN[-1]] * GC[:, i]) + RcN[-1]))
                    back = 1
                    go = 1
                else:
                    AcN = np.intersect1d(np.argwhere(lAPcN == np.min(lAPcN[NcN == 1])).reshape(-1), np.flatnonzero(NcN))
                    cN = AcN[np.argmax(np.sum(GC[NgN == 1][:, AcN], axis=0) * g + np.sum(GC[:, AcN], axis=0))]
                    maxcN = max(maxcN, c + 1 - np.sum(NcN))
                    pN = 0
            if np.sum(NcN) == 0:
                ScT = np.ones((s, wpn), int) * c
                SiT = np.ones((s, wpn), int) * l
                for c1 in RcN:
                    order = np.argwhere(RcN == cN)
                    Selt2 = SelT(KC[c1], APTType(RAiP[order], c1, RTType[order]), RpN[order])
                    l1 = PIL(ICL, c1)[0]
                    carray = np.flatnonzero(SC[:, c1])
                    plusing(ScT, carray, Selt2, c1)
                    for i in carray:
                        for j in Selt2:
                            SiT[i][j] += Simp[i][l1]
                ScT = ScT[:, LWP(ScT)]
                SiT = SiT[:, LWP(ScT)]
                for d2 in range(weekly):
                    if DPN[d2] > 3:
                        p4 = IPD[d2]
                        if np.sum(SiT[p4:p4 + 2]) < np.sum(SiT[p4 + 2:p4 + 4]):
                            ScT = swr(ScT, p4, 4)
                        if DPN[d2] == 6 and np.sum(SiT[p4 + 4]) < np.sum(SiT[p4 + 5]):
                            ScT = swr(ScT, p4 + 4, 2)
                qdsds = np.array([], int)
                for i in range(nds):
                    basic = np.arange(wpn)
                    basic1 = np.arange(wpn)
                    for j in range(weekly):
                        np.put(basic, np.arange(DPN[j]) + IPD[j], basic1[IPD[ADS[j]]:IPD[ADS[j]] + DPN[j]])
                    ScT1 = ScT[:, basic]
                    qdsd = 0
                    for cn in range(ICK[3]):
                        b = np.transpose(np.nonzero(ScT1 == cn))[:, 1]
                        kind = PIL(ICK, cn)
                        for i in range(s):
                            qdsd += PIL(IPD, b[(4 - kind) * i + (3 - kind)]) - PIL(IPD, b[(4 - kind) * i])
                    qdsds = np.append(qdsds, qdsd)
                BSP = ADS[np.argmin(qdsds)]
                Timechange = np.arange(wpn)
                basetime = np.arange(wpn)
                for j in range(weekly):
                    np.put(Timechange, np.arange(DPN[j]) + IPD[j], basetime[IPD[BSP[j]]:IPD[BSP[j]] + DPN[j]])
                ScT = ScT[:, Timechange]
                tdw = 0
                sdw = np.array([])
                for s1 in range(s):
                    sctd = ScT[s1]
                    for t1 in DP:
                        l1 = PIL(ICL, sctd[t1 - 1])
                        l2 = PIL(ICL, sctd[t1])
                        tdw += DI[l1][l2]
                        sdw = np.append(sdw, tdw - sdw[-1])
                workbookname = str(tdw) + '가 총 이동거리인 시간표'
                makingsheetdone(ScT, workbookname, StS, sdw)
            if go == 0:
                lN = PIL(ICL, cN)[0]
                kind = KL[lN]
                AP = APTType(AiP, cN, TType)
            while go == 0:
                go = 1
                if pN == len(AP) or len(AP) == 0:
                    back = 1
                else:
                    SelP = SelT(kind, AP, pN)
            if back == 0:
                RNgN = np.append(RNgN, [NgN], axis=0)
                np.put(NgN, np.flatnonzero(GC[:, cN]), 0)
                RNcN = np.append(RNcN, [NcN], axis=0)
                np.put(NcN, cN, 0)
                RcN = np.append(RcN, cN)
                RpN = np.append(RpN, pN)
                RLP = np.append(RLP, [LP], axis=0)
                plusing(LP, [lN], SelP, 1)
                str1 = ''
                for i in range(wpn):
                    if i not in SelP:
                        str1 = str1 + ' ' * len(str(i)) + ' '
                    else:
                        str1 = str1 + str(i) + ' '
                CP = np.append(CP, str1)
                RGP = np.append(RGP, [GP], axis=0)
                plusing(GP, np.flatnonzero(GC[:, cN]), SelP, 1)
                RAiP = np.append(RAiP, [AiP], axis=0)
                AiP = CAiP(AiP, cN, kind, TType, SelP)
                RTType = np.append(RTType, [TType], axis=0)
                TType = CTType(kind, SelP, TType)
                htcc = np.flatnonzero(np.sum(GC[np.flatnonzero(GC[:, cN])], axis=0))
                for i in np.flatnonzero(NcN):
                    if PIL(ICL, i) == PIL(ICL, cN) or i in htcc:
                        lAPcN[i] = LPTType(AiP, i, TType)
                RlAPcN = np.append(RlAPcN, [lAPcN], axis=0)
                """"""
                printmtt(Found, len(RSC))
                """
                fig.clear()
                gRcN = np.append(gRcN, len(RcN))
                if len(gRcN) > 100:
                    gRcN = gRcN[-100:]
                plt.ylim([np.min(gRcN) - 1, maxcN + 1])
                x = np.flip(-np.arange(len(gRcN) - figinum) - 1)
                y1 = gRcN[figinum:]
                plt.title('RcN')
                plt.plot(x, y1, x, np.mean(y1) * np.ones(len(y1)), 'r-')
                plt.text(np.mean(x), np.mean(y1), np.mean(y1), fontsize=9, color='blue', horizontalalignment='center', verticalalignment='bottom')
                plt.pause(10 ** -100)
                plt.draw()
                """
            old = 0
            go = 0
            backspace = 0
            while back == 1:
                back = 0
                backspace -= 1
                if backspace == -len(RNcN):
                    out = 1
                else:
                    NcN = np.copy(RNcN[backspace])
                    cN = RcN[-1]
                    RcN = np.delete(RcN, -1)
                    lAPcN = np.copy(RlAPcN[-1])
                    RlAPcN = np.copy(RlAPcN[:-1])
                    pN = RpN[-1] + 1
                    RpN = np.delete(RpN, -1)
                    if pN == lAPcN[RcN[-1]]:
                        back = 1
                        for i in np.flatnonzero(GC[:, np.argmax(NcN)]):
                            RE = np.unique(np.append(RE, s * np.flatnonzero(GC[i]) + i))
                    else:
                        old = 1
                        NgN = np.copy(RNgN[backspace])
                        RNgN = RNgN[:backspace]
                        RNcN = RNcN[:backspace]
                        GP = np.copy(RGP[backspace])
                        RGP = RGP[:backspace]
                        TType = np.copy(RTType[backspace])
                        RTType = RTType[:backspace]
                        LP = np.copy(RLP[backspace])
                        RLP = RLP[:backspace]
                        CP = CP[:backspace]
                        AiP = np.copy(RAiP[backspace])
                        RAiP = RAiP[:backspace]
                        """"""
                        printmtt(Found, len(RSC))
                        """
                        fig.clear()
                        gRcN = np.append(gRcN, len(RcN))
                        if len(gRcN) > 100:
                            gRcN = gRcN[-100:]
                        plt.ylim([np.min(gRcN)-1, maxcN+1])
                        plt.title('numbers of cN')
                        plt.text(np.mean(x), np.mean(y1), np.mean(y1), fontsize=9, color='blue', horizontalalignment='center', verticalalignment='bottom')
                        x = np.flip(-np.arange(len(gRcN) - figinum) - 1)
                        y1= gRcN[figinum:]
                        plt.plot(x, y1, x, np.mean(y1)*np.ones(len(y1)), 'r-')
                        plt.pause(10**-100)
                        plt.draw()
                        """
    if fastsend == 0:
        delin = np.array([], int)
        for i in range(len(REGC)):
            if Allin(GC, REGC[i, :REg[i], :]) == 1:
                delin = np.append(delin, i)
        if len(delin) > 0:
            REGC = np.delete(REGC, delin, axis=0)
            REg = np.delete(REg, delin)
        REGC = np.append(REGC, [np.append(GC, np.zeros((s - g, c)), axis=0)], axis=0)
        REg = np.append(REg, g)
    fastsend = 0
    if distinguished == 0:
        SCbr = np.copy(SC)
        delind = np.array([], int)
        for i in range(len(RE)):
            classnumber = RE[i] // s
            lecture = PIL(ICL, classnumber)
            if LCN[lecture] == 1:
                LLC = np.append(LLC, lecture)
                delind = np.append(delind, i)
            elif np.sum(SC[:, classnumber]) == 1:
                LMC = np.append(LMC, lecture)
                delind = np.append(delind, i)
        if len(delind) > 0:
            RE = np.delete(RE, delind, axis=0)
        RE = np.sort(RE)
        distinguished = 1
    newface = 0
    while newface == 0:
        classnum = RE[-1] // s
        lecture = PIL(ICL, classnum)
        student = GS[RE[-1] % s]
        SC = np.copy(SCbr)
        SC[student, classnum] = 0
        SC[student, classnum + 1 - LCN[lecture] * (classnum == FCL[lecture])] = 1
        RE = np.delete(RE, -1)
        SC = sortition(SC)
        if SC in RSC:
            if len(RE) == 0:
                dead = 1
                newface = 1
        else:
            RSC = np.append(RSC, [SC], axis=0)
            out = 0
            newface = 1
print("Lectures with less classes: " + str(LLC))
print("Lectures with much classes: " + str(LMC))
