import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook, chart
from openpyxl.utils import get_column_letter
import math
import matplotlib.pyplot as plt
from collections import Counter
from tqdm import tqdm, trange
from numba import njit, int64

onegrade = np.array([0, 1, 7, 8, 11, 12, 13, 34, 57, 59, 61])
# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
"""print('학생들의 이름이 담긴 길이 s짜리 xlsx파일의 이름이 뭔가요?')"""
StudentNames = pd.read_excel('StudentNames.xlsx').values[:, 1][:8]
s = len(StudentNames)

"""print('과목들의 이름이 담긴 길이 l짜리 xlsx파일의 이름이 뭔가요?')"""
LN = pd.read_excel('LectureNames.xlsx').values[0][onegrade]
l = len(LN)

"""print('학생들의 수강 유무가 있는 s x l짜리 xlsx파일의 이름이 뭔가요?')"""
SL = pd.read_excel('SL.xlsx').values[:8][:, onegrade]

"""print('각 과목별로 시수가 있는 2 x l짜리 xlsx파일의 이름이 뭔가요?')"""
KL = pd.read_excel('KindLecture.xlsx').values[0][onegrade]

"""print('각 과목별로 종류가 있는 2 x l짜리 xlsx파일의 이름이 뭔가요?')"""
KA = pd.read_excel('AreaLecture.xlsx').values[0][onegrade]

"""print('각 과목별로 한 교시에 가능한 최대 분반의 개수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
MLP = pd.read_excel('MCPL.xlsx').values[0][onegrade]

"""print('각 과목별 분반의 수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
LCN = pd.read_excel('LCN.xlsx').values[0][onegrade]
c = np.sum(LCN)
C = np.arange(c)
CC = np.zeros((c,c))

"""print('1주일동안 수업이 진행되는 날들의 교시수들이 담긴 xlsx파일의 이름은 뭔가요?')"""
DPN = pd.read_excel('DailyPN.xlsx').values[0]

"""print('과목별로 진행되는 강의실의 기하학적 중심에서 다른 과목의 강의실로 가는데 이동해야 하는 실거리가 담긴 l x l짜리 xlsx파일의 이름이 뭔가요?')"""
DI = pd.read_excel('DI.xlsx').values[onegrade][:, onegrade]

"""print('각 학생별로 모든 과목들을 자신에게 중요하다고 생각하는 것들의 순위(0부터시작하는 순위)가 담긴 s x l짜기 xlsx파일의 이름이 뭔가요?')"""
Simp = pd.read_excel('Studentimp.xlsx').values[:8][:, onegrade]

"""print('매 요일의 이름이 담긴 xlsx파일의 이름은 뭔가요?')"""
datenames = pd.read_excel('DateNames.xlsx').values[0]
periodnames = np.array([], str)
for i in range(len(datenames)):
    for j in range(DPN[i]):
        periodnames = np.append(periodnames, datenames[i] + ' ' + str(j + 1) + '교시')

daily = np.max(DPN)
weekly = len(DPN)
IPD = np.array([0], int)
FPD = np.array([], int)
for i3 in range(weekly - 1):
    IPD = np.append(IPD, IPD[-1] + DPN[i3])
    FPD = np.append(FPD, IPD[-1] - 1)
wpn = np.sum(DPN)
P = np.arange(wpn)
FPD = np.append(FPD, wpn - 1)
StS = np.argsort(s * np.count_nonzero(SL, axis=1) + np.arange(s))
SL = SL[StS]
SC = np.zeros((s, c), int)
cnow = 0
for l1 in range(l):
    SinL = np.nonzero(SL[:, l1])[0]
    SinLC = np.array_split(SinL, LCN[l1])
    for i in SinLC:
        for j in i:
            SC[j, cnow] += 1
        cnow += 1
KC = np.repeat(KL, LCN)

ICK = np.array([], int)
for i in range(5):
    ICK = np.append(ICK, np.sum(KC < i))

RSC = np.array([np.copy(SC)], int)
REGC = np.empty((0, s, c), int)
REg = np.array([], int)
LMC = np.array([], int)
LLC = np.array([], int)
RE = np.array([], int)
SCbr = np.empty((s, c), int)

@njit(nogil=True, cache=True)
def PIL(arr1, ip):
    pil1 = np.empty(shape=(0), dtype=np.int64)
    for i in ip:
        pil1 = np.append(pil1, np.sum(arr1 <= i) - 1)
    return pil1

@njit(nogil=True, cache=True)
def date(ip):
    return np.unique(PIL(IPD, ip))

DP = np.setdiff1d(P, np.union1d(IPD, IPD + 4))
DEPN = DPN // 2
EP = np.array([], int)
for i in range(weekly):
    EP = np.append(EP, IPD[i])
    for j in range(DEPN[i] - 1):
        EP = np.append(EP, EP[-1] + 2)
VEP = EP + 1
AEP = np.setdiff1d(P, VEP)
NEP = np.setdiff1d(AEP, EP)
NEPI = date(NEP)
PE = np.arange(len(EP))

DMEP = IPD + 2 * DEPN - 2
SDPI = np.array([], int)
SDPD = np.empty((0, 2), int)
DCD = np.array([], int)
CAD = np.array([], int)

fTType = np.copy(IPD)
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            SDPI = np.append(SDPI, len(fTType))
            fTType = np.append(fTType, 0)
            SDPD = np.append(SDPD, [[i, j]], axis=0)
            DCD = np.append(DCD, i)
            CAD = np.append(CAD, j)
            fi = 0
fTType = np.append(fTType, 0)

oAiP = fTType[np.setdiff1d(np.arange(weekly), CAD, assume_unique=True)]
fAiP = -np.ones((c, wpn), int)
fAiP[:, np.sort(np.append(np.append(oAiP, oAiP + 1), NEP))] *= -1

ADS = np.arange(weekly).reshape(1, -1)
nds = 1
done1 = 1
i1111 = 0
l1111 = 0
while done1 == 1:
    j = 0
    re = 1
    while re * (nds - i1111) > 0:
        ADS1 = ADS[i1111]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = np.copy(ADS1)
            a1 = ADS1[k1]
            a2 = ADS1[k0]
            ADS2[k0] = a1
            ADS2[k1] = a2
            i1 = 0
            l1111 = len(ADS)
            while i1 < l1111:
                ads2 = ADS[i1]
                if np.array_equiv(ads2, ADS2):
                    j += 1
                    i1 = len(ADS)
                else:
                    i1 += 1
                    if i1 == len(ADS):
                        ADS = np.append(ADS, [ADS2], axis=0)
                        nds += 1
                        re = 0
        if re == 1:
            i1111 += 1
        elif i1111 < l1111:
            i1111 += 1
            j = 0
    if re == 1:
        done1 = 0


ICL = np.array([0], int)
FCL = np.array([], int)
for i6 in range(l - 1):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL = np.append(ICL, a1 + a2)
    FCL = np.append(FCL, a1 + a2 - 1)
FCL = np.append(FCL, c - 1)


@njit(nogil=True, cache=True)
def NZBF(n):
    nzbf = np.empty(shape=(0), dtype=np.int64)
    n1 = 0
    while n > 0:
        if n % 2 == 1:
            nzbf = np.append(nzbf, n1)
        n = n // 2
        n1 += 1
    return nzbf


@njit(nogil=True, cache=True)
def BFN(arr):
    return np.sum(2 ** (len(EP) - arr - 1))


@njit(nogil=True, cache=True)
def bitizettype(n):  # APTType
    bitize = np.zeros(shape=(len(EP)), dtype=np.int64)
    i = 1
    while n > 0:
        bitize[-i] = n % 2
        n = n // 2
        i += 1
    return bitize


@njit(nogil=True, cache=True)
def numizettype(arr):  # CTType
    return np.sum((2 ** (len(EP) - PE - 1)) * (arr))


@njit(nogil=True, cache=True)
def trizewpn(n):  # AiPt[i] ->AiP[i] and LPt[i]->LP[i]
    trize = np.zeros(shape=(wpn), dtype=np.int64)
    i = 1
    while n > 0:
        trize[-i] = n % 3
        n = n // 3
        i += 1
    return trize - 3 * (trize // 2)


@njit(nogil=True, cache=True)
def numizewpn(arr):  # AiP[i] -> AiPt[i] and LP[i]->LPt[i]
    return np.sum((3 ** (wpn - P - 1)) * (arr % 3))


@njit(nogil=True, cache=True)
def vectortrizewpn(arr1d):  # AiPt->AiP and LPt->LP
    vectortr = np.zeros(shape=(len(arr1d), wpn), dtype=np.int64)
    for i in range(len(arr1d)):
        vectortr[i] = trizewpn(arr1d[i])
    return vectortr


@njit(nogil=True, cache=True)
def vectornumizewpn(arr2d):  # AiP->AiPt and LP->LPt
    vectorwpn = np.zeros(shape=(len(arr2d)), dtype=np.int64)
    for i in range(len(arr2d)):
        vectorwpn[i] = numizewpn(arr2d[i])
    return vectorwpn


@njit(nogil=True, cache=True)
def union(arr1, arr2):
    return np.unique(np.append(arr1, arr2))


@njit(nogil=True, cache=True)
def setdiff(arr1, arr2):
    arr = np.empty(shape=(0), dtype=np.int64)
    for item in arr1:
        if item not in arr2:
            arr = np.append(arr, item)
    return arr


@njit(nogil=True, cache=True)
def in1d(arr1, arr2):
    arr = np.empty(shape=(0), dtype=np.int64)
    for i in np.intersect1d(arr1, arr2):
        arr = np.append(arr, np.argwhere(arr1 == i)[0])
    return arr

fAiPt = vectornumizewpn(fAiP)

# 변경 함수


# @profile
@njit(nogil=True, cache=True)
def CTType(selp1, ttype4):
    dates = date(selp1)
    kc4 = 6 - len(selp1) - len(dates)
    arr1 = np.copy(ttype4)
    for i in dates:
        if ttype4[i] < DMEP[i] and ttype4[i] in selp1:
            arr1[i] += 2
    arr1[weekly + in1d(DCD, dates)] = 1
    if kc4 == 1:
        if selp1[2] not in NEP:
            arr1[-1] = BFN(union(NZBF(ttype4[-1]), PIL(EP, np.array([selp1[2]]))))
    elif kc4 == 2 or kc4 == 4:
        arr1[-1] = BFN(union(NZBF(ttype4[-1]), PIL(EP, setdiff(selp1, NEP))))
    if kc4 == 0 or kc4 == 2:
        inarray = (np.sum(dates == SDPD, axis=1) == 2)
        if np.any(inarray):
            if ttype4[weekly + np.flatnonzero(inarray)] == 0 and selp1[0] + np.sum(DPN[dates[0]:dates[1]]) == selp1[2 - kc4 // 2]:
                arr1[weekly + np.flatnonzero(inarray)] = 0
    return arr1


# @profile
@njit(int64[:](int64[:], int64[:], int64[:], int64[:], int64[:, :], int64, int64[:]), nogil=True, cache=True)
def CAiPt(AiPt1, ttype4, ttype5, selp1, LP1, cn7, Ncn7):
    AiP2 = vectortrizewpn(AiPt1)
    ln7 = PIL(ICL, np.array([cn7]))[0]
    Dd = np.flatnonzero((ttype5 - ttype4)[weekly:-1]) + weekly
    for k in Dd:
        j = CAD[SDPI == k][0]
        i = DCD[SDPI == k][0]
        while i in CAD and ttype4[SDPI[CAD == i]] == 0:
            i = DCD[CAD == i][0]
        for num in Ncn7:
            AiP2[num][IPD[j]:FPD[j] + 1] = np.copy(AiP2[num][IPD[i]:FPD[i] + 1])
    Ck = ttype5[np.flatnonzero((ttype5 - ttype4)[:weekly])]
    AiP2[Ncn7][:, np.concatenate((Ck, Ck + 1))] *= -1
    for d1 in date(selp1):
        AiP2[cn7][IPD[d1]:FPD[d1] + 1] = 0
    if cn7 not in Ncn7:
        AiP2[cn7] = 0
        AiP2[Ncn7[CC[cn7][Ncn7] > 0]][:, selp1] = 0
        AiP2[Ncn7[PIL(ICL, Ncn7) == ln7]][:, selp1] *= (LP1[ln7][selp1] < MLP[ln7])
        AiP2[cn7][selp1] = 1
    return vectornumizewpn(AiP2)

AP1


# @profile
@njit(int64[:, :](int64[:], int64[:], int64[:, :], int64, int64), nogil=True, cache=True)
def APTType(AiPt3, ttype6, LP2, cn9, num):
    kc4 = KC[cn9]
    if num > 0:
        kc4 = 5 - num
    if kc4 == 4:
        AiP3 = vectortrizewpn(AiPt3)
        apve = VEP[(AiP3[cn9][VEP] == 1) * (bitizettype(ttype6[-1]) == 1)]
        apae = AEP[(AiP3[cn9][AEP] == 1)]
        return np.concatenate((apve, apae)).reshape(-1, 1)
    elif kc4 == 3:
        AiP3 = vectortrizewpn(AiPt3)
        ape = EP[AiP3[cn9][EP] == 1].reshape(-1, 1)
        return np.append(ape, ape + 1, axis=1)
    AP1 = np.empty(shape=(0, 1 + (kc4 < 2)), dtype=np.int64)
    AP2 = np.empty(shape=(0, 1 + (kc4 < 1)), dtype=np.int64)
    aps = np.empty(shape=(0, 1 + (kc4 < 1)), dtype=np.int64)
    AP1 = APTType(AiPt3, ttype6, LP2, cn9, 1 + (kc4 < 2))
    SAP = np.array([], int)
    for ap1 in AP1:
        ttype7 = CTType(ap1, ttype6)
        LP2[PIL(ICL, np.array([cn9]))[0]][ap1] += 1
        AP2 = APTType(CAiPt(AiPt3, ttype6, ttype7, ap1, LP2, cn9, np.array([cn9], int)), ttype7, LP2, cn9, 1 + (kc4 < 1))
        if kc4 == 1:
            aps = np.append(aps, AP2[PIL(IPD, AP2[:, 0]) != date(ap1)], axis=0)
            SAP = np.append(SAP, np.sum(PIL(IPD, AP2[:, 0]) != date(ap1)))
        else:
            aps = np.append(aps, AP2[PIL(IPD, AP2[:, 0]) > date(ap1)], axis=0)
            SAP = np.append(SAP, np.sum(PIL(IPD, AP2[:, 0]) > date(ap1)))
    return np.concatenate((np.repeat(AP1, SAP, axis=0), aps), axis=1)


# @profile
@njit(int64(int64[:], int64[:], int64[:, :], int64, int64), nogil=True, cache=True)
def LPTType(AiPt3, ttype6, LP3, cn9, num):
    kc4 = KC[cn9]
    if num > 0:
        kc4 = 5 - num
    if kc4 == 4:
        return np.sum(vectortrizewpn(AiPt3)[cn9][AEP] == 1) + np.sum((vectortrizewpn(AiPt3)[cn9][VEP] == 1) * (bitizettype(ttype6[-1]) == 1))
    elif kc4 == 3:
        return np.sum(vectortrizewpn(AiPt3)[cn9][EP] == 1)
    else:
        lp2 = 0
        AP1 = APTType(AiPt3, ttype6, LP3, cn9, 1 + (kc4 < 2))
        for ap1 in AP1:
            ttype7 = CTType(ap1, ttype6)
            AP2 = APTType(CAiPt(AiPt3, ttype6, ttype7, ap1, LP3, cn9, np.array([cn9])), ttype7, LP3, cn9, 1 + (kc4 < 1))
            if kc4 == 1:
                lp2 += np.sum(PIL(IPD, AP2[:, 0]) != date(ap1))
            else:
                lp2 += np.sum(PIL(IPD, AP2[:, 0]) > date(ap1))
        return lp2


# making classes#
def Allin(arr1, arr2):
    for i in arr1:
        if i not in arr2:
            return 0
    return 1


def Pprediction(arr1, arr2, arr3):
    for i in range(len(arr1)):
        arr4 = arr1[i, :arr2[i], ]
        for j in range(arr2[i]):
            if Allin(arr4, arr3) == 1:
                return 1
    return 0


def sortition(arr):
    colsort = np.zeros(c, int)
    for l1 in range(l):
        ind = np.arange(LCN[l1]) + ICL[l1]
        np.put(colsort, ind, np.argsort(np.argmax(arr[:, ind], axis=0)) + ICL[l1])
    return arr[:, colsort]


def swr(arr, indi, n):
    csort = np.arange(len(arr.T))
    np.put(csort, np.arange(indi, indi + n), csort[indi:indi + n][::-1])
    return arr[:, csort]


def plusing(arr, xind, yind, num):
    for i in xind:
        for j in yind:
            arr[i, j] += num


# 교시교환 관련 함수들#
def LW1P(da1, arr):
    a = IPD[da1]
    tw = np.zeros((2, 2), int)
    for k in range(4):
        i = k // 2
        j = k % 2
        for s1 in range(s):
            c1 = arr[s1][a + 1 - i]
            c2 = arr[s1][a + 2 + j]
            if max(c1, c2) < c:
                l1 = PIL(ICL, np.array([c1]))[0]
                l2 = PIL(ICL, np.array([c2]))[0]
                tw[i][j] += DI[l1][l2]
    p2 = np.argmin(tw)
    y = p2 % 2
    x = p2 // 2
    mwp = np.array([a + x, a + 1 - x, a + 2 + y, a + 3 - y], int)
    if DPN[da1] % 2 == 1:
        mwp = np.append(mwp, a + 4)
    return mwp


def LW2P(da2, arr):
    a = IPD[da2]
    tw = np.zeros((3, 2, 2), int)
    for t in range(12):
        i = t // 4
        j = (t // 2) % 2
        k = t % 2
        for s1 in range(s):
            c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
            c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
            if max(c1, c2) < c:
                l1 = PIL(ICL, np.array([c1]))[0]
                l2 = PIL(ICL, np.array([c2]))[0]
                tw[i][j] += DI[l1][l2]
    p3 = np.argmin(tw)
    z = p3 % 2
    y = (p3 // 2) % 2
    x = p3 // 4
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1], int)


def LW3P(da3, arr):
    a = IPD[da3]
    tw = np.zeros((3, 2, 2, 2), int)
    for t in range(24):
        i = t // 8
        j = (t // 4) % 2
        k = (t // 2) % 2
        o = t % 2
        for s1 in range(s):
            c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
            c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
            c3 = arr[s1][a + 2 * ((i + 2) % 3) + 1 - o]
            c4 = arr[s1][a + 6]
            if max(c1, c2) < c:
                l1 = PIL(ICL, np.array([c1]))[0]
                l2 = PIL(ICL, np.array([c2]))[0]
                tw[i][j] += DI[l1][l2]
            if max(c3, c4) < c:
                l1 = PIL(ICL, np.array([c3]))[0]
                l2 = PIL(ICL, np.array([c4]))[0]
                tw[i][j] += DI[l1][l2]
    p4 = np.argmin(tw)
    w = p4 % 2
    z = (p4 // 2) % 2
    y = (p4 // 4) % 2
    x = p4 // 8
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6], int)


def LWP(arr):
    tmwp = np.array([], int)
    for i in range(weekly):
        if DPN[i] < 6:
            tmwp = np.append(tmwp, LW1P(i, arr))
        elif DPN[i] == 6:
            tmwp = np.append(tmwp, LW2P(i, arr))
        else:
            tmwp = np.append(tmwp, LW3P(i, arr))
    return tmwp


# 그래프 색깔#
KCA = np.zeros((c,), int)
for c1 in range(c):
    l1 = PIL(ICL, np.array([c1]))[0]
    KCA[c1] = KA[l1]
SPC = np.zeros((12,), int)
for i in range(12):
    SPC[i] = np.sum(KCA == i)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
CAL = []
for i in range(12):
    CAL.append(np.argwhere(KCA == 0).tolist())


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = np.array([], int)
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1 = np.append(RGBset1, i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                B = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[B], 'x'))
        else:
            for G in range(k - R, -1, -1):
                B = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[B], 'x'))
    return RGB


PTTcolorset = ['ffe4e1', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff', 'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']


# 엑셀 조작 관련#
def style(workbook, n, i, j, arr0, arr1):
    newsheet = workbook[sheetname(n, arr1)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif arr0[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(arr0.reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == arr0[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[c4], fill_type='solid')


def graphstyle(workbook, a, arr1, arr2):
    newsheet = workbook[GT[a]]
    for c1 in arr1:
        arr3 = np.flatnonzero(np.sum(arr2 == c1, axis=0))
        for p1 in arr3:
            newsheet.cell(c1 + 2, p1 + 2).fill = PatternFill(start_color=CColorSet(arr1)[c1], end_color=CColorSet(arr1)[c1], fill_type='solid')


def sheetname(b, arr):
    return str(arr[np.where(arr == b)[0]]) + '번째 학생'


def makingsheetdone(arr, wbname, arr1, arr2):
    wb = openpyxl.Workbook()
    for i in range(len(arr)):
        wb.create_sheet(sheetname(i, arr1))
    for area in range(12):
        wb.create_sheet(GT[area])
    wb.create_sheet('student walking distance')
    list3 = []
    for s3 in range(len(arr)):
        list0 = arr[s3].tolist()
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
    list4 = list3[:]
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(arr)):
            for i in range(weekly):
                for j in range(daily):
                    c5 = list3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, np.array([c5]))[0]
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LN[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0, arr1))
        for area0 in range(12):
            df = pd.DataFrame(index=list(map(str, CAL[area0])), columns=periodnames)
            df.to_excel(writer, sheet_name=GT[area0])
        dfc = pd.DataFrame(arr2, columns=StudentNames[arr1])
        dfc.to_excel(writer, sheet_name='student walking distance')
        chart = openpyxl.chart.LineChart()
        chart.title = 'student walking distance graph'
        chart.x_axis.title = 'student'
        chart.y_axis.title = 'walking distance'
        ws1 = wb['student walking distance']
        datas = openpyxl.chart.Reference(ws1, min_col=2, min_row=2, max_col=s + 1, max_row=2)
        chart.add_data(datas, from_rows=True, titles_from_data=False)
        cats = openpyxl.chart.Reference(ws1, min_col=2, min_row=1, max_col=s + 1, max_row=1)
        chart.set_categories(cats)
        ws1.add_chart(chart, "A1")
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(arr)):
        i = 0
        s2 = wb[sheetname(s1, arr1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list4[s1], arr1)
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    for a1 in range(12):
        graphstyle(wb, a1, CAL[a1], arr)
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


dead = 0
out = 0
distinguished = 0
old = 0
newface = 0
fastsend = 0

# numba jit용
PIL(np.array([0, 2], int), np.array([0], int))
date(np.array([0], int))
BFN(np.array([0], int))
in1d(np.array([0], int), np.array([0], int))
NZBF(0)
CTType(np.array([0, 1], int), fTType)
union(np.array([0], int), np.array([0], int))
setdiff(np.array([0], int), np.array([0], int))
bitizettype(0)
numizettype(np.zeros(len(EP), int))
trizewpn(0)
numizewpn(np.zeros(wpn, int))
vectornumizewpn(np.zeros((l, wpn), int))
vectortrizewpn(np.zeros(l, int))
fAiPt = vectornumizewpn(fAiP)

while dead == 0:
    GC = np.empty((0, c), int)
    GS = np.array([], int)
    CC = np.zeros((c, c), int)
    for s1 in range(s):
        SS1 = SC[s1]
        CC += np.outer(SS1, SS1, out=None)
    for s2 in range(s):
        NZ = np.flatnonzero(SC[s2])
        if np.min(CC[NZ][:, NZ]) > 1:
            SS2 = SC[s2]
            CC -= np.outer(SS2, SS2, out=None)
        else:
            GS = np.append(GS, s2)
            GC = np.append(GC, [SC[s2]], axis=0)
    g = len(GS)
    GC = sortition(GC)
    out = Pprediction(REGC, REg, GC)
    if out == 1:
        fastsend = 1
    else:
        distinguished = 0
        old = 0
        RE = np.array([], int)
        SuNcN = np.arange(c).reshape(1, -1)
        SuRcN = np.empty((1, 0), int)
        SuLP = np.zeros((1, l, wpn), int)
        SuAiPt = np.copy(fAiPt).reshape(1, -1)
        SuTType = np.copy(fTType).reshape(1, -1)
        SucN = np.array([], int)
        fLlcN = np.array([], int)

        # numba jit 용
        PIL(np.array([0, 2], int), np.array([0], int))
        date(np.array([0], int))
        BFN(np.array([0], int))
        in1d(np.array([0], int), np.array([0], int))
        NZBF(0)
        CTType(np.array([0, 1], int), fTType)
        union(np.array([0], int), np.array([0], int))
        setdiff(np.array([0], int), np.array([0], int))
        bitizettype(0)
        numizettype(np.zeros(len(EP), int))
        trizewpn(0)
        numizewpn(np.zeros(wpn, int))
        vectornumizewpn(np.zeros((l, wpn), int))
        vectortrizewpn(np.zeros(l, int))
        APTType(fAiPt, fTType, np.zeros((l, wpn), int), 0, 0)
        LPTType(fAiPt, fTType, np.zeros((l, wpn), int), 0, 0)
        CAiPt(fAiPt, fTType, fTType, np.array([0], int), np.zeros((l, wpn), int), 0, C)

        for i in range(c):
            fLlcN = np.append(fLlcN, LPTType(SuAiPt[0], SuTType[0], SuLP[0], i, 0))
        SuLlcN = np.array([fLlcN]).reshape(1, -1)
        delpath = np.array([], int)
        while out == 0:
            if len(SuNcN.T) == 0:
                for indexs in range(len(SuAiPt)):
                    CP = vectortrizewpn(SuAiPt[indexs])
                    RcN = SuRcN[indexs]
                    ScT = np.ones((s, wpn), int) * c
                    SiT = np.ones((s, wpn), int) * l
                    for c1 in range(c):
                        c2 = RcN[c1]
                        order = np.argwhere(RcN == c2)
                        l1 = PIL(ICL, np.array([c2]))[0]
                        carray = np.flatnonzero(SC[:, c2])
                        plusing(ScT, carray, np.flatnonzero(CP[c2]), c2 - c)
                        for i in carray:
                            for j in np.flatnonzero(CP[c2]):
                                SiT[i][j] = Simp[i][l1]
                    ScT = ScT[:, LWP(ScT)]
                    SiT = SiT[:, LWP(ScT)]
                    for d2 in range(weekly):
                        if DPN[d2] > 3:
                            p4 = IPD[d2]
                            if np.sum(SiT[p4:p4 + 2]) < np.sum(SiT[p4 + 2:p4 + 4]):
                                ScT = swr(ScT, p4, 4)
                            if DPN[d2] == 6 and np.sum(SiT[p4 + 4]) < np.sum(SiT[p4 + 5]):
                                ScT = swr(ScT, p4 + 4, 2)
                    qdsds = np.array([], int)
                    for i in range(nds):
                        basic = np.copy(P)
                        basic1 = np.copy(P)
                        for j in range(weekly):
                            np.put(basic, np.arange(DPN[j]) + IPD[j], basic1[IPD[ADS[j]]:IPD[ADS[j]] + DPN[j]])
                        ScT1 = ScT[:, basic]
                        qdsd = 0
                        for cn in range(ICK[3]):
                            b = np.transpose(np.nonzero(ScT1 == cn))[:, 1]
                            kind = PIL(ICK, np.array([cn]))[0]
                            for i in range(s):
                                qdsd += PIL(IPD, np.array([b[(4 - kind) * i + (3 - kind)]]))[0] - PIL(IPD, np.array([b[(4 - kind) * i]]))[0]
                        qdsds = np.append(qdsds, qdsd)
                    BSP = ADS[np.argmin(qdsds)]
                    Timechange = np.copy(P)
                    basetime = np.copy(P)
                    for j in range(weekly):
                        np.put(Timechange, np.arange(DPN[j]) + IPD[j], basetime[IPD[BSP[j]]:IPD[BSP[j]] + DPN[j]])
                    ScT = ScT[:, Timechange]
                    tdw = 0
                    sdw = np.array([])
                    for sctd in ScT:
                        for t1 in DP:
                            if np.all(sctd[t1 - 1, t1] < c):
                                tdw += DI[PIL(ICL, np.array([sctd[t1 - 1, t1]]))[0]][PIL(ICL, np.array([sctd[t1 - 1, t1]]))[1]]
                                sdw = np.append(sdw, tdw - sdw[-1])
                    workbookname = str(tdw) + '가 총 이동거리인 시간표'
                    makingsheetdone(ScT, workbookname, StS, sdw)
            if len(SuLlcN) > 0:
                arg = np.argsort(SuLlcN, axis=1)
                for i in range(len(SuLlcN)):
                    SuNcN[i] = SuNcN[i, arg[i]]
                    SuLlcN[i] = SuLlcN[i, arg[i]]
                SucN = SuNcN[:, 0]
                SuRcN = np.append(SuRcN, SucN.reshape(-1, 1), axis=1)
                SuNcN = np.delete(SuNcN, 0, axis=1)
                Renum = np.copy(SuLlcN[:, 0])
                SuLlcN = np.delete(SuLlcN, 0, axis=1)
                APcN1d = np.array([], int)
                IAD = np.array([], int)
                for i in tqdm(range(len(SuTType)), desc=str(len(SuRcN.T)) + ' APcN1d', ascii=True, leave=False):
                    IAD = np.append(IAD, np.arange(Renum[i]) * (5 - KC[SucN[i]] - (KC[SucN[i]] < 3)) + len(APcN1d))
                    APcN1d = np.append(APcN1d, APTType(SuAiPt[i], SuTType[i], SuLP[i], SucN[i], 0).reshape(-1))
                IAD = np.append(IAD, len(APcN1d))
                SuAiPt = np.repeat(SuAiPt, Renum, axis=0)
                SuTType = np.repeat(SuTType, Renum, axis=0)
                SuTTypep = np.copy(SuTType)
                SuLP = np.repeat(SuLP, Renum, axis=0)
                SuLlcN = np.repeat(SuLlcN, Renum, axis=0)
                SuNcN = np.repeat(SuNcN, Renum, axis=0)
                SuRcN = np.repeat(SuRcN, Renum, axis=0)
                SucN = np.repeat(SucN, Renum, axis=0)
                SulN = PIL(ICL, SucN)
                delpath = np.array([], int)
                for i in tqdm(range(sum(Renum)), desc=str(len(SuRcN.T)), ascii=True):
                    AP = APcN1d[IAD[i]:IAD[i + 1]]
                    SuTType[i] = CTType(AP, SuTType[i])
                    SuAiPt[i] = CAiPt(SuAiPt[i], SuTTypep[i], SuTType[i], AP, SuLP[i], SucN[i], SuNcN[i])
                    SuLP[i, SulN[i], AP] += 1
                    if not np.all(SuTType[i] == SuTTypep[i]):
                        for j in SuNcN[i]:
                            SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] = LPTType(SuAiPt[i], SuTType[i], SuLP[i], j, 0)
                            if SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] == 0:
                                delpath = np.append(delpath, i)
                                RE = np.unique(np.append(RE, s * j + np.flatnonzero(GC[:, j])))
                    else:
                        if np.any(SuLP[i, SulN[i], AP] == MLP[SulN[i]]):
                            for j in SuNcN[i, (PIL(ICL, SuNcN[i]) == SulN[i])]:
                                SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] = LPTType(SuAiPt[i], SuTType[i], SuLP[i], j, 0)
                                if SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] == 0:
                                    delpath = np.append(delpath, i)
                                    RE = np.unique(np.append(RE, s * j + np.flatnonzero(GC[:, j])))
                        for j in SuNcN[i][CC[SucN[i], SuNcN[i]] > 0]:
                            SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] = LPTType(SuAiPt[i], SuTType[i], SuLP[i], j, 0)
                            if SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] == 0:
                                delpath = np.append(delpath, i)
                                RE = np.unique(np.append(RE, s * j + np.flatnonzero(GC[:, j])))
                delpath = np.unique(delpath)
                SuTType = np.delete(SuTType, delpath, axis=0)
                SuAiPt = np.delete(SuAiPt, delpath, axis=0)
                SuLP = np.delete(SuLP, delpath, axis=0)
                SuLlcN = np.delete(SuLlcN, delpath, axis=0)
                SuNcN = np.delete(SuNcN, delpath, axis=0)
                SuRcN = np.delete(SuRcN, delpath, axis=0)
                SucN = np.delete(SucN, delpath, axis=0)
            else:
                out = 1
    if fastsend == 0:
        delin = np.array([], int)
        for i in range(len(REGC)):
            if Allin(GC, REGC[i, :REg[i], :]) == 1:
                delin = np.append(delin, i)
        if len(delin) > 0:
            REGC = np.delete(REGC, delin, axis=0)
            REg = np.delete(REg, delin)
        REGC = np.append(REGC, [np.append(GC, np.zeros((s - g, c)), axis=0)], axis=0)
        REg = np.append(REg, g)
    fastsend = 0
    if distinguished == 0:
        SCbr = np.copy(SC)
        delind = np.array([], int)
        for i in range(len(RE)):
            classnumber = RE[i] // s
            lecture = PIL(ICL, np.array([classnumber]))[0]
            if LCN[lecture] == 1:
                LLC = np.append(LLC, lecture)
                delind = np.append(delind, i)
            elif np.sum(SC[:, classnumber]) == 1:
                LMC = np.append(LMC, lecture)
                delind = np.append(delind, i)
        if len(delind) > 0:
            RE = np.delete(RE, delind, axis=0)
        RE = np.sort(RE)
        distinguished = 1
    newface = 0
    while newface == 0:
        classnum = RE[-1] // s
        lecture = PIL(ICL, np.array([classnum]))[0]
        student = GS[RE[-1] % s]
        SC = np.copy(SCbr)
        SC[student, classnum] = 0
        SC[student, classnum + 1 - LCN[lecture] * (classnum == FCL[lecture])] = 1
        RE = np.delete(RE, -1)
        SC = sortition(SC)
        if SC in RSC:
            if len(RE) == 0:
                dead = 1
                newface = 1
        else:
            RSC = np.append(RSC, [SC], axis=0)
            out = 0
            newface = 1
print("Lectures with less classes: " + str(LLC))
print("Lectures with much classes: " + str(LMC))
