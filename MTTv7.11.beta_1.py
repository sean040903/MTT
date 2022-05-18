import numpy as np
import pandas as pd
import math
import openpyxl
from numba import njit
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numba

# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
"""print('학생들의 이름이 담긴 길이 s짜리 xlsx파일의 이름이 뭔가요?')"""
StudentNamesFile_path = 'StudentNames.xlsx'
dfStudentNames = pd.read_excel(StudentNamesFile_path)
StudentNames = dfStudentNames.values[0]
s = len(StudentNames)
S = np.arange(s)

"""print('과목들의 이름이 담긴 길이 l짜리 xlsx파일의 이름이 뭔가요?')"""
LectureNamesFile_path = 'LectureNames.xlsx'
dfLectureNames = pd.read_excel(LectureNamesFile_path)
LectureNames = dfLectureNames.values[0]
l = len(LectureNames)


"""print('학생들의 수강 유무가 있는 s x l짜리 xlsx파일의 이름이 뭔가요?')"""
StudentLectureFile_path = 'SL.xlsx'
dfStudentLecture = pd.read_excel(StudentLectureFile_path)
SL = dfStudentLecture.values

"""print('각 과목별로 시수,종류가 있는 l x 2짜리 xlsx파일의 이름이 뭔가요?')"""
KLFile_path = 'KindLecture.xlsx'
dfKindLecture = pd.read_excel(KLFile_path)
KL = dfKindLecture.values

"""print('각 과목별로 한 교시에 가능한 최대 분반의 개수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
MCPLFile_path = 'MCPL.xlsx'
dfMCPL = pd.read_excel(MCPLFile_path)
MCPL = dfMCPL.values[0]

"""print('각 과목별 분반의 수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
LCNFile_path = 'LCN.xlsx'
dfLCN = pd.read_excel(LCNFile_path)
LCN = dfLCN.values[0]

"""print('1주일동안 수업이 진행되는 날들의 교시수들이 담긴 xlsx파일의 이름은 뭔가요?')"""
DPNFile_path = 'DailyPN.xlsx'
dfDailyPN = pd.read_excel(DPNFile_path)
DPN = dfDailyPN.values[0]

"""print('과목별로 진행되는 강의실의 기하학적 중심에서 다른 과목의 강의실로 가는데 이동해야 하는 실거리가 담긴 l x l짜리 xlsx파일의 이름이 뭔가요?')"""
DistanceFile_path = 'DI.xlsx'
dfDistance = pd.read_excel(DistanceFile_path)
DI = dfDistance.values

"""print('각 학생별로 모든 과목들을 자신에게 중요하다고 생각하는 것들의 순위(0부터시작하는 순위)가 담긴 s x l짜기 xlsx파일의 이름이 뭔가요?')"""
StudentimpFile_path = 'Studentimp.xlsx'
dfStudentimp = pd.read_excel(StudentimpFile_path)
Simp = dfStudentimp.values

"""print('매 요일의 이름이 담긴 xlsx파일의 이름은 뭔가요?')"""
datenamesFile_path = 'DateNames.xlsx'
dfdatenames = pd.read_excel(datenamesFile_path)
datenames = dfdatenames.values[0]

"""print('소수 500개 담긴 xlsx파일의 이름은 뭔가요?')"""
PrimeFile_path = 'PrimeList.xlsx'
dfprimes = pd.read_excel(PrimeFile_path)
Prime = dfprimes.values[0]



# 추가 함수

def SelT(k, arr1, n):
    if k == 2:
        return np.array([arr1[n]])
    elif k == 0:
        return np.array([arr1[n], arr1[n] + 1])
    elif k == 4:
        return np.array([arr1[n][0], arr1[n][1]])
    elif k == 3:
        return np.array([arr1[n][0], arr1[n][0] + 1, arr1[n][1]])
    else:
        return np.array([arr1[n][0], arr1[n][0] + 1, arr1[n][1], arr1[n][1] + 1])


def ProS(arr1, arr2):
    pros1 = 1
    for i in range(len(arr2)):
        pros1 *= arr1[arr2[i]]
    return pros1


def Primes(arr1):
    primes1 = np.array([],int)
    for i in range(len(arr1)):
        primes1 = np.append(primes1, [Prime[arr1[i]]])
    return primes1


def DalO(n, arr1, arr2):
    for i in range(len(arr2)):
        if n % arr1[arr2[i]] == 0:
            return 0
    return 1


def MulS(arr1, arr2, n):
    arr = np.copy(arr1)
    for i in range(len(arr2)):
        arr[arr2[i]] *= n
    return arr


def APTType(k, arr):
    if k == 2:
        return Odd(arr)
    elif k == 0:
        return Even(arr)
    elif k == 4:
        O1 = Odd(arr)
        ap1 = np.empty((0, 2), int)
        for p1 in range(len(O1)):
            arr1 = CTType(1, O1[p1], arr)
            O2 = Odd(arr1)
            for p2 in range(len(O2)):
                if date(O1[p1]) < date(O2[p2]):
                    ap1 = np.append(ap1, np.array([[O1[p1], O2[p2]]]), axis = 0)
        return ap1
    elif k == 3:
        E1 = Even(arr)
        ap2 = np.empty((0, 2), int)
        for p1 in range(len(E1)):
            mat1 = CTType(2, E1[p1], arr)
            O1 = Odd(mat1)
            for p2 in range(len(O1)):
                if date(O1[p2]) != date(E1[p1]):
                    ap2 = np.append(ap2, np.array([[E1[p1], O1[p2]]]), axis = 0)
        return ap2
    else:
        E1 = Even(arr)
        ap3 = np.empty((0, 2), int)
        for p1 in range(len(E1)):
            arr1 = CTType(2, E1[p1], arr)
            E2 = Even(arr1)
            for p2 in range(len(E2)):
                if date(E1[p1]) < date(E2[p2]):
                    ap3 = np.append(ap3, np.array([[E1[p1], E2[p2]]]), axis = 0)
        return ap3


def NPTType(k, arr):
    return len(APTType(k, arr))


def MEx(arr, n):
    mex1 = 0
    for i in range(len(arr)):
        mex1 = max(mex1, Pm(arr[i], n))
    return mex1


def SelM(arr1, arr2):
    arr = np.array([],int)
    for i in range(len(arr2)):
        arr = np.append(arr, [arr1[arr2[i]]])
    return arr


def CSE(arr1, arr2, m, n):
    arr = np.copy(arr1)
    for i in range(len(arr2)):
        arr[m][arr2[i]] = n
    return arr


def kitok(ki):
    if ki == 2:
        return 1
    elif ki == 0:
        return 2
    elif ki == 4:
        return 11
    elif ki == 3:
        return 12
    else:
        return 22


# 행렬-원소 선택 함수들#
def PIL(arr, n):
    if arr[-1] <= n:
        return len(arr) - 1
    else:
        for i in range(len(arr) - 1):
            if arr[i] <= n < arr[i + 1]:
                return i


def NED(n, arr):
    for i2 in range(len(arr)):
        if n % arr[i2] == 0:
            return 0
    return 1


# DPN 활용#
daily = max(DPN)
weekly = len(DPN)
IPD = np.array([0])
FPD = np.array([],int)
for i3 in range(weekly - 1):
    IPD = np.append(IPD, [IPD[-1] + DPN[i3]])
    FPD = np.append(FPD, [IPD[-1] - 1])
wpn = sum(DPN)
FPD = np.append(FPD, [wpn])


def dates(arr):
    Dates = np.array([],int)
    for i4 in range(len(arr)):
        Dates = np.append(Dates, [PIL(IPD, arr[i4])])
    return Dates


def date(n):
    return PIL(IPD, n)


def CTType(ki, SelP, arr):
    arr2 = np.copy(arr)
    for i5 in range(weekly):
        if arr[i5] < DMEP[i5]:
            if SelP.size == 1:
                if SelP == EP[PIL(EP, arr[i5])]:
                    arr2[i5] += 2
            else:
                print(SelP)
                if SelP[0] == EP[PIL(EP, arr[i5])]:
                    arr2[i5] += 2
                if SelP[1] == EP[PIL(EP, arr[i5])]:
                    arr2[i5] += 2
    if SelP.size == 1:
        Selp = np.array([SelP])
    else:
        Selp = np.copy(SelP)
    DATES = np.copy(dates(Selp))
    k = kitok(ki)
    bfs = format(26 - k, 'b')
    for j in range(len(Selp)):
        if DATES[j] in DCD:
            arr2[weekly + np.where(arr2 == 2)[0]] = 1
        if bfs[-j - 1] == 1:
            if Selp[j] in EP or Selp[j] in VEP:
                arr2[-1] = BFN(np.union1d(NZBF(arr[-1]), np.array([len(EP) - PIL(EP, Selp[j])])))
    if len(Selp) > 1:
        if Selp[0] + np.sum(DPN[DATES[0]:DATES[1]]) == Selp[1] and DATES in SDPD and bfs[-1] == bfs[-2]:
            i0 = np.where(DCD == DATES[0])[0] + weekly
            if arr[i0] == 0:
                arr2[i0] = 0
    return arr2


# making classes#
ICL = np.array([0])
FCL = np.array([],int)
for i6 in range(l):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL = np.append(ICL, [a1 + a2])
    FCL = np.append(FCL, [a1 + a2 - 1])
NSL = np.sum(SL, axis = 0)



def CGS(i, arr):
    cgs1 = np.array([],int)
    for j in range(len(arr)):
        if arr[j][i] == 1:
            cgs1 = np.append(cgs1, [j])
    return cgs1


c = np.sum(LCN)
C = np.arange(c)
SC = np.zeros((s, c))
for l1 in range(l):
    strangearr = np.array_split(np.arange(NSL[l1]),LCN[l1])
    for i in range(LCN[l1]):
        for j in range(len(strangearr[i])):
            arr8 = CGS(l1, SL)
            alpha2 = strangearr[i][j]
            SC[arr8[alpha2]][ICL[l1] + i] = 1
KC = np.empty((0, 2), int)
for c1 in range(c):
    KC = np.append(KC,[KL[PIL(ICL, c1)-1]],axis=0)
# SC sorting#
SCN = np.sum(SC, axis = 1)
dfSC = pd.DataFrame(SC, columns=np.arange(c))
dfSC['SCN'] = SCN
dfSC['S'] = np.arange(s)
dfSC = dfSC.sort_values('SCN')
StS = dfSC.values[:, -1]
dfSC.drop(['SCN', 'S'], axis = 1, inplace=True)
SC = dfSC.values


# 반 분류 ICLK 생성#
ICLK = np.array([0])
for i in range(c - 1):
    if KC[i + 1][0] != KC[i][0]:
        ICLK = np.append(ICLK, [i + 1])


# 소수,진법 변환 관련 함수들#
def Pm(n1, n2):
    multiply = 0
    while n1 % (Prime[n2] ** multiply) == 0:
        multiply += 1
        return multiply - 1


def NZBF(n):
    nzbf = np.array([],int)
    nbf = np.binary_repr(n)
    for i in range(0, len(nbf)):
        if nbf[i] == '1':
            nzbf = np.append(nzbf, [i])
    return nzbf


def BFN(arr):
    n = 0
    for i in range(0, len(arr)):
        n += 2 ** arr[i]
    return n


# 행렬 연속 몇개 선택 함수들#
def MIO(arr, i, n):  # Multiply In Order#
    mio = 1
    for j in range(i, i + n):
        mio *= arr[j]
    return mio


def SelIO(arr, i, n):
    return arr[i:i + n]


def C2N(x, y):
    return np.arange(y)[x:y:2]


# 행렬-행렬관계 함수#
def CIO(arr1, arr2):
    for i in range(0, len(arr1)):
        if arr1[i] > arr2[i]:
            return 1
        elif arr2[i] > arr1[i]:
            return 2
    return 0


# 교시교환 관련 함수들#
def LW1P(da, arr):
    a = IPD[da]
    tw = np.zeros((2, 2))
    for i in range(2):
        for j in range(2):
            for s1 in range(s):
                c1 = arr[s1][a + 1 - i]
                c2 = arr[s1][a + 2 + j]
                if max(c1, c2) < c:
                    l1 = PIL(ICL, c1)
                    l2 = PIL(ICL, c2)
                    tw[i][j] += DI[l1][l2]
    x = np.where(tw == np.min(tw))[0][0]
    y = np.where(tw == np.min(tw))[1][0]
    mwp = np.array([a + x, a + 1 - x, a + 2 + y, a + 3 - y])
    if DPN[da] % 2 == 1:
        mwp = np.append(mwp, [a + 4])
    return mwp


def LW2P(da, arr):
    a = IPD[da]
    tw = np.zeros((3, 2, 2))
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for s1 in range(s):
                    c1 = arr[s1][a + 2 * (i % 3) + 1 - i]
                    c2 = arr[s1][a + 2 * ((i + 1) % 3) + j]
                    if max(c1, c2) < c:
                        l1 = PIL(ICL, c1)
                        l2 = PIL(ICL, c2)
                        tw[i][j] += DI[l1][l2]
    x = np.where(tw == np.min(tw))[0][0]
    y = np.where(tw == np.min(tw))[1][0]
    z = np.where(tw == np.min(tw))[2][0]
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1])


def LW3P(da, arr):
    a = IPD[da]
    tw = np.zeros((3, 2, 2, 2))
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for o in range(2):
                    for s1 in range(s):
                        c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
                        c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
                        c3 = arr[s1][a + 2 * ((i + 2) % 3) + 1 - o]
                        c4 = arr[s1][a + 6]
                        if max(c1, c2) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
                        if max(c3, c4) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
    x = np.where(tw == np.min(tw))[0][0]
    y = np.where(tw == np.min(tw))[1][0]
    z = np.where(tw == np.min(tw))[2][0]
    w = np.where(tw == np.min(tw))[3][0]
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6])


def LWP(arr):
    tmwp = np.array([],int)
    for i in range(weekly):
        if DPN[i] < 6:
            tmwp = np.append(tmwp, LW1P(i, arr))
        elif DPN[i] == 6:
            tmwp = np.append(tmwp, LW2P(i, arr))
        else:
            tmwp = np.append(tmwp, LW3P(i, arr))
    return tmwp


# EP,VEP관련 함수들#
def EPs(arr):
    eps = np.array([],int)
    for i in range(len(arr)):
        eps = np.append(eps, [EP[arr[i]]])
    return eps


def VEPs(arr):
    veps = np.array([],int)
    for i in range(len(arr)):
        veps = np.append(veps, [VEP[arr[i]]])
    return veps


def Even(arr):
    eap = np.array([],int)
    for i1 in range(weekly):
        if i1 not in CAD or arr[SDPI[np.where(CAD == i1)[0][0]]] == 1:
            eap = np.append(eap, C2N(IPD[i1], arr[i1] + 2))
    return eap


def Odd(arr):
    oap = np.array([],int)
    for i10 in range(weekly):
        if i10 not in CAD or arr[SDPI[np.where(CAD == i10)[0][0]]] == 1:
            oap = np.append(oap, np.arange(arr[i10] + 2)[IPD[i10]:arr[i10] + 2:2])
            if DPN[i10] % 2 == 1:
                oap = np.append(oap, [NEP[np.where(NEPI == i10)[0][0]]])
    for i11 in range(len(NZBF(arr[-1]))):
        oap = np.append(oap, [VEP[NZBF(arr[-1])[i11]]])
    return oap


# 그래프 색깔#
"""SPC = np.zeros(12)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
Math = []
Phy = []
Che = []
CoSc = []
LiSc = []
EaSc = []
ConS = []
ScEx = []
CCSL = []
Lan = []
Soc = []
MAS = []


def CSP(n):
    if n == 0:
        return Math
    elif n == 1:
        return Phy
    elif n == 2:
        return Che
    elif n == 3:
        return CoSc
    elif n == 4:
        return LiSc
    elif n == 5:
        return EaSc
    elif n == 6:
        return ConS
    elif n == 7:
        return ScEx
    elif n == 8:
        return CCSL
    elif n == 9:
        return Lan
    elif n == 10:
        return Soc
    else:
        return MAS


for c1 in range(c):
    partn = KC[c1][1]
    CSP(partn).append(c1)
    SPC[partn] += 1


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = []
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1.append(i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
        else:
            for G in range(k - R, -1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
    return RGB
"""
PTTcolorset = np.array(
    ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff', 'b8ffe4',
     'd8bfd8', 'ffca99', 'ff69b4'])

# TType 설정 코드#
UDP = np.union1d(IPD, IPD + 4)
DEPN = DPN // 2
EP = np.array([],int)
for i in range(weekly):
    EP = np.append(EP, IPD[i])
    for j in range(DEPN[i] - 1):
        EP = np.append(EP, [EP[-1] + 2])
VEP = EP + 1
NEP = np.setdiff1d(np.arange(wpn), np.union1d(EP, VEP))
NEPI = dates(NEP)
DMEP = np.add(IPD, 2 * DEPN) - 2
SUDPI = np.array([],int)
SDPI = np.array([],int)
SDPD = np.empty((0, 2), int)
SUDPD = np.array([],int)
DCD = np.array([],int)
CAD = np.array([],int)
fTType = np.copy(IPD)
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            if DPN[i] % 2 == 1:
                SUDPI = np.append(SUDPI, [len(fTType)])
                SUDPD = np.append(SUDPD, [j])
            SDPI = np.append(SDPI, [len(fTType)])
            fTType = np.append(fTType, [0])
            SDPD = np.append(SDPD, [[i, j]],axis = 0)
            DCD = np.append(DCD, [i])
            CAD = np.append(CAD, [j])
            fi = 0
fTType = np.append(fTType, [0])
ADS = np.empty((0, weekly), int)
ADS = np.append(ADS, [np.arange(weekly)],axis = 0)
nds = 1
done1 = 1
while done1 == 1:
    i = 0
    j = 0
    re = 1
    while re * (nds - i) > 0:
        ADS1 = ADS[i]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = ADS1
            ADS2[k0] = ADS1[k1]
            ADS2[k1] = ADS1[k0]
            if not ADS2 in ADS:
                ADS = np.append(ADS, [ADS2], axis = 0)
                re = 0
                nds += 1
            else:
                j += 1
        if re == 1:
            i += 1
    if re == 1:
        done1 = 0


# 엑셀 조작 관련#
def style(workbook, n, i, j, arr5):  # workbook: wb, n:학생,nsn:새로운 시트,i:요일,j:교시, list2=list3[s1]
    newsheet = workbook[sheetname(n)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif arr5[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = np.unique(arr5)
        containedclasses = np.delete(containedclasses, -1, 0)
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == arr5[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[0],
                                                               fill_type='solid')


def sheetname(b):
    return str[b] + '번째 학생'


def makingsheetdone(arr, wbname):
    wb = openpyxl.Workbook()
    for i in range(s):
        wb.create_sheet(sheetname(i))
    arr3 = np.empty((0, weekly, daily), int)
    for s3 in range(s):
        arr2 = np.empty((0, daily), int)
        for i in range(weekly):
            arr2 = np.append(arr2, [np.append(arr[s3][IPD[i]:FPD[i]], [c] * (daily - DPN[i]))],axis =1)
        arr3 = np.append(arr3, [arr2],axis = 0)
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(s):
            for i in range(weekly):
                for j in range(daily):
                    c5 = arr3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, c5)
                        pc = (c5 - ICL[l1] + 1)
                        arr3[s0] = LectureNames[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(arr3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0))
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(arr)):
        i = 0
        s2 = wb[sheetname(s1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, arr3[s1])
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


# highest priority control codes#
control = 0
back = 0
came = 0
Found = 0
sdca = 1
# highest priority control codes#


# main variables define code#
Rg = np.array([],int)
RGSCP = np.array([],int)
IGSCP = np.array([],int)
RWG = np.array([],int)
RWC = np.array([],int)
onestudentlecture = np.array([],int)
oneclasslecture = np.array([],int)
CpTs = np.empty((0, c, wpn), int)
FScTs = np.array([],int)
RFScT = np.array([],int)  # 가장 최종적으로 얻고자 하는 것들 중에서 하나#
Rtottdw = np.array([],int)
RSC = np.empty((0, s, c), int)
while control == 0:
    RpL = np.array([],int)
    RpN = np.array([],int)
    pN = 0
    cN = 0
    gN = 0
    CC = np.zeros((c, c))
    for s1 in range(s):
        SS = np.zeros((c, c))
        SS[s1] = np.copy(SC[s1])
        SS1 = np.copy(SS)
        SS2 = np.copy(SS.T)
        CC += SS2 @ SS1
    for i in range(c):
        for j in range(i):
            if CC[i][j] == 0:
                CC[i][j] = s
                CC[j][i] = s
    g = 0
    GS = np.array([],int)
    GSC = np.empty((0, c), int)
    GSCPlist = []
    for s2 in range(s):
        SS = np.zeros((c, c))
        SS[s2] = np.copy(SC[s2])
        SS1 = np.copy(SS)
        SS2 = np.copy(SS.T)
        if np.min(CC - SS2 @ SS1) > 0:
            CC -= SS2 @ SS1
        else:
            GS = np.append(GS, [s2])
            GSC = np.append(GSC, [SC[s2]],axis = 0)
            g += 1
            gscp = 1
            for i in range(c):
                if SC[s2][i] == 1:
                    gscp = gscp * (Prime[i])
            GSCPlist.append(gscp)
    GSCP = np.array(GSCPlist)
    if g in Rg:
        i = 0
        while i < len(IGSCP) - 1:
            PGSCP = RGSCP[IGSCP[i]:IGSCP[i + 1]]
            if np.sort(PGSCP) == np.sort(GSCP):
                cN = c
                i = len(IGSCP) - 1
            else:
                i += 1
    else:
        Rg = np.append(Rg, [g])
        RGSCP = np.append(RGSCP, GSCP)
        IGSCP = np.append(IGSCP, [len(RGSCP)])
    if cN == 0:
        RGSTP = np.array([np.ones(g)])
        RTType = np.array([fTType])
        RCLP = np.array([np.ones(g)])
        RWC = np.array([],int)
        RWG = np.array([],int)
        while cN < c:
            ki = PIL(ICLK, cN)
            RpL = np.append(RpL,[NPTType(ki, RTType[-1])])
            while pN < RpL[-1] + 1:
                if cN == c:
                    CpT = np.zeros((c, wpn))
                    SpT = np.zeros((s, wpn))
                    ScT = np.ones((s, wpn)) * c
                    SaT = np.ones((s, wpn)) * 12
                    SiT = np.ones((s, wpn)) * l
                    for s1 in range(s):
                        for c1 in range(c):
                            if SC[s1][c1] == 1:
                                TType = RTType[c1]
                                ki = PIL(ICLK, c1)
                                Selt2 = SelT(ki, APTType(ki, TType), RpN[c1])
                                p1 = Prime[c1]
                                CpT = CSE(CpT, Selt2, c1, p1)
                                l1 = PIL(ICL, c1)
                                a1 = KC[c1][1]
                                im1 = Simp[s1][l1]
                                SpT = CSE(SpT, Selt2, s1, p1)
                                SaT = CSE(SaT, Selt2, s1, a1)
                                ScT = CSE(ScT, Selt2, s1, c1)
                                SiT = CSE(SiT, Selt2, s1, im1)
                    s2 = 0
                    while s2 < s * sdca:
                        d1 = 0
                        while d1 < weekly * sdca:
                            t1 = 1
                            while t1 < DPN[d1] * sdca:
                                t2 = 0
                                while 0 < (t1 - t2) * sdca:
                                    p2 = IPD[d1] + t1
                                    p3 = IPD[d1] + t2
                                    if SaT[s2][p2] == SaT[s2][p3] and ScT[s2][p2] != ScT[s2][p3]:
                                        sdca = 0
                                    t2 += 1
                                t1 += 1
                            d1 += 1
                        s2 += 1
                    if sdca == 1:
                        MTDW = LWP(ScT)
                        MAPF = np.copy(MTDW)
                        for d2 in range(weekly):
                            if DPN[d2] > 3:
                                p4 = IPD[d2]
                                Im12 = 0
                                Im34 = 0
                                Selt3 = SelT(0, MAPF, p4)
                                Selt4 = SelT(0, MAPF, p4 + 2)
                                for s3 in range(s):
                                    Im12 += np.sum(SelM(SiT[s3], Selt3))
                                    Im34 += np.sum(SelM(SiT[s3], Selt4))
                                Ims = np.array([Im12, Im34])
                                m1 = np.where(Ims == np.min(Ims))[0][0]
                                MAPF[p4] = MTDW[p4 + 3 - 3 * m1]
                                MAPF[p4 + 1] = MTDW[p4 + 2 - m1]
                                MAPF[p4 + 2] = MTDW[p4 + 1 + m1]
                                MAPF[p4 + 3] = MTDW[p4 + 3 * m1]
                                if DPN[d2] == 6:
                                    p5 = p4 + 4
                                    Im5 = 0
                                    Im6 = 0
                                    for s1 in range(s):
                                        Im5 += SiT[s1][MAPF[p5]]
                                        Im6 += SiT[s1][MAPF[p5 + 1]]
                                    if Im6 > Im5:
                                        MAPF[p5] = MTDW[p5 + 1]
                                        MAPF[p5 + 1] = MTDW[p5]
                        CTSN = np.array([],int)
                        for c1 in range(c):
                            CTSN = np.append(CTSN, [len(CGS(c1, SC))])
                        SCDSD = np.array([],int)
                        for i in range(nds):
                            ADS3 = np.copy(ADS[i])
                            qdsd = 0
                            tdsd = 0
                            pdsd = 0
                            for c1 in range(ICLK[1], ICLK[2]):
                                APs1 = APTType(1, RTType[c1])
                                p6 = RpN[c1]
                                qdsd += abs(np.where(ADS3 == date(APs1[p6][0])) - np.where(ADS3 == date(APs1[p6][1]))) * \
                                        CTSN[c1]
                            for c2 in range(ICLK[3], ICLK[4]):
                                APs2 = APTType(3, RTType[c2])
                                p7 = RpN[c2]
                                tdsd += abs(np.where(ADS3 == date(APs2[p7][0])) - np.where(ADS3 == date(APs2[p7][1]))) * \
                                        CTSN[c2]
                            for c3 in range(ICLK[4], c):
                                APs3 = APTType(4, RTType[c3])
                                p8 = RpN[c3]
                                pdsd += abs(np.where(ADS3 == date(APs3[p8][0])) - np.where(ADS3 == date(APs3[p8][1]))) * \
                                        CTSN[c3]
                            scdsd = qdsd + tdsd + pdsd
                            SCDSD = np.append(SCDSD, [scdsd])
                        BSP = ADS[np.argmin(SCDSD)]
                        CDSD = np.array([],int)
                        for d3 in range(weekly):
                            for p9 in range(IPD[BSP[d3]], FPD[BSP[d3]]):
                                CDSD = np.append(CDSD, [MAPF[p9]])
                        dfScTd = pd.DataFrame(ScT, columns=np.arange(len(CDSD)))
                        ScTD = dfScTd.values
                        CpTs = np.append(CpTs, [CpT],axis = 0)
                        tottdw = 0
                        for s1 in range(s):
                            sctd = ScTD[s1]
                            for t1 in range(wpn):
                                c1 = sctd[t1]
                                c2 = sctd[t1 + 1]
                                l1 = PIL(ICL, c1)
                                l2 = PIL(ICL, c2)
                                tottdw += (DI[l1][l2])
                            for i in range(len(UDP)):
                                c3 = sctd[UDP[i]]
                                c4 = sctd[UDP[i] + 1]
                                l3 = PIL(ICL, c3)
                                l4 = PIL(ICL, c4)
                                tottdw -= (DI[l3][l4])
                        RFScT = np.append(RFScT, [ScTD.reshape(-1)],axis = 0)
                        FScTs = np.append(FScTs, [np.append(ScTD.reshape(-1), [tottdw])],axis = 0)
                        Rtottdw = np.append(Rtottdw, [tottdw])
                        RSC = np.append(RSC, [SC],axis = 0)
                while back == 1:
                    if Found == 1:
                        print("Found One")
                    else:
                        printone = '\r' + 'class:' + str(cN) + ',' + " ".join(map(str, RpN))
                        print(printone, end='')
                    cN -= 1
                    RpL = np.delete(RpL, -1, 0)
                    came = 1
                    pN = RpN[-1] + 1
                    RpN = np.delete(RpN, -1, 0)
                    RGSTP = np.delete(RGSTP, -1, 0)
                    RTType = np.delete(RTType, -1, 0)
                    RCLP = np.delete(RCLP, -1, 0)
                    """print(RpN)"""
                    back = 0
                    if pN == RpL[-1]:
                        if cN == 0:
                            control = 1
                            cN = c
                            pN += 1
                        else:
                            back = 1
                if control == 0:
                    GSTP = np.copy(RGSTP[-1])
                    CLP = np.copy(RCLP[-1])
                    TType = np.copy(RTType[-1])
                    gN = 0
                    APs = APTType(ki, TType)
                    lN = PIL(ICL, cN)
                    i = 0
                    list111 = CGS(cN, GSC)
                    if came == 0:
                        pN = 0
                    while i < len(list111):
                        gN = list111[i]
                        Selt1 = SelT(ki, APs, pN)
                        if pN == len(APs):
                            back = 1
                            i += 1
                        else:
                            if DalO(GSTP[gN], Prime, Selt1) == 0:
                                RWC = np.append(RWC, [cN])
                                RWG = np.append(RWG, [gN])
                                pN += 1
                                i = 0
                                if pN == RpL[-1]:
                                    back = 1
                                    i = len(list111) + 1
                            else:
                                i += 1
                        if i == len(list111):
                            if MEx(SelM(CLP, SelT(ki, APs, pN)), lN) == MCPL[lN]:
                                pN += 1
                                if pN == RpL[-1]:
                                    back = 1
                                    gN = g + 1
                            else:
                                CLP = MulS(CLP, Selt1, Prime[lN])
                                RCLP = np.append(RCLP, [CLP],axis = 0)
                                GSTP = MulS(GSTP, CGS(cN, GSC), ProS(Prime, Selt1))
                                RGSTP = np.append(RGSTP, [GSTP],axis = 0)
                                RpN = np.append(RpN, [pN])
                                TType = CTType(ki, APs[pN], TType)
                                RTType = np.append(RTType, [TType],axis = 0)
                                cN += 1
                                if came == 1:
                                    came = 0
                                pN = 0
                                ki = PIL(ICLK, cN)
                                RpL = np.append(RpL, [NPTType(ki, RTType[-1])])
                                print(RpN)
                                printone = '\r' + 'class:' + str(cN) + ',' + " ".join(map(str, RpN))
                                print(printone, end='')
    if control == 1:
        end = 1
        i = 0
        while i < len(RWC):
            c4 = RWC[-i - 1]
            if CGS(c4, GSC)[0] > 1 and LCN[PIL(ICL, c4)] > 1:
                SC[RWG[-i - 1]][RWC[-i - 1]] = 0
                if RWC[-i - 1] in FCL:
                    SC[RWG[-i - 1]][ICL[np.where(FCL == RWC[-i - 1])[0][0]]] = 1
                else:
                    SC[RWG[-i - 1]][RWC[-i - 1] + 1] = 1
                end = 0
                i = len(RWC)
                control = 0
            else:
                if CGS(c4, GSC)[0] == LCN[PIL(ICL, c4)]:
                    onestudentlecture = np.append(onestudentlecture, [PIL(ICL, c4)])
                    oneclasslecture = np.append(oneclasslecture, [PIL(ICL, c4)])
                else:
                    if CGS(c4, GSC)[0] == 1:
                        onestudentlecture = np.append(onestudentlecture, [PIL(ICL, c4)])
                    else:
                        oneclasslecture = np.append(oneclasslecture, [PIL(ICL, c4)])
        if end == 1:
            if len(FScTs) == 0:
                print("This lectures haves too many classes and make error.", onestudentlecture)
                print("This lectures haves too few classes and make error.", oneclasslecture)
            else:
                FScTs = FScTs[np.lexsort(FScTs.T)]
                sortFScTs = np.delete(FScTs, -1, axis = 0)
                for i in range(len(sortFScTs)):
                    workbookname = str(i + 1) + '번째 시간표'
                    ScT10 = sortFScTs[i].reshape(s, wpn)
                    makingsheetdone(ScT10, workbookname)