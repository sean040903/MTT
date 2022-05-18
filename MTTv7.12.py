import numpy as np
import pandas as pd
import openpyxl
from numba import njit
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
"""print('학생들의 이름이 담긴 길이 s짜리 xlsx파일의 이름이 뭔가요?')"""
dfStudentNames = pd.read_excel('StudentNames.xlsx')
StudentNames = dfStudentNames.values.tolist()[0]
s = len(StudentNames)

"""print('과목들의 이름이 담긴 길이 l짜리 xlsx파일의 이름이 뭔가요?')"""
dfLectureNames = pd.read_excel('LectureNames.xlsx')
LectureNames = dfLectureNames.values.tolist()[0]
l = len(LectureNames)

"""print('학생들의 수강 유무가 있는 s x l짜리 xlsx파일의 이름이 뭔가요?')"""
dfStudentLecture = pd.read_excel('SL.xlsx')
SL = dfStudentLecture.values.tolist()

"""print('각 과목별로 시수,종류가 있는 l x 2짜리 xlsx파일의 이름이 뭔가요?')"""
dfKindLecture = pd.read_excel('KindLecture.xlsx')
KL = dfKindLecture.values.tolist()

"""print('각 과목별로 한 교시에 가능한 최대 분반의 개수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
dfMCPL = pd.read_excel('MCPL.xlsx')
MCPL = dfMCPL.values.tolist()[0]

"""print('각 과목별 분반의 수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
dfLCN = pd.read_excel('LCN.xlsx')
LCN = dfLCN.values.tolist()[0]

"""print('1주일동안 수업이 진행되는 날들의 교시수들이 담긴 xlsx파일의 이름은 뭔가요?')"""
dfDailyPN = pd.read_excel('DailyPN.xlsx')
DPN = dfDailyPN.values.tolist()[0]

"""print('과목별로 진행되는 강의실의 기하학적 중심에서 다른 과목의 강의실로 가는데 이동해야 하는 실거리가 담긴 l x l짜리 xlsx파일의 이름이 뭔가요?')"""
dfDistance = pd.read_excel('DI.xlsx')
DI = dfDistance.values.tolist()

"""print('각 학생별로 모든 과목들을 자신에게 중요하다고 생각하는 것들의 순위(0부터시작하는 순위)가 담긴 s x l짜기 xlsx파일의 이름이 뭔가요?')"""
dfStudentimp = pd.read_excel('Studentimp.xlsx')
Simp = dfStudentimp.values.tolist()

"""print('매 요일의 이름이 담긴 xlsx파일의 이름은 뭔가요?')"""
dfdatenames = pd.read_excel('DateNames.xlsx')
datenames = dfdatenames.values.tolist()[0]

"""print('소수 500개 담긴 xlsx파일의 이름은 뭔가요?')"""
dfprimes = pd.read_excel('PrimeList.xlsx')
Prime = dfprimes.values.tolist()[0]


# 추가 함수

def SelT(k, mat1, n):
    if k == 2:
        return [mat1[n]]
    elif k == 0:
        return [mat1[n], mat1[n] + 1]
    elif k == 4:
        return [mat1[n][0], mat1[n][1]]
    elif k == 3:
        return [mat1[n][0], mat1[n][0] + 1, mat1[n][1]]
    else:
        return [mat1[n][0], mat1[n][0] + 1, mat1[n][1], mat1[n][1] + 1]



def ProS(mat1, mat2):
    pros1 = 1
    for i in range(len(mat2)):
        pros1 *= mat1[mat2[i]]
    return pros1



def Primes(mat1):
    primes1 = []
    for i in range(len(mat1)):
        primes1.append(Prime[mat1[i]])
    return primes1



def DalO(n, mat1, mat2):
    for i in range(len(mat2)):
        if n % mat1[mat2[i]] == 0:
            return 0
    return 1



def MulS(mat1, mat2, n):
    mat = mat1[:]
    for i in range(len(mat2)):
        mat[mat2[i]] *= n
    return mat



def APTType(k, mat):
    if k == 2:
        return Odd(mat)
    elif k == 0:
        return Even(mat)
    elif k == 4:
        O1 = Odd(mat)
        ap1 = []
        for p1 in range(len(O1)):
            mat1 = CTType(1, O1[p1], mat)
            O2 = Odd(mat1)
            for p2 in range(len(O2)):
                if date(O1[p1]) < date(O2[p2]):
                    ap1.append([O1[p1], O2[p2]])
        return ap1
    elif k == 3:
        E1 = Even(mat)
        ap2 = []
        for p1 in range(len(E1)):
            mat1 = CTType(2, E1[p1], mat)
            O1 = Odd(mat1)
            for p2 in range(len(O1)):
                if date(O1[p2]) != date(E1[p1]):
                    ap2.append([E1[p1], O1[p2]])
        return ap2
    else:
        E1 = Even(mat)
        ap3 = []
        for p1 in range(len(E1)):
            mat1 = CTType(2, E1[p1], mat)
            E2 = Even(mat1)
            for p2 in range(len(E2)):
                if date(E1[p1]) < date(E2[p2]):
                    ap3.append([E1[p1], E2[p2]])
        return ap3



def NPTType(k, mat):
    return len(APTType(k, mat))



def MEx(mat, n):
    mex1 = 0
    for i in range(len(mat)):
        mex1 = max(mex1, Pm(mat[i], n))
    return mex1



def SelM(mat1, mat2):
    mat = []
    for i in range(len(mat2)):
        mat.append(mat1[mat2[i]])
    return mat



def CSE(mat1, mat2, m, n):
    mat = mat1[:]
    for i in range(len(mat2)):
        mat[m][mat2[i]] = n
    return mat


def kitok(ki):
    if ki == 2:
        return 1
    elif ki == 0:
        return 2
    elif ki == 4:
        return 11
    elif ki == 3:
        return 12
    else:
        return 22


# 행렬-원소 선택 함수들#

def PIL(mat, n):
    if mat[-1] <= n:
        return len(mat) - 1
    else:
        for i in range(len(mat) - 1):
            if mat[i] <= n:
                if mat[i + 1] > n:
                    return i



def NED(n, mat):
    for i2 in range(0, len(mat)):
        if n % mat[i2] == 0:
            return 0
    return 1


# DPN 활용#
daily = max(DPN)
weekly = len(DPN)
IPD = [0]
FPD = []
for i3 in range(0, weekly - 1):
    IPD.append(IPD[-1] + DPN[i3])
    FPD.append(IPD[-1] - 1)
wpn = sum(DPN)
FPD.append(wpn)



def dates(mat):
    Dates = []
    for i4 in range(0, len(mat)):
        Dates.append(PIL(IPD, mat[i4]))
    return Dates



def date(n):
    return PIL(IPD, n)



def CTType(ki, SelP, mat):
    mat2 = mat[:]
    for i5 in range(0, weekly):
        if mat[i5] < DMEP[i5]:
            if type(SelP) is int:
                if SelP == EP[PIL(EP, mat[i5])]:
                    mat2[i5] += 2
            else:
                if SelP[0] == EP[PIL(EP, mat[i5])]:
                    mat2[i5] += 2
                if SelP[1] == EP[PIL(EP, mat[i5])]:
                    mat2[i5] += 2
    if type(SelP) is int:
        Selp = [SelP]
    else:
        Selp = SelP
    DATES = dates(Selp)[:]
    k = kitok(ki)
    bfs = format(26 - k, 'b')
    for j in range(0, len(Selp)):
        if not list({DATES[j]} - set(DCD)):
            mat2[weekly + DCD.index(DATES[j])] = 1
        if bfs[-j - 1] == 1:
            if not list({Selp[j]} - set(EP) - set(VEP)):
                mat2[-1] = BFN(np.union1d(np.array(NZBF(mat[-1])), np.array([len(EP) - PIL(EP, Selp[j])])))
    if len(Selp) > 1:
        if bfs[-1] == bfs[-2]:
            if not list({tuple(DATES)} - {tuple(i) for i in SDPD}):
                i0 = DCD.index(DATES[0]) + weekly
                if mat[i0] == 0:
                    if Selp[0] + sum(SelIO(DPN, DATES[0], DATES[1] - DATES[0])) == Selp[1]:
                        mat2[i0] = 0
    return mat2


# making classes#
ICL = [0]
FCL = []
for i6 in range(0, l):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL.append(a1 + a2)
    FCL.append(a1 + a2 - 1)
LTSarr = np.zeros(l)
for i7 in range(0, s):
    LTSarr += np.array(SL[i7])
LTS = LTSarr.tolist()



def DSL(n,mat):
    dsl = []
    for i8 in range(0, mat[n]):
        dsl.append(np.floor(LTS[n] * (i8 + 1) / mat[n]) - np.floor(LTS[n] * i8 / mat[n]))
    return dsl



def MDI(mat1, mat2):
    for i9 in range(0, len(mat1)):
        if mat1[i9] != mat2[i9]:
            return i9
    return len(mat1)


c = sum(LCN)
C = np.arange(c)
SC = np.zeros((s, c))
for l1 in range(0, l):
    MS = np.zeros(LCN[l1])
    for s1 in range(0, s):
        if SL[s1][l1] == 1:
            SC[s1][ICL[l1] + MDI(MS, DSL(l1,LCN))] = 1
            MS[MDI(MS, DSL(l1,LCN))] += 1
KC = []
for c1 in range(0, c):
    KC.append(KL[PIL(ICL, c1) - 1])

# SC sorting#

def CCS(n):
    return list(map(str, np.arange(0, n)))


SLN = np.zeros(s)
for s1 in range(0, s):
    SLN[s1] = sum(SC[s1])
dfSC = pd.DataFrame(SC, columns=CCS(c))
dfSC['SLN'] = SLN
dfSC['S'] = np.arange(0, s)
dfSC.sort_values('SLN', ascending=True)
StS = np.zeros(s)
for i in range(0, s):
    StS[i] = dfSC.iloc[i]['S']
dfSC.drop('SLN', axis=1, inplace=True)
dfSC.drop('S', axis=1, inplace=True)
dfSC = dfSC.reset_index(drop=True)
SC = dfSC.values.tolist()



def TotSC(mat):
    TotSC = np.zeros(c)
    for s1 in range(0, s):
        TotSC += [mat[s1]]
    return TotSC



def CGS(i,mat):
    arr1 = np.array(mat)[:,i]
    return np.nonzero(arr1 > 0)[0]



# 반 분류 ICLK 생성#
ICLK = [0]
for i in range(c - 1):
    if KC[i + 1][0] > KC[i][0]:
        ICLK.append(i + 1)


# 소수,진법 변환 관련 함수들#

def Pm(n1, n2):
    multiply = 0
    while n1 % (Prime[n2] ** multiply) == 0:
        multiply += 1
        return multiply - 1



def NBF(n):
    nbf = format(n, 'b')
    return list(map(int, nbf))



def NZBF(n):
    nzbf = []
    nbf = format(n, 'b')
    for i in range(0, len(nbf)):
        if nbf[i] == '1':
            nzbf.append(i)
    return nzbf



def BFN(mat):
    n = 0
    for i in range(0, len(mat)):
        n += 2 ** mat[i]
    return n


# 행렬 연속 몇개 선택 함수들#

def MIO(mat, i, n):  # Multiply In Order#
    mio = 1
    for j in range(i, i + n):
        mio *= mat[j]
    return mio



def SelIO(mat, i, n):
    selio = []
    for j in range(i, i + n):
        selio.append(mat[j])
    return selio



def C2N(x, y):
    c2n = []
    for i in range(x, y, 2):
        c2n.append(i)
    return c2n


# 행렬-행렬관계 함수#

def CIO(mat1, mat2):
    for i in range(0, len(mat1)):
        if mat1[i] > mat2[i]:
            return 1
        elif mat2[i] > mat1[i]:
            return 2
    return 0


# 교시교환 관련 함수들#

def LW1P(da, mat):
    a = IPD[da]
    tw = np.zeros((2, 2))
    for i in range(0, 2):
        for j in range(0, 2):
            for s1 in range(0, s):
                c1 = mat[s1][a + 1 - i]
                c2 = mat[s1][a + 2 + j]
                if max(c1, c2) < c:
                    l1 = PIL(ICL, c1)
                    l2 = PIL(ICL, c2)
                    tw[i][j] += DI[l1][l2]
    x, y = np.where(tw == np.min(tw))
    mwp = [a + x, a + 1 - x, a + 2 + y, a + 3 - y]
    if DPN[da] % 2 == 1:
        mwp.append(a + 4)
    return mwp



def LW2P(da, mat):
    a = IPD[da]
    tw = np.zeros((3, 2, 2))
    for i in range(0, 3):
        for j in range(0, 2):
            for k in range(0, 2):
                for s1 in range(0, s):
                    c1 = mat[s1][a + 2 * (i % 3) + 1 - i]
                    c2 = mat[s1][a + 2 * ((i + 1) % 3) + j]
                    if max(c1, c2) < c:
                        l1 = PIL(ICL, c1)
                        l2 = PIL(ICL, c2)
                        tw[i][j] += DI[l1][l2]
    x, y, z = np.where(tw == np.min(tw))
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return [a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1]



def LW3P(da, mat):
    a = IPD[da]
    tw = np.zeros((3, 2, 2, 2))
    for i in range(0, 3):
        for j in range(0, 2):
            for k in range(0, 2):
                for o in range(0, 2):
                    for s1 in range(0, s):
                        c1 = mat[s1][a + 2 * (i % 3) + 1 - j]
                        c2 = mat[s1][a + 2 * ((i + 1) % 3) + k]
                        c3 = mat[s1][a + 2 * ((i + 2) % 3) + 1 - o]
                        c4 = mat[s1][a + 6]
                        if max(c1, c2) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
                        if max(c3, c4) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
    x, y, z, w = np.where(tw == np.min(tw))
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return [a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6]



def LWP(mat):
    tmwp = []
    for i in range(0, weekly):
        if DPN[i] < 6:
            tmwp += LW1P(i, mat)
        elif DPN[i] == 6:
            tmwp += LW2P(i, mat)
        else:
            tmwp += LW3P(i, mat)
    return tmwp


# EP,VEP관련 함수들#

def EPs(mat):
    eps = []
    for i in range(0, len(mat)):
        eps.append(EP[mat[i]])
    return eps



def VEPs(mat):
    veps = []
    for i in range(0, len(mat)):
        veps.append(VEP[mat[i]])
    return veps



def Even(mat):
    eap = []
    for i1 in range(0, weekly):
        if not {i1} - set(CAD):
            if mat[SDPI[CAD.index(i1)]] == 1:
                eap += C2N(IPD[i1], mat[i1] + 2)
        else:
            eap += C2N(IPD[i1], mat[i1] + 2)
    return eap



def Odd(mat):
    oap = []
    for i10 in range(0, weekly):
        if not {i10} - set(CAD):
            if mat[SDPI[CAD.index(i10)]] == 1:
                oap += C2N(IPD[i10], mat[i10] + 2)
                if DPN[i10] % 2 == 1:
                    oap.append(NEP[NEPI.index(i10)])
        else:
            if IPD[i10] == mat[i10]:
                oap.append(mat[i10])
            else:
                oap += C2N(IPD[i10], mat[i10] + 2)
            if DPN[i10] % 2 == 1:
                oap.append(NEP[NEPI.index(i10)])
    for i11 in range(0, len(NZBF(mat[-1]))):
        oap.append(VEP[NZBF(mat[-1])[i11]])
    return oap


# 그래프 색깔#
"""SPC = np.zeros(12)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
Math = []
Phy = []
Che = []
CoSc = []
LiSc = []
EaSc = []
ConS = []
ScEx = []
CCSL = []
Lan = []
Soc = []
MAS = []


def CSP(n):
    if n == 0:
        return Math
    elif n == 1:
        return Phy
    elif n == 2:
        return Che
    elif n == 3:
        return CoSc
    elif n == 4:
        return LiSc
    elif n == 5:
        return EaSc
    elif n == 6:
        return ConS
    elif n == 7:
        return ScEx
    elif n == 8:
        return CCSL
    elif n == 9:
        return Lan
    elif n == 10:
        return Soc
    else:
        return MAS


for c1 in range(c):
    partn = KC[c1][1]
    CSP(partn).append(c1)
    SPC[partn] += 1


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = []
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1.append(i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
        else:
            for G in range(k - R, -1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
    return RGB
"""
PTTcolorset = ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff',
               'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']


# 기타#

def SkDc(SkT, SCT):
    for s1 in range(s):
        for t1 in range(wpn):
            dat = date(t1)
            k1 = IPD[dat]
            k2 = IPD[dat + 1]
            for t2 in range(k1, k2):
                if SkT[s1][t1] == SkT[s1][t2]:
                    if SCT[s1][t1] != SCT[s1][t2]:
                        return 1
    return 0


# TType 설정 코드#
UDP = np.union1d(np.array(IPD), np.array(IPD) + 4).tolist()
DEPN = []
for i in range(weekly):
    DEPN.append(DPN[i] // 2)
EP = []
for i in range(weekly):
    EP.append(IPD[i])
    for j in range(DEPN[i] - 1):
        EP.append(EP[-1] + 2)
VEP = (np.array(EP) + 1).tolist()
NEP = np.setdiff1d(np.arange(wpn), np.union1d(EP, VEP)).tolist()
NEPI = dates(NEP)
DMEP = []
for i in range(weekly):
    DMEP.append(IPD[i] + 2 * DEPN[i] - 2)
SUDPI = []
SDPI = []
SUDPD = []
SDPD = []
DCD = []
CAD = []
fTType = IPD[:]
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            if DPN[i] % 2 == 1:
                SUDPI.append(len(fTType))
                SUDPD.append(j)
            SDPI.append(len(fTType))
            fTType.append(0)
            SDPD.append([i, j])
            DCD.append(i)
            CAD.append(j)
            fi = 0
fTType.append(0)
ADS = [np.arange(weekly).tolist()]
nds = 1
done1 = 1
while done1 == 1:
    i = 0
    j = 0
    re = 1
    while re * (nds - i) > 0:
        ADS1 = ADS[i]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = ADS1
            ADS2[k0] = ADS1[k1]
            ADS2[k1] = ADS1[k0]
            if list({tuple(ADS2)} - {tuple(i) for i in ADS}):
                ADS.append(ADS2)
                re = 0
                nds += 1
            else:
                j += 1
        if re == 1:
            i += 1
    if re == 1:
        done1 = 0


# 엑셀 조작 관련#

def style(workbook, n, i, j, list5):  # workbook: wb, n:학생,nsn:새로운 시트,i:요일,j:교시, list2=list3[s1]
    newsheet = workbook[sheetname(n)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif list5[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(np.array(list5).reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == list5[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[0],
                                                               fill_type='solid')



def sheetname(b):
    return str[b] + '번째 학생'

@njit(parrallel=True)
def makingsheetdone(mat, wbname):
    global list4
    wb = openpyxl.Workbook()
    for i in range(len(mat)):
        wb.create_sheet(sheetname(i))
    list3 = []
    for s3 in range(len(mat)):
        list0 = mat[s3]
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
        list4 = list3[:]
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(mat)):
            for i in range(weekly):
                for j in range(daily):
                    c5 = list3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, c5)
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LectureNames[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0))
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(mat)):
        i = 0
        s2 = wb[sheetname(s1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list4[s1])
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


# highest priority control codes#
control = 0
back = 0
came = 0
Found = 0
sdca = 1
# highest priority control codes#


# main variables define code#
Rg = []
RGSCP = []
IGSCP = []
RWG = []
RWC = []
onestudentlecture = []
oneclasslecture = []
CpTs = []
FScT = []
FScTs = []
RFScT = []  # 가장 최종적으로 얻고자 하는 것들 중에서 하나#
RAvgtdw = []
RSC = []
while control == 0:
    RpL = []
    RpN = []
    pN = 0
    cN = 0
    gN = 0
    CC = np.zeros((c, c))
    for s1 in range(s):
        SS = np.zeros((c, c))
        SS[s1] = np.array(SC[s1][:])
        SS1 = SS[:]
        SS2 = SS.T[:]
        CC += SS2 @ SS1
    for i in range(c):
        for j in range(i):
            if CC[i][j] == 0:
                CC[i][j] = s
                CC[j][i] = s
    g = 0
    GS = []
    GSC = []
    GSCP = []
    for s2 in range(s):
        SS = np.zeros((c, c))
        SS[s2] = np.array(SC[s2][:])
        SS1 = SS[:]
        SS2 = SS.T[:]
        if np.min(CC - SS2 @ SS1) > 0:
            CC -= SS2 @ SS1
        else:
            GS.append(s2)
            GSC.append(SC[s2][:])
            g += 1
            gscp = 1
            for i in range(c):
                if SC[s2][i] == 1:
                    gscp *= Prime[i]
            GSCP.append(gscp)
    if not {g} - set(Rg):
        i = 0
        while i < len(IGSCP) - 1:
            PGSCP = RGSCP[IGSCP[i]:IGSCP[i + 1]]
            if sorted(PGSCP) == sorted(GSCP):
                cN = c
                i = len(IGSCP) - 1
            else:
                i += 1
    else:
        Rg.append(g)
        RGSCP += GSCP
        IGSCP.append(len(RGSCP))
    if cN == 0:
        RGSTP = [[1] * g]
        RTType = [fTType[:]]
        RCLP = [[1] * g]
        RWC = []
        RWG = []
        while cN < c:
            ki = PIL(ICLK, cN)
            RpL.append(NPTType(ki, RTType[-1]))
            while pN < RpL[-1] + 1:
                if cN == c:
                    CpT = np.zeros((c, wpn)).tolist()
                    SpT = np.zeros((s, wpn)).tolist()
                    bsct = np.ones((s, wpn)) * c
                    ScT = bsct.tolist()
                    bsat = np.ones((s, wpn)) * 12
                    SaT = bsat.tolist()
                    bsit = np.ones((s, wpn)) * l
                    SiT = bsit.tolist()
                    for s1 in range(s):
                        for c1 in range(c):
                            TType = RTType[c1]
                            ki = PIL(ICLK, c1)
                            Selt2 = SelT(ki, APTType(ki, TType), RpN[c1])
                            p1 = Prime[c1]
                            CpT = CSE(CpT, Selt2, c1, p1)
                            l1 = PIL(ICL, c1)
                            if SC[s1][c1] == 1:
                                a1 = KC[c1][1]
                                im1 = Simp[s1][l1]
                                SpT = CSE(SpT, Selt2, s1, p1)
                                SaT = CSE(SaT, Selt2, s1, a1)
                                ScT = CSE(ScT, Selt2, s1, c1)
                                SiT = CSE(SiT, Selt2, s1, im1)
                    s2 = 0
                    while s2 < s * sdca:
                        d1 = 0
                        while d1 < weekly * sdca:
                            t1 = 1
                            while t1 < DPN[d1] * sdca:
                                t2 = 0
                                while 0 < (t1 - t2) * sdca:
                                    p2 = IPD[d1] + t1
                                    p3 = IPD[d1] + t2
                                    if SaT[s2][p2] == SaT[s2][p3]:
                                        if ScT[s2][p2] != ScT[s2][p3]:
                                            sdca = 0
                                    t2 += 1
                                t1 += 1
                            d1 += 1
                        s2 += 1
                    if sdca == 1:
                        MTDW = LWP(ScT)
                        MAPF = MTDW[:]
                        for d2 in range(weekly):
                            if DPN[d2] > 3:
                                p4 = IPD[d2]
                                Im12 = 0
                                Im34 = 0
                                Selt3 = SelT(0, MAPF, p4)
                                Selt4 = SelT(0, MAPF, p4 + 2)
                                for s3 in range(s):
                                    Im12 += sum(SelM(SiT[s3], Selt3))
                                    Im34 += sum(SelM(SiT[s3], Selt4))
                                Ims = [Im12, Im34]
                                m1 = Ims.index(min(Ims))
                                MAPF[p4] = MTDW[p4 + 3 - 3 * m1]
                                MAPF[p4 + 1] = MTDW[p4 + 2 - m1]
                                MAPF[p4 + 2] = MTDW[p4 + 1 + m1]
                                MAPF[p4 + 3] = MTDW[p4 + 3 * m1]
                                if DPN[d2] == 6:
                                    p5 = p4 + 4
                                    Im5 = 0
                                    Im6 = 0
                                    for s1 in range(s):
                                        Im5 += SiT[s1][MAPF[p5]]
                                        Im6 += SiT[s1][MAPF[p5 + 1]]
                                    if Im6 > Im5:
                                        MAPF[p5] = MTDW[p5 + 1]
                                        MAPF[p5 + 1] = MTDW[p5]
                        CTSN = []
                        for c1 in range(c):
                            CTSN.append(len(CGS(c1, SC)))
                        SCDSD = []
                        for i in range(nds):
                            ADS3 = ADS[i]
                            qdsd = 0
                            for c1 in range(ICLK[1], ICLK[2]):
                                APs1 = APTType(1, RTType[c1])
                                p6 = RpN[c1]
                                oqdsd = abs(ADS3.index(date(APs1[p6][0])) - ADS3.index(date(APs1[p6][1])))
                                qdsd += oqdsd * CTSN[c1]
                            tdsd = 0
                            for c2 in range(ICLK[3], ICLK[4]):
                                APs2 = APTType(3, RTType[c2])
                                p7 = RpN[c2]
                                otdsd = abs(ADS3.index(date(APs2[p7][0])) - ADS3.index(date(APs2[p7][1])))
                                tdsd += otdsd * CTSN[c2]
                            pdsd = 0
                            for c3 in range(ICLK[4], c):
                                APs3 = APTType(4, RTType[c3])
                                p8 = RpN[c3]
                                opdsd = abs(ADS3.index(date(APs3[p8][0])) - ADS3.index(date(APs3[p8][1])))
                                pdsd += opdsd * CTSN[c3]
                            scdsd = qdsd + tdsd + pdsd
                            SCDSD.append(scdsd)
                        BSP = ADS[SCDSD.index(min(SCDSD))]
                        CDSD = []
                        for d3 in range(weekly):
                            for p9 in range(IPD[BSP[d3]], FPD[BSP[d3]]):
                                CDSD.append(MAPF[p9])
                        dfScTd = pd.DataFrame(ScT, columns=list(map(str, CDSD)))
                        ScTD = dfScTd.values.tolist()
                        CpTs.append(CpT)
                        Avgtdw = 0
                        for s1 in range(s):
                            sctd = ScTD[s1]
                            for t1 in range(wpn):
                                c1 = sctd[t1]
                                c2 = sctd[t1 + 1]
                                l1 = PIL(ICL, c1)
                                l2 = PIL(ICL, c2)
                                Avgtdw += (DI[l1][l2]) / s
                            for i in range(len(UDP)):
                                c3 = sctd[UDP[i]]
                                c4 = sctd[UDP[i] + 1]
                                l3 = PIL(ICL, c3)
                                l4 = PIL(ICL, c4)
                                Avgtdw -= (DI[l3][l4]) / s
                        FScT = np.array(ScTD).reshape(-1).tolist()
                        RFScT.append(FScT)
                        FScT.append(Avgtdw)
                        FScTs.append(FScT)
                        FScTs.sort(key=lambda x: x[-1])
                        sortFScTs = np.delete(np.array(FScTs), -1, axis=1)
                        RAvgtdw.append(Avgtdw)
                        RSC.append(SC)
                while back == 1:
                    if Found == 1:
                        print("Found One")
                    else:
                        printone = '\r' + 'class:' + str(cN) + ',' + " ".join(map(str, RpN))
                        print(printone, end='')
                    cN -= 1
                    del RpL[-1]
                    came = 1
                    pN = RpN[-1] + 1
                    del RpN[-1]
                    del RGSTP[-1]
                    del RTType[-1]
                    del RCLP[-1]
                    """print(RpN)"""
                    back = 0
                    if pN == RpL[-1]:
                        if cN == 0:
                            control = 1
                            cN = c
                            pN += 1
                        else:
                            back = 1
                if control == 0:
                    GSTP = RGSTP[-1][:]
                    CLP = RCLP[-1][:]
                    TType = RTType[-1][:]
                    gN = 0
                    APs = APTType(ki, TType)
                    lN = PIL(ICL, cN)
                    i = 0
                    list111 = CGS(cN, GSC)
                    if came == 0:
                        pN = 0
                    while i < len(list111):
                        gN = list111[i]
                        if pN == len(APs):
                            back = 1
                            i += 1
                        else:
                            Selt1 = SelT(ki, APs, pN)
                            if DalO(GSTP[gN], Prime, Selt1) == 0:
                                RWC.append(cN)
                                RWG.append(gN)
                                pN += 1
                                i = 0
                                if pN == RpL[-1]:
                                    back = 1
                                    i = len(list111) + 1
                            else:
                                i += 1
                        if i == len(list111):
                            if MEx(SelM(CLP, SelT(ki, APs, pN)), lN) == MCPL[lN]:
                                pN += 1
                                if pN == RpL[-1]:
                                    back = 1
                                    gN = g + 1
                            else:
                                CLP = MulS(CLP, Selt1, Prime[lN])
                                RCLP.append(CLP)
                                GSTP = MulS(GSTP, CGS(cN, GSC), ProS(Prime, Selt1))
                                RGSTP.append(GSTP)
                                RpN.append(pN)
                                TType = CTType(ki, APs[pN], TType)
                                RTType.append(TType)
                                cN += 1
                                if came == 1:
                                    came = 0
                                pN = 0
                                ki = PIL(ICLK, cN)
                                RpL.append(NPTType(ki, RTType[-1]))
                                printone = '\r' + 'class:' + str(cN) + ',' + " ".join(map(str, RpN))
                                print(printone, end='')
    if control == 1:
        end = 1
        i = 0
        while i < len(RWC):
            c4 = RWC[-i - 1]
            if (CGS(c4, GSC)[0] - 1) * (LCN[PIL(ICL, c4)] - 1) > 0:
                SC[RWG[-i - 1]][RWC[-i - 1]] = 0
                if not {RWC[-i - 1]} - set(FCL):
                    SC[RWG[-i - 1]][ICL[FCL.index(RWC[-i - 1])]] = 1
                else:
                    SC[RWG[-i - 1]][RWC[-i - 1] + 1] = 1
                end = 0
                i = len(RWC)
                control = 0
            else:
                if CGS(c4, GSC)[0] == LCN[PIL(ICL, c4)]:
                    onestudentlecture.append(PIL(ICL, c4))
                    oneclasslecture.append(PIL(ICL, c4))
                else:
                    if CGS(c4, GSC)[0] == 1:
                        onestudentlecture.append(PIL(ICL, c4))
                    else:
                        oneclasslecture.append(PIL(ICL, c4))
        if end == 1:
            if len(sortFScTs) == 0:
                print("This lectures haves too many classes and make error.", onestudentlecture)
                print("This lectures haves too few classes and make error.", oneclasslecture)
            else:
                for i in range(len(sortFScTs)):
                    workbookname = str(i + 1) + '번째 시간표'
                    ScT10 = np.array(sortFScTs[i]).reshape(s, wpn).tolist()
                    makingsheetdone(ScT10, workbookname)