import numpy as np
import pandas as pd
import math
import openpyxl
from numba import njit
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sympy import Derivative, symbols
import matplotlib.pyplot as plt

print("What is Data File's name")
DataFile_path = input()
dfData = pd.read_excel(DataFile_path)
Data = dfData.values
N = len(Data)
ND = len(Data.T)
m = np.mean(Data)
"""
1. numpy의 장점-1
stdn = np.sqrt(np.mean((Data - np.mean(Data))**2))
이것은 그냥 일반적인 2D 행렬의 데이터의 표준편차를 구하는 식인데 그냥 저렇게 해도 되는 것임
이러한 것 때문에 numpy를 쓰는 것 이기도 해. numpy는 array**2을 하면 각 원소를 제곱한 
결과가 나와. https://pybasall.tistory.com/126 참고하셈.
"""
print("가우시안 분포의 표준편차는 뭘로 할꺼야?")
stdn = float(input())


def dstsqr(a, b):
    return np.sum((a - b) ** 2) / (2 * stdn ** 2)


def dstsqr2(a, b):
    return np.sum(np.sum((a - b) ** 2, axis=1)) / (2 * stdn ** 2)


"""
일단은 그냥 a,b가 1D array인 경우로 짰는데 혹시 나중에 2d를 1d로 낮추고 쓸 계획이였을 
가능이 있으니까 2D array인 경우의 dstsqr를 그 코드 밑에 짠거야
"""


def remfac(a, b):
    return np.e ** (-dstsqr(a, b) / 2)


def remfac2(a, b):
    return np.e ** (-dstsqr2(a, b) / 2)


def remfac3(a):
    l = len(a)
    A = np.zeros((l, l))
    for i in range(l):
        for j in range(l):
            A[i][j] = remfac(a[i], a[j])
    return A

def remfac4(a,b):
    k=0
    p=0
    for i in range(len(b)):
        k+=dstsqr(a,b[i])*remfac(a,b[i])
        p+=remfac(a,b[i])
    return k/(p+10**-100)


"""
도대체 코드가 겹치면서 하나가 다른 하나를 포함하는데 그냥 앞에서 정의한 함수를 쓰지
왜 또 계산 코드를 짠거야. 최대한 코드 줄수를 줄일 수록 속도는 조금이라도 올라가. 근데
인공지능은 무한 루틴이니까 조금이라도 줄이면 결론적으로 많이 줄으니까 무조건 코드 줄수른 줄이는
것이 중요함. 그리고 쓸데 없이 중간에라도 변수를 정의하지 않은 것이 좋아. 그냥 바로 return을 
하는 것이 가장 최상의 함수 코드야.
"""

print("반복 횟수는 얼마로 할꺼야?")
rep = int(input())
print("시간 간격은 얼마로?")
tin = float(input())

colorset = ['b','g','r','c','m','y','k','w']
color = []
for i in range(N):
    color.append(colorset[i%len(colorset)])

changeData = np.copy(Data)
"""
2. indexing
코드를 보니까 인덱싱을 할려고 하는 계획처럼 보이던데 만약 Data가 list형식이라면 그냥 
changeData = Data[:]
를 하면 되는 것이고 만약 Data가 array 형식이라면 그냥 
changeData = np.copy(Data)
하면 되는 것임.ㅋㅋㅋㅋㅋ
"""

for n in range(rep):
    eigMatrix = np.linalg.eig(remfac3(changeData))
    eig = np.copy(eigMatrix[0])
    """
    3. 일회성 함수 
    일단은 이번 한번만 정의할꺼면 저렇게 함수안에 함수를 넣은 것으로 저장공간을 줄여.
    저장공간의 감소도 연산속도를 늘리니까. 
    4. numpy의 장점-2 & pandas의 장점-1 
    list의 내가 아는 한 모든 기능을 numpy에서 할 수 있고 그냥 쓰는 방법만 다를 뿐이지.
    오히러 더욱 많은 기능이랑 편리성을 가지고 있어서 특별한 경우 아니고선 list보다는 
    numpy를 활용하는 것을 추천할께. 데이터를 처리, 가공하는 것은 pandas가 매우 좋고
    행렬 계산을 할 때는 numpy를 써.
    """
    Nmg1 = np.copy(eigMatrix[1])
    Nmg = np.zeros((N, N))
    for i in range(N):
        Nmg[:, i] = Nmg1[i]/((eig[i] * np.sum(Nmg1[i] ** 2)) ** 0.5)
    """
    5. transpose
    그냥 0행렬 array를 만들고 거기의 i열을 바꿔가면서 하는 것이 가장 간단할 것 같아서 
    transpose가 아니라 그냥 열을 바꾸게 해놨어. 근데 append를 써서 할 수도 있는데
    array의 append는 axis때문에 좀 헷갈리고 초기에 시작하는 것이 좀 난 까다로워서
    그러한 것은 나중에 익숙해지면 하는 것이 좋을꺼야.
    6. tolist하고 난 직후 다시 np.array를 썼더라 ㅋㅋㅋㅋ 어차피 array가 list의 모든
    기능을 가지니까 list는 쓰지 않는 것이 좋음.ㅋㅋㅋ
    """
    Nmgrev = np.linalg.inv(Nmg)
    Nmgh = np.copy(Nmg.T)
    Nmghrev = np.copy(Nmgrev.T)
    Hm = np.zeros((N,N))
    for i in range(N):
        for j in range(N):
            mid = (changeData[i]+changeData[j])/2
            Hm[i][j]=-dstsqr(changeData[i],changeData[j])/4*remfac(changeData[i],changeData[j])+remfac4((changeData[i]+changeData[j])/2,changeData)
    Hpm = np.copy((Nmgh.dot(Hm)).dot(Nmg))
    Xm = np.zeros((N,N,ND))
    for i in range(N):
        for j in range(N):
            A = changeData[i]+changeData[j]
            Xm[i][j] = A/2*remfac(changeData[i],changeData[j])
    Xpmim = np.copy(Nmgh.dot(Xm))
    Xpm = np.copy(Xpmim.T @ Nmg)
    Heigmatrix = np.linalg.eig(Hpm)
    Heig = np.copy(Heigmatrix[0])
    Hmgh = np.copy(Heigmatrix[1])
    Hmg = np.copy(Hmgh.T)
    Hmgrev = np.linalg.inv(Hmg)
    Xppmim = np.copy(Hmgh.dot(Xpm))
    Xppm = np.copy(Xppmim @ Hmg)
    HNm = np.copy(Hmgrev.dot(Nmgrev))
    HNmh = np.copy(HNm.T)
    Xmcos = np.zeros((N,N))
    for i in range(N):
        for j in range(N):
            Xmcos[i][j] = Xppm[i][j]*np.cos(tin*(Heig[i]-Heig[j]))
    HXpmim = np.copy(HNmh.dot(Xmcos))
    Xupdate = np.copy(Xpmim @ HNm)
    changeData = np.diag(Xupdate)
    if n%10==0:
        for i in range(N):
            plt.scatter(changeData[:,0],changeData[:,1],s=1,c=color)
        plt.pause(0.001)
plt.show()