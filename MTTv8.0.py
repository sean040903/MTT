import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math

# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
"""print('학생들의 이름이 담긴 길이 s짜리 xlsx파일의 이름이 뭔가요?')"""
StudentNamesFile_path = 'StudentNames.xlsx'
dfStudentNames = pd.read_excel(StudentNamesFile_path)
StudentNames = dfStudentNames.values[:, 1]
s = len(StudentNames)

"""print('과목들의 이름이 담긴 길이 l짜리 xlsx파일의 이름이 뭔가요?')"""
LectureNamesFile_path = 'LectureNames.xlsx'
dfLectureNames = pd.read_excel(LectureNamesFile_path)
LectureNames = dfLectureNames.values[0]
l = len(LectureNames)

"""print('학생들의 수강 유무가 있는 s x l짜리 xlsx파일의 이름이 뭔가요?')"""
StudentLectureFile_path = 'SL.xlsx'
dfStudentLecture = pd.read_excel(StudentLectureFile_path)
SL = dfStudentLecture.values

"""print('각 과목별로 시수,종류가 있는 l x 2짜리 xlsx파일의 이름이 뭔가요?')"""
KLFile_path = 'KindLecture.xlsx'
dfKindLecture = pd.read_excel(KLFile_path)
KL = dfKindLecture.values

"""print('각 과목별로 한 교시에 가능한 최대 분반의 개수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
MCPLFile_path = 'MCPL.xlsx'
dfMCPL = pd.read_excel(MCPLFile_path)
MCPL = dfMCPL.values[0]

"""print('각 과목별 분반의 수가 담긴 l짜리 xlsx파일의 이름이 뭔가요?')"""
LCNFile_path = 'LCN.xlsx'
dfLCN = pd.read_excel(LCNFile_path)
LCN = dfLCN.values[0]

"""print('1주일동안 수업이 진행되는 날들의 교시수들이 담긴 xlsx파일의 이름은 뭔가요?')"""
DPNFile_path = 'DailyPN.xlsx'
dfDailyPN = pd.read_excel(DPNFile_path)
DPN = dfDailyPN.values[0]

"""print('과목별로 진행되는 강의실의 기하학적 중심에서 다른 과목의 강의실로 가는데 이동해야 하는 실거리가 담긴 l x l짜리 xlsx파일의 이름이 뭔가요?')"""
DistanceFile_path = 'DI.xlsx'
dfDistance = pd.read_excel(DistanceFile_path)
DI = dfDistance.values

"""print('각 학생별로 모든 과목들을 자신에게 중요하다고 생각하는 것들의 순위(0부터시작하는 순위)가 담긴 s x l짜기 xlsx파일의 이름이 뭔가요?')"""
StudentimpFile_path = 'Studentimp.xlsx'
dfStudentimp = pd.read_excel(StudentimpFile_path)
Simp = dfStudentimp.values

"""print('매 요일의 이름이 담긴 xlsx파일의 이름은 뭔가요?')"""
datenamesFile_path = 'DateNames.xlsx'
dfdatenames = pd.read_excel(datenamesFile_path)
datenames = dfdatenames.values[0]

"""print('소수 500개 담긴 xlsx파일의 이름은 뭔가요?')"""
PrimeFile_path = 'PrimeList.xlsx'
dfprimes = pd.read_excel(PrimeFile_path)
Prime = dfprimes.values[0]


# 추가 함수
def SelT(k, arr1, n):
    if k == 4:
        return np.array([arr1[n]])
    elif k == 3:
        return np.array([arr1[n], arr1[n] + 1])
    elif k == 2:
        return np.array([arr1[n][0], arr1[n][1]])
    elif k == 1:
        return np.array([arr1[n][0], arr1[n][0] + 1, arr1[n][1]])
    else:
        return np.array([arr1[n][0], arr1[n][0] + 1, arr1[n][1], arr1[n][1] + 1])


def ProS(arr1, arr2):
    list1000 = arr1[arr2]
    setint32 = np.array([1],float)
    produ = setint32[0]
    for i in list1000:
        produ *= i
        print(produ)
    production = np.prod(arr1[arr2])

    return np.prod(arr1[arr2])


def DalO(n, arr1, arr2):
    return np.sum(n / arr1[arr2] == n // arr1[arr2]) > 0


def MulS(arr1, arr2, n):
    np.put(arr1, arr2, arr1[arr2] * n)


def APTType(k, arr0):
    if k == 4:
        return Even(arr0)
    elif k == 3:
        return Odd(arr0)
    else:
        ap = np.empty((0, 2), int)
        AP1 = APTType(4 - k // 2, arr0)
        for p1 in AP1:
            arr1 = CTType(k // 2, p1, arr0)
            AP2 = APTType(4 - (k + 1) // 2, arr1)
            for p2 in AP2:
                if date(p1) < date(p2):
                    ap = np.append(ap, [[p1, p2]], axis=0)
                elif k == 3 and date(p1) > date(p2):
                    ap = np.append(ap, [[p1, p2]], axis=0)
        return ap


def NPTType(k, arr):
    return len(APTType(k, arr))


def MEx(arr, n):
    mex1 = 0
    for i in arr:
        mex1 = max(mex1, Pm(i, n))
    return mex1


def SelM(arr1, arr2):
    return arr1[arr2]


def CSE(arr1, arr2, m, n):
    np.put(arr1, len(arr1.T) * m + arr2, n)


# 행렬-원소 선택 함수들#
def PIL(arr1, ip):
    arr2 = np.array(ip, int).reshape(-1)
    pil1 = np.array([], int)
    for i in arr2:
        pil1 = np.append(pil1, np.sum(arr1 <= i)-1)
    return pil1


# DPN 활용#
daily = np.max(DPN)
weekly = len(DPN)
IPD = np.array([0], int)
FPD = np.array([], int)
for i3 in range(weekly - 1):
    IPD = np.append(IPD, IPD[-1] + DPN[i3])
    FPD = np.append(FPD, IPD[-1] - 1)
wpn = np.sum(DPN)
FPD = np.append(FPD, wpn)


def date(ip):
    return PIL(IPD, np.array(ip).reshape(-1))


def CTType(k, SelP, arr):
    selp = np.array(SelP).reshape(-1)
    arr1 = np.copy(arr)
    for i in range(weekly):
        if arr[i] < DMEP[i]:
            for j in selp:
                if j == EP[PIL(EP, arr[i])]:
                    arr1[i] += 2
    dates = date(SelP)
    for i in dates:
        if i in DCD:
            ind = np.where(DCD == i)[0]
            arr1[weekly + ind] = 1
    if k == 1:
        if selp[1] in EP or selp[1] in VEP:
            arr1[-1] = BFN(np.union1d(NZBF(arr[-1]), len(EP) - PIL(EP, selp[1])))
    elif k == 2 or k == 4:
        for i in selp:
            if i in EP or i in VEP:
                arr1[-1] = BFN(np.union1d(NZBF(arr[-1]), len(EP) - PIL(EP, i)))
    if k == 0 or k == 2:
        if dates in SDPD:
            i0 = np.where(DCD == dates[0])[0] + weekly
            if arr[i0] == 0:
                if len(selp)>1 and SelP[0] + np.sum(DPN[dates[0]:dates[1]]) == SelP[1]:
                    arr1[i0] = 0
    return arr1


# making classes#
ICL = np.array([0], int)
FCL = np.array([], int)
for i6 in range(l):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL = np.append(ICL, a1 + a2)
    FCL = np.append(FCL, a1 + a2 - 1)
LTS = np.sum(SL, axis=1)
S = np.arange(s)
SL = SL[LTS.argsort()]
StS = S[S[LTS.argsort()].argsort()]

c = np.sum(LCN)
SC = np.zeros((s, c), int)
cnow = 0
for l1 in range(l):
    SinL = np.nonzero(SL[:, l1])[0]
    SinLC = np.array_split(SinL, LCN[l1])
    for i in range(LCN[l1]):
        SinC = SinLC[i]
        for j in range(len(SinC)):
            SC[SinC[j]][cnow] = 1
        cnow += 1

KC = np.repeat(KL, LCN, axis=0)


def CGS(i, arr):
    return np.nonzero(arr[:, i])[0]


# 반 분류 ICLK 생성#
ICLK = np.array([], int)
for i in range(np.max(KC[:, 0])):
    ICLK = np.append(ICLK, np.sum(KC[:, 0] < i))


# 소수,진법 변환 관련 함수들#
def Pm(n1, n2):
    m = 0
    while n1 % Prime[n2] == 0:
        m += 1
        n1 = n1 // Prime[n2]
    return m


def NZBF(n):
    nzbf = np.array([], int)
    nbf = format(n, 'b')
    for i in range(len(nbf)):
        if nbf[i] == '1':
            nzbf = np.append(nzbf, i)
    return nzbf


def BFN(arr):
    n = 0
    for i in arr:
        n += 2 ** i
    return int(n)


# 행렬 연속 몇개 선택 함수들#
def C2N(x, y):
    c2n = np.array([], int)
    for i in range(x, y, 2):
        c2n = np.append(c2n, i)
    return c2n


# 교시교환 관련 함수들#
def LW1P(da, arr):
    a = IPD[da]
    tw = np.zeros((2, 2), int)
    for i in range(2):
        for j in range(2):
            for s1 in range(s):
                c1 = arr[s1][a + 1 - i]
                c2 = arr[s1][a + 2 + j]
                if max(c1, c2) < c:
                    l1 = PIL(ICL, c1)
                    l2 = PIL(ICL, c2)
                    tw[i][j] += DI[l1][l2]
    x1, y1 = np.where(tw == np.min(tw))
    x = x1[0]
    y = y1[0]
    mwp = np.array([a + x, a + 1 - x, a + 2 + y, a + 3 - y])
    if DPN[da] % 2 == 1:
        mwp = np.append(mwp, a + 4)
    return mwp


def LW2P(da, arr):
    a = IPD[da]
    tw = np.zeros((3, 2, 2), int)
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for s1 in range(s):
                    c1 = arr[s1][a + 2 * (i % 3) + 1 - i]
                    c2 = arr[s1][a + 2 * ((i + 1) % 3) + j]
                    if max(c1, c2) < c:
                        l1 = PIL(ICL, c1)
                        l2 = PIL(ICL, c2)
                        tw[i][j] += DI[l1][l2]
    x4, y4, z4 = np.where(tw == np.min(tw))
    x = x4[0]
    y = y4[0]
    z = z4[0]
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1])


def LW3P(da, arr):
    a = IPD[da]
    tw = np.zeros((3, 2, 2, 2), int)
    for i in range(3):
        for j in range(2):
            for k in range(2):
                for o in range(2):
                    for s1 in range(s):
                        c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
                        c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
                        c3 = arr[s1][a + 2 * ((i + 2) % 3) + 1 - o]
                        c4 = arr[s1][a + 6]
                        if max(c1, c2) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
                        if max(c3, c4) < c:
                            l1 = PIL(ICL, c1)
                            l2 = PIL(ICL, c2)
                            tw[i][j] += DI[l1][l2]
    x4, y4, z4, w4 = np.where(tw == np.min(tw))
    x = x4[0]
    y = y4[0]
    z = z4[0]
    w = w4[0]
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6])


def LWP(arr):
    tmwp = np.array([], int)
    for i in range(weekly):
        if DPN[i] < 6:
            tmwp = np.append(tmwp, LW1P(i, arr))
        elif DPN[i] == 6:
            tmwp = np.append(tmwp, LW2P(i, arr))
        else:
            tmwp = np.append(tmwp, LW3P(i, arr))
    return tmwp


# EP,VEP관련 함수들#
def Even(arr):
    eap = np.array([], int)
    for i in range(weekly):
        if i not in CAD or arr[SDPI[np.where(CAD == i)[0]]] == 1:
            eap = np.append(eap, C2N(IPD[i], arr[i] + 2))
    return np.sort(eap)


def Odd(arr):
    oap = np.array([], int)
    for i in range(weekly):
        n1 = arr[SDPI[np.where(CAD == i)[0]]]
        n2 = NEP[np.where(NEPI == i)[0]]
        if i in CAD:
            if n1 == 1:
                oap = np.append(oap, C2N(IPD[i], arr[i] + 2))
        else:
            if IPD[i] == arr[i]:
                oap = np.append(oap, arr[i])
            else:
                oap = np.append(oap, C2N(IPD[i], arr[i] + 2))
        if i not in CAD or n1 == 1 and DPN[i] % 2 == 1:
            oap = np.append(oap, n2)
    for i in NZBF(arr[-1]):
        oap = np.append(oap, VEP[i])
    return np.sort(oap)


# 그래프 색깔#
"""SPC = np.zeros(12)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
Math = []
Phy = []
Che = []
CoSc = []
LiSc = []
EaSc = []
ConS = []
ScEx = []
CCSL = []
Lan = []
Soc = []
MAS = []


def CSP(n):
    if n == 0:
        return Math
    elif n == 1:
        return Phy
    elif n == 2:
        return Che
    elif n == 3:
        return CoSc
    elif n == 4:
        return LiSc
    elif n == 5:
        return EaSc
    elif n == 6:
        return ConS
    elif n == 7:
        return ScEx
    elif n == 8:
        return CCSL
    elif n == 9:
        return Lan
    elif n == 10:
        return Soc
    else:
        return MAS


for c1 in range(c):
    partn = KC[c1][1]
    CSP(partn).append(c1)
    SPC[partn] += 1


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = []
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1.append(i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
        else:
            for G in range(k - R, -1):
                Blue = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[Blue], 'x'))
    return RGB
"""
PTTcolorset = ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff',
               'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']

# TType 설정 코드#
DP = np.setdiff1d(np.arange(wpn), np.union1d(IPD, IPD + 4))
DEPN = DPN // 2
EP = np.array([], int)
for i in range(weekly):
    EP = np.append(EP, IPD[i])
    for j in range(DEPN[i] - 1):
        EP = np.append(EP, EP[-1] + 2)
VEP = EP + 1
NEP = np.setdiff1d(np.arange(wpn), np.union1d(EP, VEP))
NEPI = date(NEP)
DMEP = IPD + 2 * DEPN - 2
SUDPI = np.array([], int)
SDPI = np.array([], int)
SUDPD = np.array([], int)
SDPD = np.empty((0, 2), int)
DCD = np.array([], int)
CAD = np.array([], int)
fTType = np.copy(IPD)
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            if DPN[i] % 2 == 1:
                SUDPI = np.append(SUDPI, len(fTType))
                SUDPD = np.append(SUDPD, j)
            SDPI = np.append(SDPI, len(fTType))
            fTType = np.append(fTType, 0)
            SDPD = np.append(SDPD, [[i, j]], axis=0)
            DCD = np.append(DCD, i)
            CAD = np.append(CAD, j)
            fi = 0
fTType = np.append(fTType, 0)
ADS = np.arange(weekly).reshape(1, -1)
nds = 1
done1 = 1
i1111 = 0
l1111 = 0
while done1 == 1:
    j = 0
    re = 1
    while re * (nds - i1111) > 0:
        ADS1 = ADS[i1111]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = np.copy(ADS1)
            a1 = ADS1[k1]
            a2 = ADS1[k0]
            ADS2[k0] = a1
            ADS2[k1] = a2
            i1 = 0
            l1111 = len(ADS)
            while i1 < l1111:
                ads2 = ADS[i1]
                if np.array_equiv(ads2, ADS2):
                    j += 1
                    i1 = len(ADS)
                else:
                    i1 += 1
                    if i1 == len(ADS):
                        ADS = np.append(ADS, ADS2.reshape(1,-1), axis=0)
                        nds += 1
                        re = 0
        if re == 1:
            i1111 += 1
        elif i1111 < l1111:
            i1111 += 1
            j = 0
    if re == 1:
        done1 = 0


# 엑셀 조작 관련#
def style(workbook, n, i, j, arr0, arr1):
    newsheet = workbook[sheetname(n, arr1)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif arr0[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(arr0.reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == arr0[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[0],
                                                               fill_type='solid')


def sheetname(b, arr):
    return str(arr[np.where(arr == b)[0]]) + '번째 학생'


def makingsheetdone(arr, wbname, arr1):
    wb = openpyxl.Workbook()
    for i in range(len(arr)):
        wb.create_sheet(sheetname(i, arr1))
    list3 = []
    for s3 in range(len(arr)):
        list0 = arr[s3].tolist()
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
    list4 = list3[:]
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(arr)):
            for i in range(weekly):
                for j in range(daily):
                    c5 = list3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, c5)
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LectureNames[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0, arr1))
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(arr)):
        i = 0
        s2 = wb[sheetname(s1, arr1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list4[s1], arr1)
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


def printmtt(n1, n2, arr1):
    printone = 'Found:' + str(n1) + ' ' * (
            170 - math.floor(np.log10(max(1, n1))) - 1) + 'class:' + str(n2) + ' ' * (
                       172 - math.floor(
                   np.log10(max(1, n2))) - 1) + '각 분반별 배정된 시간:' + " ".join(map(str, arr1))
    print(printone, end='\r')


# highest priority control codes#
control = 0
back = 0
came = 0
Found = 0
sdca = 1
# highest priority control codes#


# main variables define code#
Rg = np.array([], int)
RGSCP = np.array([], int)
IGSCP = np.array([], int)
RWG = np.array([], int)
RWC = np.array([], int)
osl = np.array([], int)
ocl = np.array([], int)
while control == 0:
    RpL = np.array([], int)
    RpN = np.array([], int)
    pN = 0
    cN = 0
    gN = 0
    CC = np.zeros((c, c), int)
    for s1 in range(s):
        SS1 = SC[s1]
        CC += np.outer(SS1, SS1, out=None)
    for i in range(c):
        for j in range(c):
            if CC[i][j] == 0:
                CC[i][j] = s
    g = 0
    GS = np.array([], int)
    GSC = np.empty((0, c), int)
    GSCP = np.array([], int)
    for s2 in range(s):
        SS2 = SC[s2]
        if np.min(CC - np.outer(SS2, SS2, out=None)) > 0:
            CC -= np.outer(SS2, SS2, out=None)
        else:
            g += 1
            GS = np.append(GS, s2)
            GSC = np.append(GSC, [SC[s2]], axis=0)
            selection = Prime[np.nonzero(SC[s2])[0]]
            prod1 = ProS(Prime,np.nonzero(SC[s2])[0])
            i1000 = prod1%Prime[np.nonzero(SC[s2])[0][0]]
            GSCP = np.append(GSCP, ProS(Prime,np.nonzero(SC[s2])[0]))
    if g in Rg:
        i = 0
        while i < len(IGSCP) - 1:
            PGSCP = RGSCP[IGSCP[i]:IGSCP[i + 1]]
            if np.array_equiv(np.sort(PGSCP), np.sort(GSCP)):
                cN = c
                i = len(IGSCP) - 1
            else:
                i += 1
    else:
        Rg = np.append(Rg, g)
        RGSCP = np.append(RGSCP, GSCP)
        IGSCP = np.append(IGSCP, len(RGSCP))
    if cN == 0:
        RGSTP = np.ones(g, int).reshape((1, -1))
        RTType = np.copy(fTType).reshape((1, -1))
        RCLP = np.ones(g, int).reshape((1, -1))
        RWC = np.array([], int)
        RWG = np.array([], int)
        while cN < c:
            ki = PIL(ICLK, cN)
            RpL = np.append(RpL, NPTType(ki, RTType[-1]))
            while pN < RpL[-1] + 1:
                if cN == c:
                    ScT = np.ones((s, wpn), int) * c
                    SaT = np.ones((s, wpn), int) * 12
                    SiT = np.ones((s, wpn), int) * l
                    for s1 in range(s):
                        for c1 in range(c):
                            TType = RTType[c1]
                            ki = PIL(ICLK, c1)
                            Selt2 = SelT(ki, APTType(ki, TType), RpN[c1])
                            p1 = Prime[c1]
                            l1 = PIL(ICL, c1)
                            if SC[s1][c1] == 1:
                                a1 = KC[c1][1]
                                im1 = Simp[s1][l1]
                                CSE(SaT, Selt2, s1, a1)
                                CSE(ScT, Selt2, s1, c1)
                                CSE(SiT, Selt2, s1, im1)
                    s2 = 0
                    while s2 < s * sdca:
                        d1 = 0
                        while d1 < weekly * sdca:
                            t1 = 1
                            while t1 < DPN[d1] * sdca:
                                t2 = 0
                                while 0 < (t1 - t2) * sdca:
                                    p2 = IPD[d1] + t1
                                    p3 = IPD[d1] + t2
                                    if SaT[s2][p2] == SaT[s2][p3] and ScT[s2][p2] != ScT[s2][p3]:
                                        sdca = 0
                                    t2 += 1
                                t1 += 1
                            d1 += 1
                        s2 += 1
                    if sdca == 1:
                        MTDW = LWP(ScT)
                        MAPF = np.copy(MTDW)
                        for d2 in range(weekly):
                            if DPN[d2] > 3:
                                p4 = IPD[d2]
                                Im12 = 0
                                Im34 = 0
                                Selt3 = SelT(0, MAPF, p4)
                                Selt4 = SelT(0, MAPF, p4 + 2)
                                for s3 in range(s):
                                    Im12 += np.sum(SelM(SiT[s3], Selt3))
                                    Im34 += np.sum(SelM(SiT[s3], Selt4))
                                Ims = np.array(Im12, Im34)
                                m1 = np.argmin(Ims)
                                MAPF[p4] = MTDW[p4 + 3 - 3 * m1]
                                MAPF[p4 + 1] = MTDW[p4 + 2 - m1]
                                MAPF[p4 + 2] = MTDW[p4 + 1 + m1]
                                MAPF[p4 + 3] = MTDW[p4 + 3 * m1]
                                if DPN[d2] == 6:
                                    p5 = p4 + 4
                                    Im5 = 0
                                    Im6 = 0
                                    for s1 in range(s):
                                        Im5 += SiT[s1][MAPF[p5]]
                                        Im6 += SiT[s1][MAPF[p5 + 1]]
                                    if Im6 > Im5:
                                        MAPF[p5] = MTDW[p5 + 1]
                                        MAPF[p5 + 1] = MTDW[p5]
                        CTSN = np.sum(SC, axis=0)
                        SCDSD = np.array([], int)
                        for i in range(nds):
                            ADS3 = ADS[i]
                            qdsd = 0
                            for c1 in range(ICLK[3]):
                                APs1 = APTType(PIL(ICLK, c1), RTType[c1])[RpN[c1]]
                                qdsd += abs(np.where(ADS3 == date(APs1[0]))[0] - np.where(ADS3 == date(APs1[1]))) * CTSN[c1]
                            SCDSD = np.append(SCDSD, qdsd)
                        BSP = ADS[np.argmin(SCDSD)]
                        CDSD = np.array([], int)
                        for d3 in range(weekly):
                            for p9 in range(IPD[BSP[d3]], FPD[BSP[d3]] + 1):
                                CDSD = np.append(CDSD, MAPF[p9])
                        ScTD = ScT[:, CDSD]
                        tdw = 0
                        for s1 in range(s):
                            sctd = ScTD[s1]
                            for t1 in DP:
                                l1 = PIL(ICL, sctd[t1 - 1])
                                l2 = PIL(ICL, sctd[t1])
                                tdw += DI[l1][l2]
                        workbookname = str(tdw) + '가 총 이동거리인 시간표'
                        makingsheetdone(ScTD, workbookname, StS)
                while back == 1:
                    if len(RpL) == 1:
                        control = 1
                        back = 0
                        pN = RpL[-1]+1
                        cN = c
                    else:
                        cN -= 1
                        printmtt(Found, cN, RpN)
                        RpL = np.delete(RpL, -1)
                        came = 1
                        pN = RpN[-1] + 1
                        RpN = np.delete(RpN, -1)
                        RGSTP = np.delete(RGSTP, -1)
                        RTType = np.delete(RTType, -1)
                        RCLP = np.delete(RCLP, -1)
                        back = 0
                        if pN == RpL[-1]:
                            if cN == 0:
                                control = 1
                                cN = c
                                pN += 1
                            else:
                                back = 1
                if control == 0:
                    GSTP = np.copy(RGSTP[-1])
                    CLP = np.copy(RCLP[-1])
                    TType = np.copy(RTType[-1])
                    gN = 0
                    APs = APTType(ki, TType)
                    lN = PIL(ICL, cN)
                    i = 0
                    arr111 = CGS(cN, GSC)
                    if came == 0:
                        pN = 0
                    while i < len(arr111):
                        gN = arr111[i]
                        Selt1 = SelT(ki, APs, pN)
                        if pN == len(APs):
                            back = 1
                            i += 1
                        else:
                            if DalO(GSTP[gN], Prime, Selt1) == 0:
                                RWC = np.append(RWC, cN)
                                RWG = np.append(RWG, gN)
                                pN += 1
                                i = 0
                                if pN == RpL[-1]:
                                    back = 1
                                    i = len(arr111) + 1
                            else:
                                i += 1
                        if i == len(arr111):
                            if MEx(SelM(CLP, SelT(ki, APs, pN)), lN) == MCPL[lN]:
                                pN += 1
                                if pN == RpL[-1]:
                                    back = 1
                                    gN = g + 1
                            else:
                                MulS(CLP, Selt1, Prime[lN])
                                RCLP = np.append(RCLP, CLP)
                                MulS(GSTP, CGS(cN, GSC), ProS(Prime, Selt1))
                                RGSTP = np.append(RGSTP, GSTP)
                                RpN = np.append(RpN, pN)
                                TType = CTType(ki, APs[pN], TType)
                                RTType = np.append(RTType, TType)
                                cN += 1
                                came = 0
                                pN = 0
                                ki = PIL(ICLK, cN)
                                RpL = np.append(RpL, NPTType(ki, RTType[-1]))
                                printmtt(Found, cN, RpN)
    if control == 1:
        end = 1
        i = 0
        while i < len(RWC):
            c4 = RWC[-i - 1]
            if (np.sum(SC[:, c4]) - 1) * (LCN[PIL(ICL, c4)] - 1) > 0:
                SC[RWG[-i - 1]][RWC[-i - 1]] = 0
                if RWC[-i - 1] not in FCL:
                    SC[RWG[-i - 1]][ICL[np.where(FCL == RWC[-i - 1])[0]]] = 1
                else:
                    SC[RWG[-i - 1]][RWC[-i - 1] + 1] = 1
                end = 0
                i = len(RWC)
                control = 0
            else:
                if np.sum(SC[:, c4]) == 1:
                    osl = np.append(osl, PIL(ICL, c4))
                if LCN[PIL(ICL, c4)] == 1:
                    ocl = np.append(ocl, PIL(ICL, c4))
        if end == 1:
            if Found == 0:
                print("The lectures that has many classes : ", osl)
                print("The lectures that has less classes : ", ocl)
