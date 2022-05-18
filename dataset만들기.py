import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

s=263

datalist=[]
for i in range(s):



def sheetname(s2):
    return list(map(str, [s2]))[0]

wb = openpyxl.Workbook()
for i in range(s):
    wb.create_sheet(sheetname(i))

with pd.ExcelWriter('test.xlsx') as writer:
    for s0 in range(s):
        arr1 = np.array(datalist[s0])
        arr2 = arr1.T
        datalist2 = arr2.tolist()
        df = pd.DataFrame(datalist2, columns=['월요일', '화요일', '수요일', '목요일'])
        df.to_excel(writer, sheet_name=sheetname(s0))

excel_filename = 'test.xlsx'
wb = load_workbook(filename=excel_filename)

for s0 in range(s):
    s1 = wb[sheetname(s0)]
    for row in range(1, len(df.index) + 1):
        s1.row_dimensions[row + 1].height = 40
    for col in range(1, len(df.columns) + 1):
        s1.column_dimensions[get_column_letter(col + 1)].width = 15
    s1.row_dimensions[1].height = 18
    s1.column_dimensions[get_column_letter(1)].width = 5
os.remove('test.xlsx')
wb.save('파이썬으로 만든 시간표(최종버전).xlsx')