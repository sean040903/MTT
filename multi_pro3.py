import time
import numpy as np
import ray

ray.init(num_cpus=4)


@ray.remote
def plusing(a1,ind1):
    a2 = np.copy(a1)
    for i in ind1:
        a2 = np.append(a2,[np.ones(2,int)*i])
    return a2

@ray.remote
def multiplying(a1, indx, indy, num):
    a2 = np.copy(a1)
    for i in indx+1:
        a2[i][indy+1] *= num
    return a2


start = time.time()
xindex = ray.put(np.arange(0,200,2))
yindex = ray.put(np.arange(0,200,4))
a_id = ray.put(np.ones((200,200), int))
b_id = ray.put(np.ones((200,200), int))
c_id = ray.get(multiplying.remote(a_id,xindex,yindex,5))
print(np.append(np.ones((200,200), int),c_id,axis = 0))
array1 = np.zeros(10,int)
array1 += 1
print(array1)
array1[0] = c_id[0,0]
print(array1)
c_id = np.copy(c_id)
c_id[0] = c_id[1]
print(c_id)
print("duration with ray.get=", time.time() - start)



"""


import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook, chart
from openpyxl.utils import get_column_letter
import math
import matplotlib.pyplot as plt
from collections import Counter
from tqdm import tqdm, trange
from numba import jit, int32

onegrade = np.array([0, 1, 7, 8, 11, 12, 13, 34, 57, 59, 61])
# 엑셀에 담긴 데이터를 가져오는 코드를 사용하기 위해 엑셀 데이터파일의 이름을 받고 월드입력변수를 받는 코드#
StudentNames = pd.read_excel('StudentNames.xlsx').values[:, 1][:8]
s = len(StudentNames)

LN = pd.read_excel('LectureNames.xlsx').values[0][onegrade]
l = len(LN)

SL = pd.read_excel('SL.xlsx').values[:8][:, onegrade]

KL = pd.read_excel('KindLecture.xlsx').values[0][onegrade]

KA = pd.read_excel('AreaLecture.xlsx').values[0][onegrade]

MLP = pd.read_excel('MCPL.xlsx').values[0][onegrade]

LCN = pd.read_excel('LCN.xlsx').values[0][onegrade]
c = np.sum(LCN)
C = np.arange(c)

DPN = pd.read_excel('DailyPN.xlsx').values[0]

DI = pd.read_excel('DI.xlsx').values[onegrade][:, onegrade]

Simp = pd.read_excel('Studentimp.xlsx').values[:8][:, onegrade]

datenames = pd.read_excel('DateNames.xlsx').values[0]
periodnames = np.array([], str)
for i in range(len(datenames)):
    for j in range(DPN[i]):
        periodnames = np.append(periodnames, datenames[i] + ' ' + str(j + 1) + '교시')


# 변경 함수
@ray.remote
def CAiP(AiP1, ttype4, selp1, LP1, cn7):  # TType을 변경하기 전에 진행을 해야 한다.
    AiP2 = np.copy(AiP1)
    ln7 = ray.get(PIL.remote(ICL,np.array([cn7])))[0]
    ttype5 = ray.get(CTType.remote(selp1,ttype4))
    Dd = np.flatnonzero((ttype5 - ttype4)[weekly:-1]) + weekly
    for k in Dd:
        j = CAD[SDPI == k][0]
        i = DCD[SDPI == k][0]
        while i in CAD and ttype4[SDPI[CAD == i]] == 0:
            i = DCD[CAD == i][0]
        AiP2[:, IPD[j]:FPD[j] + 1] = np.copy(AiP2[:, IPD[i]:FPD[i] + 1])
    Ck = ttype5[np.flatnonzero((ttype5 - ttype4)[:weekly])]
    for i in Ck:
        AiP2[NcN, i:i + 2] *= -1
    for d1 in date(selp1):
        AiP2[cn7, IPD[d1]:FPD[d1] + 1] = 0
    GinC = np.flatnonzero(GC[:, cn7])
    AiP2[np.intersect1d(np.flatnonzero(np.sum(GC[GinC], axis=0)), NcN)][:, selp1] = 0
    AiP2[np.intersect1d(np.flatnonzero(np.sum(GC[GinC], axis=0)), NcN)] *= (np.sum(GP[GinC], axis=0) == 0)
    AiP2[np.intersect1d(np.arange(ICL[ln7], FCL[ln7] + 1), NcN)] *= (LP1[ln7] < MLP[ln7])
    return AiP2


@ray.remote
def APTType(AiP3, ttype6, LP2, cn9, num=0):
    kc4 = KC[cn9]
    if num > 0:
        kc4 = 5 - num
    if kc4 == 4:
        apve = VEP[np.intersect1d(np.flatnonzero(AiP3[cn9, VEP] == 1), ray.get(NZBF.remote(ttype6[-1])))]
        apae = AEP[AiP3[cn9, AEP] == 1]
        return np.concatenate((apve, apae)).reshape(-1, 1)
    elif kc4 == 3:
        ape = EP[AiP3[cn9, EP] == 1].reshape(1, -1)
        return np.append(ape, ape + 1, axis=0).T
    elif kc4 == 2:
        ap2 = np.empty((0, 2), int)
        AP1 = ray.get(APTType.remote(AiP3, ttype6, LP2, cn9, num=1))
        for ap1 in AP1:
            AiP4 = ray.get(CAiP.remote(AiP3, ttype6, LP2, ap1, cn9))
            AP2 = ray.get(APTType.remote(AiP4, ray.get(CTType.remote(ap1, ttype6)), LP2, cn9, num=1))
            AP2 = AP2[ray.get(date.remote(AP2[:, 0])) > ray.get(date.remote(ap1))]
            for ap0 in AP2:
                ap2 = np.append(ap2, [np.append(ap1, ap0)], axis=0)
        return ap2
    elif kc4 == 1:
        ap3 = np.empty((0, 3), int)
        AP1 = ray.get(APTType.remote(AiP3, ttype6, LP2, cn9, num=2))
        for ap1 in AP1:
            AiP4 = ray.get(CAiP.remote(AiP3, ttype6, ap1, LP2, cn9))
            AP2 = ray.get(APTType.remote(AiP4, ray.get(CTType.remote(ap1, ttype6)), LP2, cn9, num=1))
            AP2 = AP2[ray.get(date.remote(AP2[:, 0])) != ray.get(date.remote(ap1))]
            for ap0 in AP2:
                ap3 = np.append(ap3, [np.append(ap1, ap0)], axis=0)
        return ap3
    else:
        ap4 = np.empty((0, 4), int)
        AP1 = ray.get(APTType.remote(AiP3, ttype6, LP2, cn9, num=2))
        for ap1 in AP1:
            AiP4 = ray.get(CAiP.remote(AiP3, ttype6, ap1, LP2, cn9))
            AP2 = ray.get(APTType.remote(AiP4, ray.get(CTType.remote(ap1, ttype6)), LP2, cn9, num=2))
            AP2 = AP2[ray.get(date.remote(AP2[:, 0])) > ray.get(date.remote(ap1))]
            for ap0 in AP2:
                ap4 = np.append(ap4, [np.append(ap1, ap0)], axis=0)
        return ap4


@ray.remote
def LPTType(AiP3, ttype6, LP3, cn9, num=0):
    kc4 = KC[cn9]
    if num > 0:
        kc4 = 5 - num
    if kc4 == 4:
        return np.sum(AiP3[cn9, AEP] == 1) + len(np.intersect1d(np.flatnonzero(AiP3[cn9, VEP] == 1), ray.get(NZBF.remote(ttype6[-1]))))
    elif kc4 == 3:
        return np.sum(AiP3[cn9, EP] == 1)
    elif kc4 == 2:
        lp2 = 0
        AP1 = ray.get(APTType.remote(AiP3, ttype6, LP3, cn9, num=1))
        for ap1 in AP1:
            AP2 = ray.get(APTType.remote(ray.get(CAiP.remote(AiP3, ttype6, ap1, LP3, cn9)), ray.get(CTType.remote(ap1, ttype6)), LP3, cn9, num=1))
            AP2 = AP2[ray.get(date.remote(AP2[:, 0])) > ray.get(date.remote(ap1))]
            lp2 += len(AP2)
        return lp2
    elif kc4 == 1:
        lp3 = 0
        AP1 = ray.get(APTType.remote(AiP3, ttype6, LP3, cn9, num=2))
        for ap1 in AP1:
            AP2 = ray.get(APTType.remote(ray.get(CAiP.remote(AiP3, ttype6, ap1, LP3, cn9)), ray.get(CTType.remote(ap1, ttype6)), LP3, cn9, num=1))
            AP2 = AP2[ray.get(date.remote(AP2[:, 0])) != ray.get(date.remote(ap1))]
            lp3 += len(AP2)
        return lp3
    else:
        lp4 = 0
        AP1 = ray.get(APTType.remote(AiP3, ttype6, LP3, cn9, num=2))
        for ap1 in AP1:
            AP2 = ray.get(APTType.remote(ray.get(CAiP.remote(AiP3, ttype6, ap1, LP3, cn9)), ray.get(CTType.remote(ap1, ttype6)), LP3, cn9, num=2))
            AP2 = AP2[ray.get(date.remote(AP2[:, 0])) > ray.get(date.remote(ap1))]
            lp4 += len(AP2)
        return lp4


# 행렬-원소 선택 함수들#
@ray.remote
def PIL(arr1, ip):
    return np.array([Counter(arr1 <= i)[True] for i in ip]) - 1


# DPN 활용#
daily = np.max(DPN)
weekly = len(DPN)
IPD = np.array([0], int)
FPD = np.array([], int)
for i3 in range(weekly - 1):
    IPD = np.append(IPD, IPD[-1] + DPN[i3])
    FPD = np.append(FPD, IPD[-1] - 1)
wpn = np.sum(DPN)
FPD = np.append(FPD, wpn - 1)

@ray.remote
def date(ip):
    return np.unique(ray.get(PIL.remote(IPD, ip)))


@ray.remote
def CTType(selp1, ttype4):
    dates = ray.get(date.remote(selp1))
    kc4 = 6 - len(selp1) - len(dates)
    arr1 = np.copy(ttype4)
    for i in dates:
        if ttype4[i] < DMEP[i] and EP[ray.get(PIL.remote(EP, np.array([ttype4[i]])))] in selp1:
            arr1[i] += 2
    for i in np.intersect1d(dates,DCD):
        arr1[weekly + np.argwhere(DCD == i)] = 1
    if kc4 == 1:
        if selp1[2] in EP or selp1[2] in VEP:
            arr1[-1] = ray.get(BFN.remote(np.union1d(ray.get(NZBF.remote(ttype4[-1])), ray.get(PIL.remote(EP, np.array([selp1[2]]))))))
    elif kc4 == 2 or kc4 == 4:
        arr1[-1] = ray.get(BFN.remote(np.union1d(ray.get(NZBF.remote(ttype4[-1])), ray.get(PIL.remote(EP, np.setdiff1d(selp1,NEP))))))
    if kc4 == 0 or kc4 == 2:
        for i in SDPD:
            if np.array_equiv(dates, i):
                i0 = np.where(DCD == dates[0])[0] + weekly
                if ttype4[i0] == 0:
                    if selp1[0] + np.sum(DPN[dates[0]:dates[1]]) == selp1[2 - kc4 // 2]:
                        arr1[i0] = 0
    return arr1


# making classes#
ICL = np.array([0], int)
FCL = np.array([], int)
for i6 in range(l - 1):
    a1 = ICL[-1]
    a2 = LCN[i6]
    ICL = np.append(ICL, a1 + a2)
    FCL = np.append(FCL, a1 + a2 - 1)
FCL = np.append(FCL, c - 1)

StS = np.argsort(s * np.count_nonzero(SL, axis=1) + np.arange(s))
SL = SL[StS]
SC = np.zeros((s, c), int)
cnow = 0
for l1 in range(l):
    SinL = np.nonzero(SL[:, l1])[0]
    SinLC = np.array_split(SinL, LCN[l1])
    for i in SinLC:
        for j in i:
            SC[j, cnow] += 1
        cnow += 1
KC = np.repeat(KL, LCN)

ICK = np.array([], int)
for i in range(5):
    ICK = np.append(ICK, np.sum(KC < i))

RSC = np.array([np.copy(SC)], int)
REGC = np.empty((0, s, c), int)
REg = np.array([], int)
LMC = np.array([], int)
LLC = np.array([], int)
RE = np.array([], int)
SCbr = np.empty((s, c), int)

@ray.remote
def Allin(arr1, arr2):
    for i in arr1:
        if i not in arr2:
            return 0
    return 1

@ray.remote
def Pprediction(arr1, arr2, arr3):
    for i in range(len(arr1)):
        arr4 = arr1[i, :arr2[i], ]
        for j in range(arr2[i]):
            if ray.get(Allin.remote(arr4, arr3)) == 1:
                return 1
    return 0

@ray.remote
def sortition(arr):
    colsort = np.zeros(c, int)
    for l1 in range(l):
        ind = np.arange(LCN[l1]) + ICL[l1]
        np.put(colsort, ind, np.argsort(np.argmax(arr[:, ind], axis=0)) + ICL[l1])
    return arr[:, colsort]

@ray.remote
def swr(arr, indi, n):
    csort = np.arange(len(arr.T))
    np.put(csort, np.arange(indi, indi + n), csort[indi:indi + n][::-1])
    return arr[:, csort]

@ray.remote
def plusing(arr, xind, yind, num):
    arr2 = np.copy(arr)
    for i in xind:
        for j in yind:
            arr2[i, j] += num
    return arr2


# 소수,진법 변환 관련 함수들#
@ray.remote
def NZBF(n):
    nzbf = np.empty(shape=(0), dtype=np.int64)
    n1 = 0
    while n > 0:
        if n % 2 == 1:
            nzbf = np.append(nzbf, n1)
        n = n // 2
        n1 += 1
    return nzbf

@ray.remote
def BFN(arr):
    return 2**(len(EP) - arr - 1)


# 교시교환 관련 함수들#
def LW1P(da1, arr):
    a = IPD[da1]
    tw = np.zeros((2, 2), int)
    for k in range(4):
        i = k // 2
        j = k % 2
        for s1 in range(s):
            c1 = arr[s1][a + 1 - i]
            c2 = arr[s1][a + 2 + j]
            if max(c1, c2) < c:
                l1 = ray.get(PIL.remote(ICL, np.array([c1])))[0]
                l2 = ray.get(PIL.remote(ICL, np.array([c2])))[0]
                tw[i][j] += DI[l1][l2]
    p2 = np.argmin(tw)
    y = p2 % 2
    x = p2 // 2
    mwp = np.array([a + x, a + 1 - x, a + 2 + y, a + 3 - y], int)
    if DPN[da1] % 2 == 1:
        mwp = np.append(mwp, a + 4)
    return mwp


def LW2P(da2, arr):
    a = IPD[da2]
    tw = np.zeros((3, 2, 2), int)
    for t in range(12):
        i = t // 4
        j = (t // 2) % 2
        k = t % 2
        for s1 in range(s):
            c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
            c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
            if max(c1, c2) < c:
                l1 = ray.get(PIL.remote(ICL, np.array([c1])))[0]
                l2 = ray.get(PIL.remote(ICL, np.array([c2])))[0]
                tw[i][j] += DI[l1][l2]
    p3 = np.argmin(tw)
    z = p3 % 2
    y = (p3 // 2) % 2
    x = p3 // 4
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2, a + x2 + 1], int)


def LW3P(da3, arr):
    a = IPD[da3]
    tw = np.zeros((3, 2, 2, 2), int)
    for t in range(24):
        i = t // 8
        j = (t // 4) % 2
        k = (t // 2) % 2
        o = t % 2
        for s1 in range(s):
            c1 = arr[s1][a + 2 * (i % 3) + 1 - j]
            c2 = arr[s1][a + 2 * ((i + 1) % 3) + k]
            c3 = arr[s1][a + 2 * ((i + 2) % 3) + 1 - o]
            c4 = arr[s1][a + 6]
            if max(c1, c2) < c:
                l1 = ray.get(PIL.remote(ICL, np.array([c1])))[0]
                l2 = ray.get(PIL.remote(ICL, np.array([c2])))[0]
                tw[i][j] += DI[l1][l2]
            if max(c3, c4) < c:
                l1 = ray.get(PIL.remote(ICL, np.array([c3])))[0]
                l2 = ray.get(PIL.remote(ICL, np.array([c4])))[0]
                tw[i][j] += DI[l1][l2]
    p4 = np.argmin(tw)
    w = p4 % 2
    z = (p4 // 2) % 2
    y = (p4 // 4) % 2
    x = p4 // 8
    x0 = 2 * (x % 3)
    x1 = 2 * ((x + 1) % 3)
    x2 = 2 * ((x + 2) % 3)
    return np.array([a + x0 + y, a + x0 + 1 - y, a + x1 + z, a + x1 + 1 - z, a + x2 + w, a + x2 + 1 - w, a + 6], int)


def LWP(arr):
    tmwp = np.array([], int)
    for i in range(weekly):
        if DPN[i] < 6:
            tmwp = np.append(tmwp, LW1P(i, arr))
        elif DPN[i] == 6:
            tmwp = np.append(tmwp, LW2P(i, arr))
        else:
            tmwp = np.append(tmwp, LW3P(i, arr))
    return tmwp


# 그래프 색깔#
KCA = np.zeros((c,), int)
for c1 in range(c):
    l1 = PIL(ICL, np.array([c1]))[0]
    KCA[c1] = KA[l1]
SPC = np.zeros((12,), int)
for i in range(12):
    SPC[i] = np.sum(KCA == i)
GT = ['Mathematics', 'Physics', 'Chemistry', 'Computer Science', 'Life Science', 'Earth Science', 'Convergence Science',
      'Science Experiment', 'Creative Convergence Special Lecture', 'Language', 'Society', 'Music-Art-Sport']
CAL = []
for i in range(12):
    CAL.append(np.argwhere(KCA == 0).tolist())


def RGBset(n):
    k = -(math.floor(1 - (8 * n + 1) ** 0.5 / 2))
    RGBset1 = np.array([], int)
    if k > 1:
        x = 255 // (k - 1)
        for i in range(k):
            RGBset1 = np.append(RGBset1, i * x)
        return RGBset1


def CColorSet(mat):
    n = len(mat)
    nRGB = RGBset(n)
    k = len(nRGB) - 1
    RGB = []
    for R in range(k, -1, -1):
        if (k - R) % 2 == 0:
            for G in range(k - R + 1):
                B = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[B], 'x'))
        else:
            for G in range(k - R, -1, -1):
                B = k - R - G
                RGB.append(format(nRGB[R] * 16 ** 4 + nRGB[G] * 16 ** 2 + nRGB[B], 'x'))
    return RGB


PTTcolorset = ['ffe4e1', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff', 'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']

# TType 설정 코드#
DP = np.setdiff1d(np.arange(wpn), np.union1d(IPD, IPD + 4))
DEPN = DPN // 2
EP = np.array([], int)
for i in range(weekly):
    EP = np.append(EP, IPD[i])
    for j in range(DEPN[i] - 1):
        EP = np.append(EP, EP[-1] + 2)
VEP = EP + 1
AEP = np.setdiff1d(np.arange(wpn), VEP)
NEP = np.setdiff1d(AEP, EP)
NEPI = date(NEP)

DMEP = IPD + 2 * DEPN - 2
SDPI = np.array([], int)
SDPD = np.empty((0, 2), int)
DCD = np.array([], int)
CAD = np.array([], int)

fTType = np.copy(IPD)
for i in range(weekly - 1):
    j = i
    fi = 1
    while (weekly - j) * fi > 1:
        j += 1
        if DPN[i] == DPN[j]:
            SDPI = np.append(SDPI, len(fTType))
            fTType = np.append(fTType, 0)
            SDPD = np.append(SDPD, [[i, j]], axis=0)
            DCD = np.append(DCD, i)
            CAD = np.append(CAD, j)
            fi = 0
fTType = np.append(fTType, 0)

oAiP = fTType[np.setdiff1d(np.arange(weekly), CAD, assume_unique=True)]
fAiP = -np.ones((c, wpn), int)
fAiP[:, np.sort(np.concatenate((oAiP, oAiP + 1, NEP)))] *= -1

ADS = np.arange(weekly).reshape(1, -1)
nds = 1
done1 = 1
i1111 = 0
l1111 = 0
while done1 == 1:
    j = 0
    re = 1
    while re * (nds - i1111) > 0:
        ADS1 = ADS[i1111]
        while re * (len(DCD) - j) > 0:
            k0 = DCD[j]
            k1 = CAD[j]
            ADS2 = np.copy(ADS1)
            a1 = ADS1[k1]
            a2 = ADS1[k0]
            ADS2[k0] = a1
            ADS2[k1] = a2
            i1 = 0
            l1111 = len(ADS)
            while i1 < l1111:
                ads2 = ADS[i1]
                if np.array_equiv(ads2, ADS2):
                    j += 1
                    i1 = len(ADS)
                else:
                    i1 += 1
                    if i1 == len(ADS):
                        ADS = np.append(ADS, [ADS2], axis=0)
                        nds += 1
                        re = 0
        if re == 1:
            i1111 += 1
        elif i1111 < l1111:
            i1111 += 1
            j = 0
    if re == 1:
        done1 = 0


# 엑셀 조작 관련#
def style(workbook, n, i, j, arr0, arr1):
    newsheet = workbook[sheetname(n, arr1)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif arr0[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(arr0.reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == arr0[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[c4], fill_type='solid')


def graphstyle(workbook, a, arr1, arr2):
    newsheet = workbook[GT[a]]
    for c1 in arr1:
        arr3 = np.flatnonzero(np.sum(arr2 == c1, axis=0))
        for p1 in arr3:
            newsheet.cell(c1 + 2, p1 + 2).fill = PatternFill(start_color=CColorSet(arr1)[c1], end_color=CColorSet(arr1)[c1], fill_type='solid')


def sheetname(b, arr):
    return str(arr[np.where(arr == b)[0]]) + '번째 학생'


def makingsheetdone(arr, wbname, arr1, arr2):
    wb = openpyxl.Workbook()
    for i in range(len(arr)):
        wb.create_sheet(sheetname(i, arr1))
    for area in range(12):
        wb.create_sheet(GT[area])
    wb.create_sheet('student walking distance')
    list3 = []
    for s3 in range(len(arr)):
        list0 = arr[s3].tolist()
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
    list4 = list3[:]
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(arr)):
            for i in range(weekly):
                for j in range(daily):
                    c5 = list3[s0][i][j]
                    if c5 < c:
                        l1 = PIL(ICL, np.array([c5]))[0]
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LN[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0, arr1))
        for area0 in range(12):
            df = pd.DataFrame(index=list(map(str, CAL[area0])), columns=periodnames)
            df.to_excel(writer, sheet_name=GT[area0])
        dfc = pd.DataFrame(arr2, columns=StudentNames[arr1])
        dfc.to_excel(writer, sheet_name='student walking distance')
        chart = openpyxl.chart.LineChart()
        chart.title = 'student walking distance graph'
        chart.x_axis.title = 'student'
        chart.y_axis.title = 'walking distance'
        ws1 = wb['student walking distance']
        datas = openpyxl.chart.Reference(ws1, min_col=2, min_row=2, max_col=s + 1, max_row=2)
        chart.add_data(datas, from_rows=True, titles_from_data=False)
        cats = openpyxl.chart.Reference(ws1, min_col=2, min_row=1, max_col=s + 1, max_row=1)
        chart.set_categories(cats)
        ws1.add_chart(chart, "A1")
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(arr)):
        i = 0
        s2 = wb[sheetname(s1, arr1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list4[s1], arr1)
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    for a1 in range(12):
        graphstyle(wb, a1, CAL[a1], arr)
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


dead = 0
out = 0
distinguished = 0
old = 0
go = 0
back = 0
newface = 0
fastsend = 0

cN = -1
Found = 0
pN = 0
SbD = np.zeros((s, weekly), int)
RSbD = np.empty((0, s, weekly), int)
CP = np.array([], str)
AiP = np.copy(fAiP).reshape(1, c, wpn)
RAiP = np.empty((0, c, wpn), int)
maxcN = -1
RcN = np.array([], int)
gRcN = np.array([], int)
fig = plt.figure()
NcN = np.arange(c)

while dead == 0:
    GC = np.empty((0, c), int)
    GS = np.array([], int)
    CC = np.zeros((c, c), int)
    for s1 in range(s):
        SS1 = SC[s1]
        CC += np.outer(SS1, SS1, out=None)
    for s2 in range(s):
        NZ = np.flatnonzero(SC[s2])
        if np.min(CC[NZ][:, NZ]) > 1:
            SS2 = SC[s2]
            CC -= np.outer(SS2, SS2, out=None)
        else:
            GS = np.append(GS, s2)
            GC = np.append(GC, [SC[s2]], axis=0)
    g = len(GS)
    GC = ray.get(sortition.remote(GC))
    out = ray.get(Pprediction.remote(REGC, REg, GC))
    if out == 1:
        fastsend = 1
    else:
        distinguished = 0
        old = 0
        RE = np.array([], int)
        SuCP = np.zeros((1, c, wpn), int)
        maxcN = -1
        figinum = 0
        SuNcN = np.arange(c).reshape(1,-1)
        SuRcN = np.empty((1, 0), int)
        SuLP = np.zeros((1, l, wpn), int)
        SuGP = np.zeros((1, g, wpn), int)
        SuAiP = np.copy(fAiP).reshape(1, c, wpn)
        SuTType = np.copy(fTType).reshape(1, -1)
        SucN = np.array([], int)
        fLlcN = np.array([], int)
        GP = np.zeros((g, wpn), int)
        for i in range(c):
            fLlcN = np.append(fLlcN, ray.get(LPTType.remote(SuAiP[0], SuTType[0], SuLP[0], i)))
        SuLlcN = np.array([fLlcN]).reshape(1, -1)
        while out == 0:
            if len(SuRcN.T) > 0:
                out = 1
                dead = 1
            delpass = np.array([], int)
            for i in range(len(SuLlcN)):
                if 0 in SuLlcN[i]:
                    for j in NcN[SuLlcN[i] == 0]:
                        RE = np.unique(np.append(RE, s * np.flatnonzero(GC[:, cN] * GC[:, j]) + cN))
                    delpass = np.append(delpass, i)
            SuTType = np.delete(SuTType, delpass, axis=0)
            SuAiP = np.delete(SuAiP, delpass, axis=0)
            SuLP = np.delete(SuLP, delpass, axis=0)
            SuLlcN = np.delete(SuLlcN, delpass, axis=0)
            SuCP = np.delete(SuCP, delpass, axis=0)
            SuNcN = np.delete(SuNcN, delpass, axis=0)
            SuRcN = np.delete(SuRcN, delpass, axis=0)
            SucN = np.delete(SucN, delpass, axis=0)
            if len(SuLlcN) > 0:
                for i in range(len(SuLlcN)):
                    SuNcN[i] = SuNcN[i, np.argsort(SuLlcN[i])]
                    SuLlcN[i] = SuLlcN[i, np.argsort(SuLlcN[i])]
                SucN = SuNcN[:, 0]
                SuRcN = np.append(SuRcN, SucN.reshape(-1,1), axis=1)
                SuNcN = np.delete(SuNcN, 0, axis=1)
                Renum = np.copy(SuLlcN[:, 0])
                SuLlcN = np.delete(SuLlcN, 0, axis=1)
                APcN1d = np.array([], int)
                IAD = np.array([0], int)
                for i in range(len(SuTType)):
                    GP = SuGP[i]
                    IAD = np.append(IAD, np.arange(1, Renum[i]+1) * (5 - KC[SucN[i]] - (KC[SucN[i]] < 3)) + len(APcN1d))
                    APcN1d = np.append(APcN1d, ray.get(APTType.remote(SuAiP[i], SuTType[i], SuLP[i], SucN[i])).reshape(-1))
                SuAiP = np.repeat(SuAiP, Renum, axis=0)
                SuTType = np.repeat(SuTType, Renum, axis=0)
                SuLP = np.repeat(SuLP, Renum, axis=0)
                SuGP = np.repeat(SuGP, Renum, axis=0)
                SuLlcN = np.repeat(SuLlcN, Renum, axis=0)
                SuCP = np.repeat(SuCP, Renum, axis=0)
                SuNcN = np.repeat(SuNcN, Renum, axis=0)
                SuRcN = np.repeat(SuRcN, Renum, axis=0)
                SucN = np.repeat(SucN, Renum, axis=0)
                SulN = ray.get(PIL.remote(ICL, SucN))
                for i in tqdm(range(sum(Renum)), desc=str(len(SuRcN.T)), ascii=True):
                    AP = APcN1d[IAD[i]:IAD[i+1]]
                    SuCP[i, SucN[i],AP] += 1
                    SuLP[i, SulN[i],AP] += 1
                    SuGP[i, np.flatnonzero(GC[:, SucN[i]])][AP] += 1
                    GP = SuGP[i]
                    SuAiP[i] = np.copy(ray.get(CAiP.remote(SuAiP[i], SuTType[i], AP, SuLP[i], SucN[i])))
                    SuTType[i] = np.copy(ray.get(CTType.remote(AP, SuTType[i])))
                    for j in SuNcN[i]:
                        SuLlcN[i, np.flatnonzero(SuNcN[i] == j)[0]] = ray.get(LPTType.remote(SuAiP[i], SuTType[i], SuLP[i], j))
"""
