import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook, chart
from openpyxl.utils import get_column_letter
import math
import matplotlib.pyplot as plt

data1 = np.array([], int)
for i in range(100):
    data1 = np.append(data1, i**2)
    plt.close()
    plt.plot(np.arange(i+1),data1)
    plt.show()