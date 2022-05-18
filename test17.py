import numpy as np
import pandas as pd
import math
import openpyxl
from numba import jit, prange,njit
from numba import int64 as i8
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
@jit(nopython=True,cache=True)
def CGS1(i1, arr3):
    return np.nonzero(arr3[:, i1])[0]
arr1 = np.array([[1,1],[0,1],[1,1]])
start = time.time()
print(CGS1(1,arr1))
print(CGS1(1,arr1))
print(CGS1(1,arr1))
print(CGS1(1,arr1))
print(CGS1(1,arr1))
print(CGS1(1,arr1))
print(CGS1(1,arr1))
print(CGS1(0,arr1))
print("time1 :", time.time() - start)