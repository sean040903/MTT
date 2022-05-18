import numpy as np
import pandas as pd
import math
import openpyxl
from numba import jit, prange
from numba import int64 as i8
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
arr2 = np.array([[1,1],[0,1],[1,1]])
def CGS2(i4, arr4):
    return np.nonzero(arr4[:, i4])[0]



start2 = time.time()
print(CGS2(1,arr2))
print(CGS2(1,arr2))
print(CGS2(1,arr2))
print(CGS2(1,arr2))
print(CGS2(1,arr2))
print(CGS2(1,arr2))
print(CGS2(1,arr2))
print(CGS2(0,arr2))
print("time2 :", time.time() - start2)