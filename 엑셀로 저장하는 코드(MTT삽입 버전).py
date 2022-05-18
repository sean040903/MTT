import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

PTTcolorset = ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff',
               'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']

datenames = list(input())


def style(workbook, n, i, j, list5):  # workbook: wb, n:학생,nsn:새로운 시트,i:요일,j:교시, list2=list3[s1]
    newsheet = workbook[sheetname(n)]
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    elif list5[i][j] == c:
        newsheet.cell(i + 2, j + 2).value = ""
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        containedclasses = sorted(list(np.unique(np.array(list5).reshape(-1))))
        del containedclasses[-1]
        for c4 in range(len(containedclasses)):
            if containedclasses[c4] == list5[i][j]:
                newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[c4], end_color=PTTcolorset[0],
                                                               fill_type='solid')


def sheetname(b):
    return str[b] + '번째 학생'


def makingsheetdone(mat,wbname):
    wb = openpyxl.Workbook()
    for i in range(len(mat)):
        wb.create_sheet(sheetname(i))
    list3 = []
    for s3 in range(len(mat)):
        list0 = mat[s3]
        list2 = []
        for i in range(weekly):
            list1 = list0[IPD[i]:FPD[i]]
            for j in range(daily - DPN[i]):
                list1.append(c)
            list2.append(list1)
        list3.append(list2)
    with pd.ExcelWriter('SnappyData.xlsx') as writer:
        for s0 in range(len(mat)):
            for i in range(weekly):
                for j in range(daily):
                    c5=list3[s0][i][j]
                    if c5 < c:
                        l1=PIL(ICL,c5)
                        pc = (c5 - ICL[l1] + 1)
                        list3[s0] = LectureNames[l1] + '(' + str(pc) + '반)'
            df = pd.DataFrame(list3[s0], columns=datenames)
            df.to_excel(writer, sheet_name=sheetname(s0))
    wb = load_workbook(filename='SnappyData.xlsx')
    for s1 in range(len(mat)):
        i = 0
        s2 = wb[sheetname(s1)]
        while i < len(df.index):
            j = -1
            while j < len(df.columns):
                style(wb, s1, i, j, list3[s1])
                j += 1
            i += 1
        for row in range(1, len(df.index) + 1):
            s2.row_dimensions[row + 1].height = 40
        for col in range(1, len(df.columns) + 1):
            s2.column_dimensions[get_column_letter(col + 1)].width = 15
        s2.row_dimensions[1].height = 18
        s2.column_dimensions[get_column_letter(1)].width = 5
    os.remove('SnappyData.xlsx')
    wb.save(wbname)


for i in range(len(sortFScTs)):
    workbookname = str(i+1) + '번째 시간표'
    ScT10 = np.array(sortFScTs[i]).reshape(s,wpn).tolist()
    makingsheetdone(ScT10,workbookname)
