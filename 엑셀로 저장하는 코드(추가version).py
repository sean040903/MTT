import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

s = 6

PTTcolorset = ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff',
               'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']


def style(n, i, j):  # n:학생,nsn:새로운 시트,i:요일,j:교시
    newsheet = wb[sheetname(n)]
    arr1 = np.array(datalist[n])
    arr2 = arr1.T
    datalist2 = arr2.tolist()
    df = pd.DataFrame(datalist2, columns=['월요일', '화요일', '수요일', '목요일'])
    list1 = df.values.tolist()
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        newsheet.cell(i + 2, 1).font = font_style2
        newsheet.cell(i + 2, 1).alignment = alignment_style2
    else:
        newsheet.cell(i + 2, j + 2).font = font_style
        newsheet.cell(i + 2, j + 2).alignment = alignment_style
        newsheet.cell(i + 2, j + 2).border = border_style
        if list1[i][j] == '고급물리학2':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[0], end_color=PTTcolorset[0],
                                                           fill_type='solid')
        elif list1[i][j] == '수학4':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[1], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '수학3':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[2], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '고급물리학1':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[3], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '화학4':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[4], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '물리2':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[5], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '화학3':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[6], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '화학1':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[7], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '수학2':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[8], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '물리1':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[9], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '화학2':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[10], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '수학1':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[11], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '로봇공학':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[12], end_color=PTTcolorset[1],
                                                           fill_type='solid')
        elif list1[i][j] == '국어1':
            newsheet.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[13], end_color=PTTcolorset[1],
                                                           fill_type='solid')


def sheetname(s2):
    return list(map(str, [s2]))[0]


datalist = [[['수학4', '고급물리학2', '고급물리학1', '화학4'], ['화학3', '화학1', '수학3', '물리2'], ['수학2', '물리1', '화학2', '수학1'],
             ['로봇공학', '로봇공학', '국어1', '국어1']],
            [[ '국어1', '국어1', '화학1', '화학4'], ['수학3', '물리2', '고급물리학2', '고급물리학1'], ['수학2', '물리1', '화학2', '수학1'],
             ['로봇공학', '로봇공학','수학4', '화학3']],
            [['수학4', '고급물리학2', '고급물리학1', '화학4'], ['수학3', '화학1', '물리2', '화학3'], ['수학2', '로봇공학', '국어1', '국어1'],
             ['로봇공학', '물리1', '화학2', '수학1']],
            [['수학4', '고급물리학2',  '물리1', '화학1'], ['화학4', '수학3', '화학3', '화학2'], ['수학2','고급물리학1', '물리2','수학1'],
             ['로봇공학', '로봇공학', '국어1', '국어1']],
            [['화학4', '로봇공학', '국어1', '물리1'], ['고급물리학1', '수학4', '화학2', '수학1'], ['수학2', '수학3', '물리2', '화학3'],
             ['로봇공학','고급물리학2', '화학1', '국어1']],
            [['수학4', '고급물리학2', '고급물리학1', '국어1'], ['수학3', '물리2', '화학2', '화학1'], ['수학2', '물리1', '화학3', '수학1'],
             ['로봇공학', '로봇공학', '국어1', '화학4']]]

wb = openpyxl.Workbook()
for i in range(s):
    wb.create_sheet(sheetname(i))

with pd.ExcelWriter('test.xlsx') as writer:
    for s0 in range(s):
        arr1 = np.array(datalist[s0])
        arr2 = arr1.T
        datalist2 = arr2.tolist()
        df = pd.DataFrame(datalist2, columns=['월요일', '화요일', '수요일', '목요일'])
        df.to_excel(writer, sheet_name=sheetname(s0))

excel_filename = 'test.xlsx'
wb = load_workbook(filename=excel_filename)

for s0 in range(s):
    i = 0
    s1 = wb[sheetname(s0)]
    while i < len(df.index):
        j = -1
        while j < len(df.columns):
            style(s0, i, j)
            j += 1
        i += 1
    for row in range(1, len(df.index) + 1):
        s1.row_dimensions[row + 1].height = 40
    for col in range(1, len(df.columns) + 1):
        s1.column_dimensions[get_column_letter(col + 1)].width = 15
    s1.row_dimensions[1].height = 18
    s1.column_dimensions[get_column_letter(1)].width = 5
os.remove('test.xlsx')
wb.save('파이썬으로 만든 시간표(최종버전).xlsx')
