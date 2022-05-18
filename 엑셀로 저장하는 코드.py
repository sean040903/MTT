import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def style(i, j):  # n:학생,i:요일,j:교시
    font_style = Font(name='맑은 고딕', bold=True, color='000000', size=15)
    alignment_style = Alignment(vertical='center', horizontal='center', wrap_text=False)
    border_style = Border(left=Side(style="thick", color='000000'), right=Side(style="thick", color='000000'),
                          top=Side(style="thick", color='000000'), bottom=Side(style="thick", color='000000'))
    font_style2 = Font(name='맑은 고딕', bold=False, color='000000', size=10)
    alignment_style2 = Alignment(vertical='center', horizontal='center', wrap_text=True)
    if j == -1:
        sheet1.cell(i + 2, 1).font = font_style2
        sheet1.cell(i + 2, 1).alignment = alignment_style2
    else:
        sheet1.cell(i + 2, j + 2).font = font_style
        sheet1.cell(i + 2, j + 2).alignment = alignment_style
        sheet1.cell(i + 2, j + 2).border = border_style
        if list1[i][j] == '고급물리학2':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[0], end_color=PTTcolorset[0],
                                                         fill_type='solid')
        elif list1[i][j] == '수학4':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[1], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '수학3':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[2], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '고급물리학1':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[3], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '화학4':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[4], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '물리2':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[5], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '화학3':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[6], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '화학1':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[7], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '수학2':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[8], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '물리1':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[9], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '화학2':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[10], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '수학1':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[11], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '로봇공학':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[12], end_color=PTTcolorset[1],
                                                         fill_type='solid')
        elif list1[i][j] == '국어1':
            sheet1.cell(i + 2, j + 2).fill = PatternFill(start_color=PTTcolorset[13], end_color=PTTcolorset[1],
                                                         fill_type='solid')


datalist = ['1101 김민재' '1201 권우재' '1301 김남혁' '1401 김도현' '1501 고민재' '1601 공윤수'
 '1701 고종환' '1801 구준우' '2101 김규민' '2102 김선웅' '2103 김시호' '2104 김영준'
 '2105 김영훈' '2106 김예성' '2107 김재성' '2108 노이헌' '2109 박준상' '2110 서규민'
 '2111 윤원준' '2112 이주협' '2113 이형석' '2114 임설호' '2115 조영인' '2201 김산'
 '2202 김진모' '2203 남승우' '2204 류재민' '2205 박성빈' '2206 박해성' '2207 오유찬'
 '2208 유지훈' '2209 윤동섭' '2210 이다연' '2211 이승환' '2212 이화인' '2213 임재현'
 '2214 정유찬' '2215 최우진' '2216 최홍우' '2301 권성안' '2302 권용하' '2303 김도균'
 '2304 김범준' '2305 박주형' '2306 손훈일' '2307 양준혁' '2308 이규동' '2309 장민기'
 '2310 전승주' '2311 정다온' '2312 정우찬' '2313 조민상' '2314 조연우' '2315 최민종'
 '2316 허재학' '2401 김동원' '2402 김성현' '2403 김온겸' '2404 김유현' '2405 김현준'
 '2406 박준원' '2407 백종현' '2408 엄휘식' '2409 이서준' '2410 이원재' '2411 이준서'
 '2412 이지후' '2413 임지언' '2414 정채원' '2415 조명기' '2416 한승우' '2501 김진하'
 '2502 박승비' '2503 박준형' '2504 박진원' '2505 박현민' '2506 서채원' '2507 윤성현'
 '2508 이민혁' '2509 이승원' '2510 이현채' '2511 장희찬' '2512 전지민' '2513 정해찬'
 '2514 조윤재' '2515 주성운' '2516 차이경' '2601 권지이' '2602 김수빈' '2603 김희재'
 '2604 박준혁' '2605 반지호' '2606 배준휘' '2607 서세준' '2608 신유범' '2609 이동훈'
 '2610 이상기' '2611 임세준' '2612 정유찬' '2613 정현우' '2614 조준우' '2615 한은혜'
 '2616 허정' '2701 김선우' '2702 문새결' '2703 박세진' '2704 박시형' '2705 박재영'
 '2706 배현서' '2707 이범석' '2708 이승현' '2709 이예준' '2710 이원재' '2711 이윤수'
 '2712 이현진' '2713 이형주' '2714 장준성' '2715 정재원' '2716 정준호' '2801 고민준'
 '2802 고윤석' '2803 김주영' '2804 류태상' '2805 박도영' '2806 박탐' '2807 방유찬'
 '2808 변지예' '2809 심이센' '2810 이관우' '2811 이정진' '2812 이지후' '2813 최재훈'
 '2814 최정윤' '2815 한승헌' '3101 김남진' '3102 김대순' '3103 김용재' '3104 김지환'
 '3105 김해정' '3106 류호원' '3107 문성빈' '3108 문준원' '3109 송민혁' '3110 이재효'
 '3111 이호준' '3112 임서윤' '3113 장윤성' '3114 최민준' '3115 최원형' '3116 한정우'
 '3201 강동우' '3202 고건' '3203 김선호' '3204 김영서' '3205 김재영' '3206 송경민'
 '3207 송재한' '3208 여현준' '3209 오태인' '3210 이재용' '3211 이주원' '3212 이지후'
 '3213 정찬우' '3214 조원준' '3215 홍재영' '3216 황태훈' '3301 구현모' '3302 김동현'
 '3303 김석환' '3304 김세헌' '3305 김수민' '3306 김영준' '3307 김재우' '3308 박진수'
 '3309 배서연' '3310 이성민' '3311 이윤서' '3312 이주한' '3313 이준서' '3314 이준우'
 '3315 정우진' '3316 홍휘택' '3401 고도영' '3402 권보민' '3403 김동규' '3404 김종우'
 '3405 민규철' '3406 송우석' '3407 안호중' '3408 양준혁' '3409 유현우' '3410 육민수'
 '3411 이동건' '3412 장현서' '3413 정진교' '3414 조성호' '3415 최성민' '3416 한서준'
 '3417 황수영' '3501 김민준' '3502 김현준' '3503 남우현' '3504 박지연' '3505 박지형'
 '3506 서한결' '3507 신명진' '3508 유연주' '3509 이동재' '3510 이서준' '3511 이준혁'
 '3512 임상우' '3513 임재아' '3514 정재원' '3515 조성길' '3516 황인환' '3601 권재욱'
 '3602 김동현' '3603 김성헌' '3604 나윤호' '3605 노세현' '3606 문수혁' '3607 박지호'
 '3608 변형준' '3609 서지민' '3610 안도현' '3611 안현태' '3612 이채운' '3613 정선'
 '3614 정정훈' '3615 차원재' '3616 함주현' '3701 강현후' '3702 고영린' '3703 김성혁'
 '3704 김찬우' '3705 민정호' '3706 박경원' '3707 박제하' '3708 송현욱' '3709 심준선'
 '3710 이경민' '3711 이현서' '3712 정시우' '3713 조석희' '3714 한상우' '3715 홍경찬'
 '3716 황아현' '3801 길아성' '3802 김동환' '3803 김민준' '3804 김성준' '3805 문승현'
 '3806 박정우' '3807 박찬우' '3808 손지훈' '3809 손형우' '3810 신다윤' '3811 유지민'
 '3812 윤동규' '3813 이준혁' '3814 이한주' '3815 최건' '3816 최원서']
df = pd.DataFrame(datalist1, columns=list(map(str,np.arange(len(datalist1)))))
with pd.ExcelWriter("test.xlsx") as writer:
    df.to_excel(writer, sheet_name="Test1")
excel_filename = 'test.xlsx'
wb = load_workbook(filename=excel_filename)
list1 = df.values.tolist()
sheet1 = wb.active
sheet2 = wb.create_sheet("2nd sheet")
PTTcolorset = ['ffe4e1', 'f7bbbb', 'ff7f50', 'ffb3d9', 'ffff00', '7fff00', 'd2b5fc', 'b4d0fd', 'afeeee', '00ffff',
               'b8ffe4', 'd8bfd8', 'ffca99', 'ff69b4']
i = 0
while i < len(df.index):
    j = -1
    while j < len(df.columns):
        style(i, j)
        j += 1
    i += 1

os.remove('test.xlsx')
for row in range(1, len(df.index) + 1):
    sheet1.row_dimensions[row + 1].height = 40
for col in range(1, len(df.columns) + 1):
    sheet1.column_dimensions[get_column_letter(col + 1)].width = 15
sheet1.row_dimensions[1].height = 18
sheet1.column_dimensions[get_column_letter(1)].width = 5

wb.save('python으로 만든 test 시간표(색,테두리,폭,글자크기 다 설정함).xlsx')
